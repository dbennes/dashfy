from __future__ import annotations

import hashlib
import re
import unicodedata
from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import Any

from django.core.cache import cache
from django.db import transaction

from .models import EngineeringMonitorImport


MONITOR_REQUIRED_HEADERS = {
    "Document Number",
    "Title",
    "Discipline",
    "Revision",
    "Directory",
    "Document Status",
    "Purpose Next Issue",
    "11-Cpy_Approval_Code",
}

MONITOR_OFFICIAL_HEADERS = {
    "Documento",
    "Contabilizar",
    "Status normalizado",
}

MONITOR_SAMPLE_ROOTS = (
    "BNO / 02-DED",
    "BNO / 02.5-FOE",
)

MONITOR_SAMPLE_EXCLUDED_FOLDERS = (
    "BNO / 02-DED / 02.000-PMS",
    "BNO / 02-DED / 02.001-DAE_TQR",
    "BNO / 02-DED / 02.003-REPORTS",
    "BNO / 02-DED / 02.004-OTHER",
    "BNO / 02-DED / 02.007-MDR_CHANGE NOTICE",
    "BNO / 02-DED / 02.300-DAE_PROCUREMENT / 02.301-DAE_TBE",
    "BNO / 02.5-FOE / FOE.00-TEMPLATES",
)

MONITOR_CANCELLED_DOCUMENT_STATUSES = frozenset({
    "CANCELED",
    "CANCELLED",
    "CANCELADO",
    "CANCELADA",
})

MONITOR_EXCLUDED_PURPOSE_NEXT_ISSUES = frozenset({
    "FORECAST",
    "FOE - RESUBMIT",
    "IFI - ISSUED FOR INFORMATION",
})

MONITOR_AFC_PURPOSE = "AFC/AFU - RELEASE"
MONITOR_IFF_PURPOSE = "IFF - ISSUED FOR FABRICATION"
MONITOR_IFA_PURPOSE = "IFA - ISSUED FOR APPROVAL"
MONITOR_IFR_PURPOSE = "IFR - ISSUED FOR REVIEW"
MONITOR_IFF_APPROVAL_CODES = ("CODE 1", "CODE 2", "CODE 3")
_MONITOR_IFF_APPROVAL_CODE_PATTERN = re.compile(r"\bCODE\s*[123]\b")

MONITOR_DISCIPLINE_GROUPS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("PMT", ("AA - PROJECT MANAGEMENT",)),
    ("CONSTRUCTION", ("BA - CONSTRUCTION",)),
    ("STRUCTURAL", ("CS - STRUCTURAL",)),
    ("ELECTRICAL", ("EA - ELECTRICAL",)),
    ("TECHNICAL SAFETY", ("HX - HSE", "HS - SAFETY")),
    ("INSTRUMENTATION", ("IN - INSTRUMENTATION",)),
    ("IT / IM", ("JA - IM", "KA - IT")),
    ("PIPING", ("MP - PIPING",)),
    ("MECHANICAL", (
        "MX - OVERAL MECHANICAL",
        "MR - MECHANICAL ROTATING",
        "MS - MECHANICAL STATIC",
        "MH - HVAC",
    )),
    ("PROCESS", ("PX - PROCESS",)),
    ("MATERIALS", ("RA - MATERIALS",)),
)

MONITOR_STATUS_ORDER = (
    "NI",
    "IFR",
    "IFA",
    "IFI",
    "AFC 1",
    "AFC 3",
    "AFC CODE 3A",
    "UNDER REVIEW",
    "N/A",
    "REJECTED",
    "UNCLASSIFIED",
)

_DISCIPLINE_LOOKUP = {
    raw: label
    for label, values in MONITOR_DISCIPLINE_GROUPS
    for raw in values
}
_DISCIPLINE_LABELS = {label for label, _values in MONITOR_DISCIPLINE_GROUPS} | {"PIPING ISOMETRIC"}


def _title_has_isometric_drawing(value: Any) -> bool:
    text = re.sub(r"[^A-Z0-9]+", " ", _norm_text(value))
    return "ISOMETRIC DRAWING" in text


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").strip().split())


def _norm_text(value: Any) -> str:
    text = unicodedata.normalize("NFD", _clean_text(value))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text.upper()).strip()


def _directory_segments(value: Any) -> tuple[str, ...]:
    text = unicodedata.normalize("NFKC", _clean_text(value)).replace("\\", "/")
    return tuple(
        normalized
        for part in text.split("/")
        if (normalized := _norm_text(part))
    )


_MONITOR_SAMPLE_ROOT_SEGMENTS = tuple(_directory_segments(path) for path in MONITOR_SAMPLE_ROOTS)
_MONITOR_SAMPLE_EXCLUDED_SEGMENTS = tuple(
    _directory_segments(path)
    for path in MONITOR_SAMPLE_EXCLUDED_FOLDERS
)


def _is_path_at_or_below(path: tuple[str, ...], root: tuple[str, ...]) -> bool:
    return len(path) >= len(root) and path[:len(root)] == root


def _monitor_sample_gate(
    directory: Any,
    document_status: Any,
    purpose_next_issue: Any = "",
    revision: Any = "",
) -> tuple[bool, str]:
    """Apply the configured global filter layers before status classification."""
    path = _directory_segments(directory)
    if not any(_is_path_at_or_below(path, root) for root in _MONITOR_SAMPLE_ROOT_SEGMENTS):
        return False, "Outside configured sample folders"
    if any(_is_path_at_or_below(path, folder) for folder in _MONITOR_SAMPLE_EXCLUDED_SEGMENTS):
        return False, "Excluded sample folder"
    if _norm_text(document_status) in MONITOR_CANCELLED_DOCUMENT_STATUSES:
        return False, "Cancelled document status"
    if _norm_text(purpose_next_issue) in MONITOR_EXCLUDED_PURPOSE_NEXT_ISSUES:
        return False, "Excluded purpose next issue"
    if "." in _clean_text(revision):
        return False, "Revision contains field mark"
    return True, ""


def _worksheet_header_row(ws: Any, required_headers: set[str]) -> int | None:
    normalized_required = {_norm_text(header) for header in required_headers}
    for row_number, values in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5), values_only=True),
        start=1,
    ):
        headers = {_norm_text(value).replace("\n", " ") for value in values if _clean_text(value)}
        if normalized_required.issubset(headers):
            return row_number
    return None


def _worksheet_has_headers(ws: Any, required_headers: set[str]) -> bool:
    return _worksheet_header_row(ws, required_headers) is not None


def _find_sheet(workbook: Any, required_headers: set[str]) -> Any | None:
    for ws in workbook.worksheets:
        if _worksheet_has_headers(ws, required_headers):
            return ws
    return None


def _revision_parts(value: Any) -> tuple[str, int]:
    text = _clean_text(value).upper()
    if not text or text == "-":
        return "", 0
    match = re.match(r"([A-Z]+)\s*0*(\d+)?", text)
    if not match:
        return text[:1], 0
    return match.group(1)[:1], int(match.group(2) or 0)


def normalize_monitor_discipline(raw_discipline: Any, title: Any = "") -> str:
    text = _norm_text(raw_discipline)
    discipline = _DISCIPLINE_LOOKUP.get(text, "")
    if not discipline and text in _DISCIPLINE_LABELS:
        discipline = text
    if discipline == "PIPING" and _title_has_isometric_drawing(title):
        return "PIPING ISOMETRIC"
    return discipline


def monitor_discipline_order(disciplines: Any = ()) -> list[str]:
    discipline_set = {str(value or "").strip().upper() for value in disciplines if str(value or "").strip()}
    order = [label for label, _values in MONITOR_DISCIPLINE_GROUPS]
    if "PIPING ISOMETRIC" in discipline_set and "PIPING ISOMETRIC" not in order:
        piping_index = order.index("PIPING") + 1 if "PIPING" in order else len(order)
        order.insert(piping_index, "PIPING ISOMETRIC")
    for discipline in sorted(discipline_set):
        if discipline not in order:
            order.append(discipline)
    return order


def _monitor_discipline(raw_discipline: Any, title: Any = "") -> str:
    return normalize_monitor_discipline(raw_discipline, title)


def _status_payload(
    document_status: Any,
    issue_status: Any,
    last_transmittal_purpose: Any,
    fabrication_ref: Any,
    directory: Any,
    document_number: Any,
    revision: Any,
    title: Any,
    purpose_next_issue: Any = "",
    approval_code: Any = "",
) -> dict[str, Any]:
    original_doc_status = _clean_text(document_status).upper()
    normalized_doc_status = _norm_text(document_status)
    is_transmittal = normalized_doc_status == "TRANSMITTAL"
    effective_doc_status = "MABU UNDER REVIEW" if is_transmittal else original_doc_status
    next_issue = _norm_text(purpose_next_issue)
    approval = _norm_text(approval_code)
    normalized_document_number = _norm_text(document_number)
    is_in_sample, sample_excluded_reason = _monitor_sample_gate(
        directory,
        document_status,
        purpose_next_issue,
        revision,
    )

    def bucket(label: str, group: str, afc_code: str = "") -> dict[str, Any]:
        return {
            "status_bucket": label,
            "doc_status": "MABU UNDER REVIEW" if label == "UNDER REVIEW" else label,
            "doc_status_group": group,
            "afc_code": afc_code,
            "document_status_effective": effective_doc_status,
            "document_status_original": original_doc_status,
            "is_transmittal": is_transmittal,
            "is_in_sample": True,
            "is_countable": True,
            "excluded_reason": "",
        }

    def excluded(reason: str) -> dict[str, Any]:
        return {
            "status_bucket": "",
            "doc_status": "",
            "doc_status_group": "",
            "afc_code": "",
            "document_status_effective": effective_doc_status,
            "document_status_original": original_doc_status,
            "is_transmittal": is_transmittal,
            "is_in_sample": False,
            "is_countable": False,
            "excluded_reason": reason,
        }

    if not is_in_sample:
        return excluded(sample_excluded_reason)

    # The engineering team's approved population treats AFC/AFU and blank
    # purposes as approved. IFF is approved only after a Code 1/2/3 return.
    if next_issue == MONITOR_AFC_PURPOSE or not next_issue:
        return bucket("AFC 1", "AFC", "1")
    if next_issue == MONITOR_IFF_PURPOSE:
        if _MONITOR_IFF_APPROVAL_CODE_PATTERN.search(approval):
            return bucket("AFC 1", "AFC", "1")
        return excluded("IFF without approved return code")

    if next_issue == MONITOR_IFA_PURPOSE:
        if "RA-7769" in normalized_document_number:
            return excluded("IFA RA-7769 excluded")
        return bucket("IFA", "IFA")

    if next_issue == MONITOR_IFR_PURPOSE:
        if "3323" in normalized_document_number:
            return excluded("IFR 3323 excluded")
        return bucket("IFR", "IFR")

    return excluded("Purpose next issue outside monitored statuses")


def _row_value(raw: dict[str, Any], header: str) -> Any:
    return raw.get(header) or raw.get(_norm_text(header)) or ""


def _first_row_value(raw: dict[str, Any], *headers: str) -> Any:
    for header in headers:
        value = _row_value(raw, header)
        if _clean_text(value):
            return value
    return ""


def _yes(value: Any) -> bool:
    text = _norm_text(value)
    return text in {"SIM", "S", "YES", "Y", "TRUE", "1", "X"}


def _normal_status_bucket(value: Any) -> str:
    text = _norm_text(value)
    text = text.replace("AFC - ", "AFC ")
    text = text.replace("CODE ", "CODE")
    if not text or text in {"NAO", "NO", "DESCONSIDERADO", "EXCLUIDO"}:
        return ""
    if text in {"NI", "NOT ISSUED"}:
        return "NI"
    if text.startswith("IFR"):
        return "IFR"
    if text.startswith("IFA"):
        return "IFA"
    if text in {"IFI", "ISSUED FOR INFORMATION"} or "ISSUED FOR INFORMATION" in text:
        return "IFI"
    if text in {"AFC 1", "AFC CODE1", "AFC CODE 1", "APPROVED FOR USE"}:
        return "AFC 1"
    if text in {"AFC 3", "AFC CODE3", "AFC CODE 3", "FOE"}:
        return "AFC 3"
    if text in {"AFC 3A", "AFC CODE3A", "AFC CODE 3A", "AFC 3 A", "AFC CODE 3 A"}:
        return "AFC CODE 3A"
    if "MABU" in text or "UNDER REVIEW" in text:
        return "UNDER REVIEW"
    if text in {"N/A", "NA", "NOT APPLICABLE", "NAO SE APLICA", "NAO APLICAVEL"}:
        return "N/A"
    if text.startswith("REJECT"):
        return "REJECTED"
    if text in {"UNCLASSIFIED", "NAO CLASSIFICADO"}:
        return "UNCLASSIFIED"
    return ""


def _status_parts_from_bucket(status_bucket: str) -> tuple[str, str, str]:
    status = _normal_status_bucket(status_bucket)
    if status == "NI":
        return "NI", "NOT ISSUED", ""
    if status in {"IFR", "IFA", "IFI"}:
        return status, status, ""
    if status == "AFC 1":
        return "AFC 1", "AFC", "1"
    if status == "AFC 3":
        return "AFC 3", "AFC", "3"
    if status == "AFC CODE 3A":
        return "AFC CODE 3A", "AFC", "3A"
    if status == "UNDER REVIEW":
        return "MABU UNDER REVIEW", "AFC", "UNDER REVIEW"
    if status == "N/A":
        return "N/A", "N/A", ""
    if status == "REJECTED":
        return "REJECTED", "REJECTED", ""
    if status == "UNCLASSIFIED":
        return "UNCLASSIFIED", "UNCLASSIFIED", ""
    return "", "", ""


def _parse_monitor_rows(ws: Any) -> list[dict[str, Any]]:
    header_row = _worksheet_header_row(ws, MONITOR_REQUIRED_HEADERS) or 1
    iterator = ws.iter_rows(min_row=header_row, values_only=True)
    raw_headers = next(iterator)
    headers = [_clean_text(value).replace("\n", " ") for value in raw_headers]
    normalized_headers = [_norm_text(header) for header in headers]
    rows: list[dict[str, Any]] = []

    try:
        for row_number, values in enumerate(iterator, start=header_row + 1):
            raw = dict(zip(headers, values))
            raw.update(dict(zip(normalized_headers, values)))
            document_number = _clean_text(_row_value(raw, "Document Number"))
            if not document_number:
                continue

            title = _clean_text(_row_value(raw, "Title"))
            revision = _clean_text(_row_value(raw, "Revision")).upper()
            revision_family, revision_number = _revision_parts(revision)
            raw_discipline = _clean_text(_row_value(raw, "Discipline")).upper()
            monitor_discipline = _monitor_discipline(raw_discipline, title)
            purpose_next_issue = _clean_text(_row_value(raw, "Purpose Next Issue")).upper()
            approval_code = _clean_text(_row_value(raw, "11-Cpy_Approval_Code")).upper()
            fabrication_ref = _clean_text(_first_row_value(
                raw,
                "14-Tr_Aol-Fabrication",
                "01-Tr_Aol-Fabrication",
            ))
            status = _status_payload(
                _row_value(raw, "Document Status"),
                _row_value(raw, "Issue Status"),
                _row_value(raw, "Last transmittal purpose"),
                fabrication_ref,
                _row_value(raw, "Directory"),
                document_number,
                revision,
                _row_value(raw, "Title"),
                purpose_next_issue,
                approval_code,
            )
            is_monitored = bool(status["is_in_sample"])
            rows.append({
                "row_number": row_number,
                "document_number": document_number,
                "title": title,
                "directory": _clean_text(_row_value(raw, "Directory")),
                "discipline": monitor_discipline or raw_discipline or "-",
                "source_discipline": raw_discipline or "-",
                "is_monitored": is_monitored,
                "revision": revision,
                "revision_family": revision_family,
                "revision_number": revision_number,
                "issue_status": _clean_text(_row_value(raw, "Issue Status")).upper(),
                "last_transmittal_purpose": _clean_text(_row_value(raw, "Last transmittal purpose")).upper(),
                "purpose_next_issue": purpose_next_issue,
                "approval_code": approval_code,
                "fabrication_ref": fabrication_ref,
                **status,
            })
    except ValueError:
        if not rows or len(rows) + 1 < int(getattr(ws, "max_row", 0) or 0):
            raise
    return rows


def _parse_official_rows(ws: Any) -> list[dict[str, Any]]:
    header_row = _worksheet_header_row(ws, MONITOR_OFFICIAL_HEADERS) or 1
    iterator = ws.iter_rows(min_row=header_row, values_only=True)
    raw_headers = next(iterator)
    headers = [_clean_text(value).replace("\n", " ") for value in raw_headers]
    normalized_headers = [_norm_text(header) for header in headers]
    rows: list[dict[str, Any]] = []

    for row_number, values in enumerate(iterator, start=header_row + 1):
        raw = dict(zip(headers, values))
        raw.update(dict(zip(normalized_headers, values)))
        document_number = _clean_text(_first_row_value(raw, "Documento", "Document Number"))
        if not document_number:
            continue

        title = _clean_text(_first_row_value(raw, "Titulo", "Título", "Title"))
        revision = _clean_text(_first_row_value(raw, "Revisao", "Revisão", "Revision", "rev")).upper()
        revision_family, revision_number = _revision_parts(revision)
        source_discipline = _clean_text(_first_row_value(raw, "Disciplina origem", "Source Discipline", "Discipline")).upper()
        raw_discipline = _clean_text(_first_row_value(raw, "Disciplina", "Disciplina AOL", "Discipline")).upper()
        discipline = (
            normalize_monitor_discipline(source_discipline or raw_discipline, title)
            or normalize_monitor_discipline(raw_discipline, title)
            or raw_discipline
        )
        status_bucket = _normal_status_bucket(_first_row_value(raw, "Status normalizado", "Status AOL", "Status", "Final Status"))
        is_countable = _yes(_first_row_value(raw, "Contabilizar", "Count", "Considerar"))
        doc_status, doc_status_group, afc_code = _status_parts_from_bucket(status_bucket)
        document_status_original = _clean_text(_first_row_value(raw, "Status documento", "Document Status")).upper()
        purpose_next_issue = _clean_text(_first_row_value(raw, "Purpose Next Issue", "Proxima emissao", "Próxima emissão")).upper()
        approval_code = _clean_text(_first_row_value(raw, "11-Cpy_Approval_Code", "Approval Code")).upper()
        is_transmittal = _norm_text(document_status_original) == "TRANSMITTAL"
        document_status_effective = "MABU UNDER REVIEW" if is_transmittal else document_status_original
        excluded_reason = _clean_text(_first_row_value(raw, "Motivo desconsiderado", "Motivo", "Excluded Reason"))
        if not is_countable and not excluded_reason:
            excluded_reason = "Desconsiderado pelo administrador"

        rows.append({
            "row_number": row_number,
            "document_number": document_number,
            "title": title,
            "directory": _clean_text(_first_row_value(raw, "Diretorio", "Diretório", "Directory")),
            "discipline": discipline or source_discipline or "-",
            "source_discipline": source_discipline or discipline or "-",
            "is_monitored": is_countable,
            "revision": revision,
            "revision_family": revision_family,
            "revision_number": revision_number,
            "issue_status": _clean_text(_first_row_value(raw, "Issue Status", "Status emissao", "Status emissão")).upper(),
            "last_transmittal_purpose": _clean_text(_first_row_value(raw, "Last transmittal purpose", "Proposito transmittal", "Propósito transmittal")).upper(),
            "purpose_next_issue": purpose_next_issue,
            "approval_code": approval_code,
            "fabrication_ref": _clean_text(_first_row_value(raw, "14-Tr_Aol-Fabrication", "Fabrication")),
            "status_bucket": status_bucket if is_countable else "",
            "doc_status": doc_status if is_countable else "",
            "doc_status_group": doc_status_group if is_countable else "",
            "afc_code": afc_code if is_countable else "",
            "document_status_effective": document_status_effective,
            "document_status_original": document_status_original,
            "is_transmittal": is_transmittal,
            "is_in_sample": is_countable,
            "is_countable": is_countable,
            "excluded_reason": "" if is_countable else excluded_reason,
        })
    return rows


def import_engineering_monitor_workbook(uploaded_file: Any, *, imported_by: Any = None) -> EngineeringMonitorImport:
    filename = Path(getattr(uploaded_file, "name", "") or "engineering-monitor.xlsx").name
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("Upload an engineering monitor workbook in .xlsx or .xlsm format.")

    content = uploaded_file.read()
    if not content:
        raise ValueError("The uploaded engineering monitor workbook is empty.")

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    official_ws = workbook["Base Engenharia"] if "Base Engenharia" in workbook.sheetnames else workbook["Base AOL"] if "Base AOL" in workbook.sheetnames else _find_sheet(workbook, MONITOR_OFFICIAL_HEADERS)
    detail_ws = None
    import_mode = "official_aol" if official_ws is not None else "raw_engineering"
    if official_ws is None:
        preferred_ws = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else None
        detail_ws = (
            preferred_ws
            if preferred_ws is not None and _worksheet_has_headers(preferred_ws, MONITOR_REQUIRED_HEADERS)
            else _find_sheet(workbook, MONITOR_REQUIRED_HEADERS)
        )
    if detail_ws is None:
        if official_ws is None:
            raise ValueError("The document detail sheet was not found in the monitor workbook.")

    rows = _parse_official_rows(official_ws) if official_ws is not None else _parse_monitor_rows(detail_ws)
    if not rows:
        raise ValueError("No valid documents were found in the monitor workbook.")

    monitored_rows = [row for row in rows if row["is_monitored"]]
    countable_rows = [row for row in rows if row["is_countable"]]
    excluded_rows = [row for row in rows if not row["is_countable"]]
    status_counts = Counter(row["status_bucket"] for row in countable_rows)
    excluded_counts = Counter(row["excluded_reason"] for row in excluded_rows)
    discipline_counts = Counter(row["discipline"] for row in countable_rows)
    source_discipline_counts = Counter(row["source_discipline"] for row in rows)

    file_hash = hashlib.sha256(content).hexdigest()
    user = getattr(imported_by, "_wrapped", imported_by)
    if not getattr(user, "is_authenticated", False):
        user = None

    discipline_order = monitor_discipline_order(discipline_counts)

    sample_rules = {
        "included_roots": list(MONITOR_SAMPLE_ROOTS),
        "excluded_folders": list(MONITOR_SAMPLE_EXCLUDED_FOLDERS),
        "cancelled_document_statuses": sorted(MONITOR_CANCELLED_DOCUMENT_STATUSES),
        "excluded_purpose_next_issues": sorted(MONITOR_EXCLUDED_PURPOSE_NEXT_ISSUES),
        "exclude_revisions_containing": ".",
        "classification": {
            "afc_afu": {
                "purpose_values": [MONITOR_AFC_PURPOSE, ""],
                "conditional_purpose": MONITOR_IFF_PURPOSE,
                "approval_codes": list(MONITOR_IFF_APPROVAL_CODES),
            },
            "ifa": {
                "purpose": MONITOR_IFA_PURPOSE,
                "exclude_document_number_contains": "RA-7769",
            },
            "ifr": {
                "purpose": MONITOR_IFR_PURPOSE,
                "exclude_document_number_contains": "3323",
            },
        },
    }
    sample_funnel = {
        "source_documents": len(rows),
        "sample_documents": len(countable_rows),
    }
    if import_mode == "raw_engineering":
        outside_roots = int(excluded_counts.get("Outside configured sample folders", 0))
        excluded_folders = int(excluded_counts.get("Excluded sample folder", 0))
        cancelled_status = int(excluded_counts.get("Cancelled document status", 0))
        excluded_purpose = int(excluded_counts.get("Excluded purpose next issue", 0))
        excluded_revision = int(excluded_counts.get("Revision contains field mark", 0))
        classification_exclusions = sum(
            int(excluded_counts.get(reason, 0))
            for reason in (
                "IFF without approved return code",
                "IFA RA-7769 excluded",
                "IFR 3323 excluded",
                "Purpose next issue outside monitored statuses",
            )
        )
        sample_funnel.update({
            "inside_allowed_roots": len(rows) - outside_roots,
            "excluded_folders": excluded_folders,
            "after_folder_exclusions": len(rows) - outside_roots - excluded_folders,
            "excluded_cancelled_status": cancelled_status,
            "after_cancelled_status": len(rows) - outside_roots - excluded_folders - cancelled_status,
            "excluded_purpose_next_issue": excluded_purpose,
            "after_purpose_next_issue": len(rows) - outside_roots - excluded_folders - cancelled_status - excluded_purpose,
            "excluded_revision_field_marks": excluded_revision,
            "after_revision_filter": len(rows) - outside_roots - excluded_folders - cancelled_status - excluded_purpose - excluded_revision,
            "excluded_by_classification": classification_exclusions,
        })

    payload = {
        "documents": rows,
        "status_order": list(MONITOR_STATUS_ORDER),
        "discipline_order": discipline_order,
        "import_mode": import_mode,
        "sample_rules": sample_rules,
    }

    with transaction.atomic():
        EngineeringMonitorImport.objects.filter(is_active=True).update(is_active=False)
        batch = EngineeringMonitorImport.objects.create(
            original_filename=filename,
            file_size=len(content),
            file_hash=file_hash,
            imported_by=user,
            detail_sheet=(official_ws or detail_ws).title,
            document_count=len(rows),
            monitored_document_count=len(countable_rows),
            discipline_count=len(discipline_counts),
            excluded_count=len(excluded_rows),
            payload=payload,
            metadata={
                "sheet_names": workbook.sheetnames,
                "total_sheet_count": len(workbook.sheetnames),
                "import_mode": import_mode,
                "raw_document_count": len(rows),
                "monitored_raw_document_count": len(monitored_rows),
                "status_counts": dict(status_counts),
                "excluded_counts": dict(excluded_counts),
                "source_discipline_counts": dict(source_discipline_counts),
                "discipline_counts": dict(discipline_counts),
                "sample_rules": sample_rules,
                "sample_funnel": sample_funnel,
            },
        )

    cache.clear()
    return batch
