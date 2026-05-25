from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

from django.core.cache import cache
from django.db import transaction

from .models import (
    EngineeringDisciplineStatus,
    EngineeringDocumentStatus,
    EngineeringStatusImport,
)


SUMMARY_HEADERS = {"DISCIPLINE", "TOTAL", "REMARKS"}
DETAIL_HEADERS = {"Document Number", "Title", "Discipline", "Revision", "Doc Status", "Document Status"}


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").strip().split())


def _text_or_dash(value: Any) -> str:
    text = _clean_text(value)
    return text if text else "-"


def _doc_status_label(value: Any) -> str:
    text = _text_or_dash(value).upper()
    if text == "MABU":
        return "UNDER REVIEW"
    if "MABU" in text and "UNDER REVIEW" in text:
        return " ".join(part for part in text.split() if part != "MABU")
    return text


def _int(value: Any) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(str(value).replace(",", ".")))
    except (TypeError, ValueError):
        return 0


def _excel_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_text(value)
    if not text:
        return None
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _find_sheet(workbook: Any, required_headers: set[str]) -> Any | None:
    for ws in workbook.worksheets:
        for values in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5), values_only=True):
            headers = {_clean_text(value).replace("\n", " ") for value in values if _clean_text(value)}
            if required_headers.issubset(headers):
                return ws
    return None


def _revision_parts(value: Any) -> tuple[str, int]:
    text = _clean_text(value).upper()
    if not text or text == "-":
        return "", 0
    match = re.match(r"([A-Z]+)\s*0*(\d+)?", text)
    if not match:
        return text[:1], 0
    family = match.group(1)[:1]
    number = int(match.group(2) or 0)
    return family, number


def _doc_status_parts(value: Any) -> tuple[str, str]:
    text = _clean_text(value).upper()
    if not text:
        return "", ""
    if text.startswith("NOT ISSUED"):
        return "NOT ISSUED", ""
    if text.startswith("IFR"):
        return "IFR", ""
    if text.startswith("IFA"):
        return "IFA", ""
    if text.startswith("AFC"):
        if "UNDER REVIEW" in text or "MABU" in text:
            return "AFC", "UNDER REVIEW"
        match = re.search(r"CODE\s*(\d+)", text)
        return "AFC", match.group(1) if match else ""
    return text.split(" - ", 1)[0], ""


def _parse_summary_rows(ws: Any) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    totals: dict[str, Any] = {}
    for row_number, values in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        discipline = _clean_text(values[0] if len(values) > 0 else "")
        if not discipline:
            continue
        row = {
            "row_number": row_number,
            "discipline": discipline,
            "not_issued": _int(values[1] if len(values) > 1 else None),
            "ifr": _int(values[2] if len(values) > 2 else None),
            "ifa": _int(values[3] if len(values) > 3 else None),
            "afc_code1": _int(values[4] if len(values) > 4 else None),
            "afc_code3": _int(values[5] if len(values) > 5 else None),
            "mabu_under_review": _int(values[6] if len(values) > 6 else None),
            "total": _int(values[7] if len(values) > 7 else None),
            "remarks": _clean_text(values[8] if len(values) > 8 else ""),
        }
        if discipline.upper() == "TOTAL":
            totals = row
            continue
        if discipline == "0":
            continue
        rows.append(row)
    return rows, totals


def _parse_detail_rows(ws: Any) -> list[dict[str, Any]]:
    iterator = ws.iter_rows(values_only=True)
    raw_headers = next(iterator)
    headers = [_clean_text(value).replace("\n", " ") for value in raw_headers]
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(iterator, start=2):
        raw = dict(zip(headers, values))
        document_number = _clean_text(raw.get("Document Number"))
        if not document_number:
            continue
        revision = _clean_text(raw.get("Revision")).upper()
        revision_family, revision_number = _revision_parts(revision)
        doc_status = _doc_status_label(raw.get("Doc Status"))
        status_group, afc_code = _doc_status_parts(doc_status)
        rows.append({
            "row_number": row_number,
            "document_number": document_number,
            "title": _clean_text(raw.get("Title")),
            "discipline": _text_or_dash(raw.get("Discipline")).upper(),
            "revision": revision,
            "revision_family": revision_family,
            "revision_number": revision_number,
            "doc_status": doc_status,
            "doc_status_group": status_group,
            "afc_code": afc_code,
            "document_status": _text_or_dash(raw.get("Document Status")).upper(),
            "workflow_start": _excel_date(raw.get("Workflow Start")),
            "workflow_end": _excel_date(raw.get("Workflow End")),
            "responsible": _clean_text(raw.get("Person Responsible for Activity")).upper(),
            "issue_status": _text_or_dash(raw.get("Issue Status")).upper(),
            "fabrication_ref": _clean_text(raw.get("14-Tr_Aol-Fabrication")),
        })
    return rows


def import_engineering_status_workbook(uploaded_file: Any, *, imported_by: Any = None) -> EngineeringStatusImport:
    filename = getattr(uploaded_file, "name", "") or "ded-status.xlsx"
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("Envie a planilha DED em formato .xlsx ou .xlsm.")

    content = uploaded_file.read()
    if not content:
        raise ValueError("O arquivo enviado esta vazio.")

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    summary_ws = workbook["Sheet2"] if "Sheet2" in workbook.sheetnames else _find_sheet(workbook, SUMMARY_HEADERS)
    detail_ws = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else _find_sheet(workbook, DETAIL_HEADERS)
    if summary_ws is None:
        raise ValueError("A aba de resumo por disciplina nao foi encontrada.")
    if detail_ws is None:
        raise ValueError("A aba detalhada de documentos nao foi encontrada.")

    summary_rows, summary_totals = _parse_summary_rows(summary_ws)
    detail_rows = _parse_detail_rows(detail_ws)
    if not detail_rows:
        raise ValueError("Nenhum documento valido foi encontrado na aba detalhada.")

    file_hash = hashlib.sha256(content).hexdigest()
    user = getattr(imported_by, "_wrapped", imported_by)
    if not getattr(user, "is_authenticated", False):
        user = None

    with transaction.atomic():
        EngineeringStatusImport.objects.filter(is_active=True).update(is_active=False)
        batch = EngineeringStatusImport.objects.create(
            original_filename=filename,
            file_size=len(content),
            file_hash=file_hash,
            imported_by=user,
            summary_sheet=summary_ws.title,
            detail_sheet=detail_ws.title,
            document_count=len(detail_rows),
            discipline_count=len({row["discipline"] for row in detail_rows}),
            summary_row_count=len(summary_rows),
            metadata={
                "sheet_names": workbook.sheetnames,
                "total_sheet_count": len(workbook.sheetnames),
                "summary_totals": summary_totals,
            },
        )
        EngineeringDisciplineStatus.objects.bulk_create([
            EngineeringDisciplineStatus(import_batch=batch, **row)
            for row in summary_rows
        ])
        EngineeringDocumentStatus.objects.bulk_create([
            EngineeringDocumentStatus(import_batch=batch, **row)
            for row in detail_rows
        ])

    cache.clear()
    return batch
