from __future__ import annotations

import hashlib
import json
import sqlite3
from collections import Counter
from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from django.conf import settings
from django.core.cache import cache
from django.db.utils import OperationalError, ProgrammingError

from apps.eclic.api_client import EclicAPIError, EclicClient
from apps.core.engineering_monitor_import import monitor_discipline_order, normalize_monitor_discipline
from apps.core.models import DatafySupplySnapshot, EngineeringMonitorImport, EngineeringStatusImport, P6CurveImport
from apps.core.supply_snapshot_filters import (
    hash_supply_snapshot_filters,
    normalized_supply_snapshot_filters,
)


ECLIC_MANAGERIAL_DISCIPLINES = [
    "MP - PIPING",
    "IN - INSTRUMENTATION",
    "EA - ELECTRICAL",
    "CS - STRUCTURAL",
    "QA - QUALITY",
    "MX - OVERAL MECHANICAL",
    "MR - MECHANICAL ROTATING",
    "MS - MECHANICAL STATIC",
    "TA - TELECOMMUNICATIONS",
    "BC - COMMISSIONING",
    "MH - HVAC",
    "SA - LOGISTICS",
]

P6_MONTHS = {
    "jan": 1,
    "janeiro": 1,
    "feb": 2,
    "fev": 2,
    "fevereiro": 2,
    "mar": 3,
    "marco": 3,
    "apr": 4,
    "abr": 4,
    "abril": 4,
    "may": 5,
    "mai": 5,
    "maio": 5,
    "jun": 6,
    "junho": 6,
    "jul": 7,
    "julho": 7,
    "aug": 8,
    "ago": 8,
    "agosto": 8,
    "sep": 9,
    "set": 9,
    "setembro": 9,
    "oct": 10,
    "out": 10,
    "outubro": 10,
    "nov": 11,
    "novembro": 11,
    "dec": 12,
    "dez": 12,
    "dezembro": 12,
}

P6_MONTH_LABELS = {
    1: "jan",
    2: "fev",
    3: "mar",
    4: "abr",
    5: "mai",
    6: "jun",
    7: "jul",
    8: "ago",
    9: "set",
    10: "out",
    11: "nov",
    12: "dez",
}


def _empty_chart(message: str = "Sem dados") -> dict:
    return {
        "data": [],
        "layout": {
            "height": 280,
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "annotations": [{
                "text": message,
                "xref": "paper",
                "yref": "paper",
                "x": 0.5,
                "y": 0.5,
                "showarrow": False,
                "font": {"size": 12, "color": "#737373"},
            }],
            "xaxis": {"visible": False},
            "yaxis": {"visible": False},
        },
    }


def _bar_chart(labels: list[str], values: list[int | float], *, color: str = "#0a0a0a") -> dict:
    if not labels:
        return _empty_chart()
    horizontal = any(len(str(label)) > 12 for label in labels)
    trace = {
        "type": "bar",
        "marker": {"color": color},
        "hovertemplate": "%{y}<br>%{x}<extra></extra>" if horizontal else "%{x}<br>%{y}<extra></extra>",
    }
    if horizontal:
        trace.update({"x": values, "y": labels, "orientation": "h"})
    else:
        trace.update({"x": labels, "y": values})
    return {
        "data": [trace],
        "layout": {
            "height": 280,
            "margin": {"l": 178 if horizontal else 42, "r": 12, "t": 8, "b": 28 if horizontal else 46},
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "font": {"family": "Inter, sans-serif", "size": 11, "color": "#525252"},
            "xaxis": {"gridcolor": "#ededed", "tickangle": 0 if horizontal else -12},
            "yaxis": {
                "gridcolor": "#ededed",
                "zerolinecolor": "#d4d4d4",
                **({"categoryorder": "total ascending"} if horizontal else {}),
            },
            "showlegend": False,
        },
    }


def _grouped_bar_chart(labels: list[str], series: list[dict[str, Any]], *, height: int = 280) -> dict:
    if not labels or not series:
        return _empty_chart()
    palette = ["#0a0a0a", "#737373", "#b45309", "#1d4ed8", "#b91c1c"]
    traces = []
    for index, item in enumerate(series):
        traces.append({
            "type": "bar",
            "name": item["name"],
            "x": labels,
            "y": item["values"],
            "marker": {"color": item.get("color") or palette[index % len(palette)]},
            "hovertemplate": "%{x}<br>%{fullData.name}: %{y}<extra></extra>",
        })
    return {
        "data": traces,
        "layout": {
            "height": height,
            "barmode": "group",
            "margin": {"l": 42, "r": 12, "t": 8, "b": 46},
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "font": {"family": "Inter, sans-serif", "size": 11, "color": "#525252"},
            "xaxis": {"gridcolor": "#ededed", "tickangle": -12},
            "yaxis": {"gridcolor": "#ededed", "zerolinecolor": "#d4d4d4"},
            "legend": {"orientation": "h", "y": -0.18},
        },
    }


def _line_chart(labels: list[str], series: list[dict[str, Any]], *, height: int = 320) -> dict:
    if not labels or not series:
        return _empty_chart()
    palette = ["#0a0a0a", "#b91c1c", "#737373", "#1d4ed8"]
    traces = []
    for index, item in enumerate(series):
        traces.append({
            "type": "scatter",
            "mode": "lines+markers",
            "name": item["name"],
            "x": labels,
            "y": item["values"],
            "line": {"color": item.get("color") or palette[index % len(palette)], "width": 2},
            "marker": {"size": 5},
            "hovertemplate": "%{x}<br>%{fullData.name}: %{y:,.0f}<extra></extra>",
        })
    return {
        "data": traces,
        "layout": {
            "height": height,
            "margin": {"l": 54, "r": 18, "t": 8, "b": 42},
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "font": {"family": "Inter, sans-serif", "size": 11, "color": "#525252"},
            "xaxis": {"gridcolor": "#ededed", "tickangle": -18},
            "yaxis": {"gridcolor": "#ededed", "zerolinecolor": "#d4d4d4"},
            "legend": {"orientation": "h", "y": -0.18},
        },
    }


def _gantt_chart(rows: list[dict[str, Any]], *, height: int = 380) -> dict:
    valid_rows = [row for row in rows if row.get("start") and row.get("finish")]
    if not valid_rows:
        return _empty_chart()
    labels = [
        f"{row.get('activity_id') or '-'} - {str(row.get('name') or '')[:42]}"
        for row in valid_rows
    ]
    x_values: list[Any] = []
    y_values: list[str] = []
    text_values: list[str] = []
    for label, row in zip(labels, valid_rows):
        x_values.extend([_fmt_date(row.get("start")), _fmt_date(row.get("finish")), None])
        y_values.extend([label, label, None])
        text_values.extend([
            f"{row.get('activity_id') or '-'}<br>{row.get('start')} / {row.get('finish')}<br>{row.get('percent_complete') or 0}% complete",
            f"{row.get('activity_id') or '-'}<br>{row.get('start')} / {row.get('finish')}<br>{row.get('percent_complete') or 0}% complete",
            "",
        ])
    return {
        "data": [{
            "type": "scatter",
            "mode": "lines+markers",
            "x": x_values,
            "y": y_values,
            "text": text_values,
            "line": {"color": "#0a0a0a", "width": 5},
            "marker": {"size": 6, "color": "#737373"},
            "hovertemplate": "%{text}<extra></extra>",
        }],
        "layout": {
            "height": height,
            "margin": {"l": 250, "r": 18, "t": 8, "b": 44},
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "font": {"family": "Inter, sans-serif", "size": 10, "color": "#525252"},
            "xaxis": {"type": "date", "gridcolor": "#ededed"},
            "yaxis": {"gridcolor": "#ededed", "autorange": "reversed", "automargin": True},
            "showlegend": False,
        },
    }


def _stacked_horizontal_chart(labels: list[str], series: list[dict[str, Any]], *, height: int = 320) -> dict:
    if not labels or not series:
        return _empty_chart()
    palette = ["#0a0a0a", "#737373", "#a3a3a3", "#b45309", "#b91c1c", "#1d4ed8", "#15803d"]
    traces = []
    for index, item in enumerate(series):
        traces.append({
            "type": "bar",
            "orientation": "h",
            "name": item["name"],
            "x": item["values"],
            "y": labels,
            "marker": {"color": item.get("color") or palette[index % len(palette)]},
            "hovertemplate": "%{y}<br>%{fullData.name}: %{x}<extra></extra>",
        })
    return {
        "data": traces,
        "layout": {
            "height": height,
            "barmode": "stack",
            "margin": {"l": 190, "r": 12, "t": 8, "b": 34},
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "font": {"family": "Inter, sans-serif", "size": 11, "color": "#525252"},
            "xaxis": {"gridcolor": "#ededed"},
            "yaxis": {"gridcolor": "#ededed"},
            "legend": {"orientation": "h", "y": -0.14},
        },
    }


def _donut_chart(labels: list[str], values: list[int | float]) -> dict:
    if not labels:
        return _empty_chart()
    return {
        "data": [{
            "type": "pie",
            "labels": labels,
            "values": values,
            "hole": 0.62,
            "sort": False,
            "marker": {"colors": ["#0a0a0a", "#737373", "#a3a3a3", "#b45309", "#b91c1c", "#1d4ed8"]},
            "textinfo": "label+percent",
            "hovertemplate": "%{label}<br>%{value}<extra></extra>",
        }],
        "layout": {
            "height": 280,
            "margin": {"l": 10, "r": 10, "t": 8, "b": 8},
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "font": {"family": "Inter, sans-serif", "size": 11, "color": "#525252"},
            "legend": {"orientation": "h", "y": -0.12},
        },
    }


def _expected_date_key(value: Any) -> tuple[int, str]:
    text = str(value or "").strip()
    if not text:
        return (0, "")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return (1, datetime.strptime(text, fmt).date().isoformat())
        except ValueError:
            continue
    return (0, text)


def _expected_date_iso(value: Any) -> str:
    ok, parsed = _expected_date_key(value)
    return parsed if ok else ""


def _is_delivery_at_yard_actual(stage: Any, kind: Any) -> bool:
    stage_text = str(stage or "").strip().lower()
    kind_text = str(kind or "").strip().lower()
    return "delivery" in stage_text and "yard" in stage_text and kind_text == "actual"


def _po_row_has_yard_actual(row: dict[str, Any]) -> bool:
    candidates: list[str] = []
    raw_payload = row.get("procurement_plan_payload")
    payload = {}
    if raw_payload:
        try:
            payload = json.loads(raw_payload) if isinstance(raw_payload, str) else raw_payload
        except (TypeError, ValueError):
            payload = {}
    timeline = payload.get("timeline") if isinstance(payload, dict) else []
    for item in timeline or []:
        if not isinstance(item, dict):
            continue
        if not _is_delivery_at_yard_actual(item.get("stage"), item.get("kind")):
            continue
        iso = _expected_date_iso(item.get("date"))
        if iso:
            candidates.append(iso)

    if _is_delivery_at_yard_actual(row.get("procurement_plan_stage"), row.get("procurement_plan_kind")):
        iso = _expected_date_iso(row.get("procurement_plan_date"))
        if iso:
            candidates.append(iso)

    today = date.today().isoformat()
    return any(value <= today for value in candidates)


def _po_delivery_date_iso(row: dict[str, Any]) -> str:
    raw_payload = row.get("procurement_plan_payload")
    payload = {}
    if raw_payload:
        try:
            payload = json.loads(raw_payload) if isinstance(raw_payload, str) else raw_payload
        except (TypeError, ValueError):
            payload = {}
    timeline = payload.get("timeline") if isinstance(payload, dict) else []
    for kind in ("Actual", "Forecast", "Planned"):
        dates = []
        for item in timeline or []:
            if not isinstance(item, dict):
                continue
            stage_text = str(item.get("stage") or "").lower()
            kind_text = str(item.get("kind") or "").lower()
            if "delivery" not in stage_text or "yard" not in stage_text or kind_text != kind.lower():
                continue
            iso = _expected_date_iso(item.get("date"))
            if iso:
                dates.append(iso)
        if dates:
            return sorted(dates)[-1]

    stage_text = str(row.get("procurement_plan_stage") or "").lower()
    if "delivery" in stage_text and "yard" in stage_text:
        iso = _expected_date_iso(row.get("procurement_plan_date"))
        if iso:
            return iso
    return _expected_date_iso(row.get("procurement_plan_date"))


def datafy_po_delivery_lookup(po_numbers: list[str] | set[str] | tuple[str, ...]) -> dict[str, dict[str, str]]:
    clean = []
    seen = set()
    for value in po_numbers or []:
        po = str(value or "").strip()
        key = po.upper()
        if not po or key in seen or key in {"NO LINKED PO", "COVERED", "-"}:
            continue
        seen.add(key)
        clean.append(po)
    if not clean:
        return {}

    placeholders = ", ".join("?" for _ in clean)
    try:
        with _datafy_conn() as conn:
            rows = _rows(conn.cursor(), f"""
                select po_number,
                       procurement_plan_stage,
                       procurement_plan_kind,
                       procurement_plan_date,
                       procurement_plan_payload
                from core_purchaseorder
                where po_number in ({placeholders})
            """, tuple(clean))
    except Exception:
        return {}

    output = {}
    for row in rows:
        po = str(row.get("po_number") or "").strip()
        if not po:
            continue
        output[po.upper()] = {
            "po": po,
            "expected_date": _po_delivery_date_iso(row),
            "stage": str(row.get("procurement_plan_stage") or ""),
            "kind": str(row.get("procurement_plan_kind") or ""),
        }
    return output


def _material_flow_summary(total: int, covered: int, at_aveon: int) -> dict[str, Any]:
    total = max(int(total or 0), 0)
    covered = max(int(covered or 0), 0)
    at_aveon = min(max(int(at_aveon or 0), 0), covered)
    po_placed = max(covered - at_aveon, 0)
    tbe = max(total - covered, 0)

    def pct(value: int) -> float:
        return round(100 * value / total, 2) if total else 0.0

    covered_pct = pct(covered)
    at_aveon_pct = pct(at_aveon)
    tbe_pct = pct(tbe)
    po_placed_pct = pct(po_placed)
    items = [
        {
            "key": "po_placed",
            "label": "PO placed",
            "value": po_placed,
            "pct": po_placed_pct,
            "pct_css": f"{po_placed_pct:.2f}",
            "color": "#2563eb",
            "icon": "bi-receipt-cutoff",
            "detail": "Pedido colocado, aguardando chegada.",
        },
        {
            "key": "at_aveon",
            "label": "Material at Aveon",
            "value": at_aveon,
            "pct": at_aveon_pct,
            "pct_css": f"{at_aveon_pct:.2f}",
            "color": "#059669",
            "icon": "bi-check2-circle",
            "detail": "Material com chegada confirmada no Aveon.",
        },
        {
            "key": "tbe",
            "label": "TBE",
            "value": tbe,
            "pct": tbe_pct,
            "pct_css": f"{tbe_pct:.2f}",
            "color": "#d97706",
            "icon": "bi-exclamation-triangle",
            "detail": "Exposicao ainda sem PO no balanceamento.",
        },
    ]
    return {
        "total": total,
        "covered": covered,
        "covered_pct": covered_pct,
        "covered_pct_css": f"{covered_pct:.2f}",
        "po_placed": po_placed,
        "po_placed_pct": po_placed_pct,
        "at_aveon": at_aveon,
        "at_aveon_pct": at_aveon_pct,
        "tbe": tbe,
        "tbe_pct": tbe_pct,
        "items": items,
    }


def _supply_scope_from_table(value: Any) -> str:
    text = str(value or "").strip().upper()
    if any(token in text for token in ("DEMOLISH", "DEMOLI", "REMOV")):
        return "other"
    if any(token in text for token in ("ERECTION", "INSTALLATION", "ONBOARD")):
        return "erection"
    if any(token in text for token in ("FABRICATION", "FABRICAC")):
        return "fabrication"
    return "fabrication"


def _supply_scope_label(value: Any) -> str:
    scope = str(value or "").strip().lower()
    if scope == "erection":
        return "Erection"
    if scope == "fabrication":
        return "Fabrication"
    return "Other"


def _supply_material_status(row: dict[str, Any]) -> tuple[str, str]:
    missing = float(row.get("missing_qty") or 0)
    allocated = float(row.get("allocated_qty") or 0)
    if missing <= 0:
        return "ok", "Covered"
    if allocated > 0:
        return "partial", "Partial"
    if int(row.get("stock_free_na") or 0):
        return "unknown", "Check"
    return "missing", "Missing"


def _truthy_flag(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value) > 0
    text = str(value or "").strip().lower()
    if not text:
        return False
    if text in {"1", "true", "yes", "sim", "y"}:
        return True
    try:
        return float(text.replace(",", ".")) > 0
    except ValueError:
        return False


def _supply_po_covering_has_po(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    normalized = text.lower()
    return normalized not in {"no linked po", "sem po", "without po", "no po", "-"}


def _supply_enrich_material_rows(
    rows: list[dict[str, Any]],
    yard_material_ids: set[Any] | None = None,
    po_material_ids: set[Any] | None = None,
) -> dict[str, int]:
    has_yard_ids = yard_material_ids is not None
    has_po_ids = po_material_ids is not None
    yard_material_ids = yard_material_ids or set()
    po_material_ids = po_material_ids or set()
    counts = {"ok": 0, "partial": 0, "missing": 0, "unknown": 0}
    for row in rows:
        status, label = _supply_material_status(row)
        counts[status] = counts.get(status, 0) + 1
        scope = row.get("scope") or _supply_scope_from_table(row.get("table_name"))
        if has_po_ids:
            has_po = row.get("material_item_id") in po_material_ids
        elif row.get("has_po") is not None:
            has_po = _truthy_flag(row.get("has_po"))
        else:
            has_po = _supply_po_covering_has_po(row.get("po_covering"))
        if has_yard_ids:
            yard_actual = row.get("material_item_id") in yard_material_ids
        else:
            yard_actual = _truthy_flag(row.get("yard_actual"))
        row["scope"] = scope
        row["scope_label"] = _supply_scope_label(scope)
        row["status"] = status
        row["status_label"] = label
        row["has_po"] = 1 if has_po else 0
        row["yard_actual"] = 1 if yard_actual else 0
        row["stock_free_label"] = "n/a" if row.get("stock_free_na") else row.get("stock_free_qty") or 0
        if not row.get("po_covering"):
            row["po_covering"] = "No linked PO" if status != "ok" else "Covered"
    return counts


def _supply_build_drawing_line_rows(
    material_rows: list[dict[str, Any]],
    *,
    limit: int = 900,
) -> list[dict[str, Any]]:
    groups: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in material_rows:
        scope = row.get("scope") or _supply_scope_from_table(row.get("table_name"))
        if scope == "other":
            continue
        drawing_number = row.get("drawing_number") or row.get("original_filename") or "-"
        campaign = _campaign_label(row.get("campaign"))
        key = (
            row.get("document_id"),
            drawing_number,
            scope,
            campaign,
            row.get("priority"),
        )
        group = groups.setdefault(key, {
            "document_id": row.get("document_id"),
            "drawing_number": drawing_number,
            "revision": row.get("revision") or "-",
            "revision_detail": row.get("revision_detail") or "-",
            "discipline": row.get("discipline") or "-",
            "scope": scope,
            "scope_label": _supply_scope_label(scope),
            "campaign": campaign,
            "priority": row.get("priority"),
            "total_items": 0,
            "active_items": 0,
            "finalized_items": 0,
            "covered": 0,
            "partial": 0,
            "pending": 0,
            "with_po": 0,
            "without_po": 0,
            "po_not_arrived": 0,
            "yard_actual": 0,
            "stage_total_items": 0,
            "stage_total_pending": 0,
            "stage_po_items": 0,
            "stage_po_pending": 0,
            "stage_no_po_items": 0,
            "stage_no_po_pending": 0,
            "stage_no_yard_items": 0,
            "stage_no_yard_pending": 0,
            "stage_yard_items": 0,
            "stage_yard_pending": 0,
            "families": set(),
            "pending_families": set(),
            "lines": set(),
        })
        line = str(row.get("line") or "").strip()
        if line and line != "-":
            group["lines"].add(line)
        group["total_items"] += 1
        if _supply_is_finalized_item(row):
            group["finalized_items"] += 1
            continue

        group["active_items"] += 1
        missing = float(row.get("missing_qty") or 0)
        allocated = float(row.get("allocated_qty") or 0)
        has_po = _truthy_flag(row.get("has_po"))
        yard_actual = _truthy_flag(row.get("yard_actual"))
        if missing <= 0:
            group["covered"] += 1
        else:
            group["pending"] += 1
            if allocated > 0:
                group["partial"] += 1
        if has_po:
            group["with_po"] += 1
        else:
            group["without_po"] += 1
        if has_po and not yard_actual:
            group["po_not_arrived"] += 1
        if yard_actual:
            group["yard_actual"] += 1

        group["stage_total_items"] += 1
        if missing > 0:
            group["stage_total_pending"] += 1
        if has_po:
            group["stage_po_items"] += 1
            if missing > 0:
                group["stage_po_pending"] += 1
        else:
            group["stage_no_po_items"] += 1
            if missing > 0:
                group["stage_no_po_pending"] += 1
        if has_po and not yard_actual:
            group["stage_no_yard_items"] += 1
            if missing > 0:
                group["stage_no_yard_pending"] += 1
        if yard_actual:
            group["stage_yard_items"] += 1
            if missing > 0:
                group["stage_yard_pending"] += 1

        family = str(row.get("family") or "").strip()
        if family and family != "-":
            group["families"].add(family)
            if missing > 0:
                group["pending_families"].add(family)

    output = []
    for index, group in enumerate(groups.values(), start=1):
        total = int(group["total_items"] or 0)
        active_total = int(group["active_items"] or 0)
        finalized_only = int(group["finalized_items"] or 0) > 0 and active_total == 0
        covered = int(group["covered"] or 0)
        pending = int(group["pending"] or 0)
        group["row_no"] = index
        group["coverage_pct"] = 100.0 if finalized_only else (round(100 * covered / active_total, 1) if active_total else 0)
        group["yard_pct"] = round(100 * int(group["yard_actual"] or 0) / active_total, 1) if active_total else 0
        group["drawing_finalized"] = 1 if finalized_only else 0
        if finalized_only:
            group["status"] = "finalized"
            group["status_label"] = "Finalized"
            group["pending_bucket"] = -1
        elif pending <= 0:
            group["status"] = "ok"
            group["status_label"] = "Covered"
            group["pending_bucket"] = 0
        elif covered or int(group["partial"] or 0):
            group["status"] = "partial"
            group["status_label"] = "Partial"
            group["pending_bucket"] = 8 if pending >= 8 else pending
        else:
            group["status"] = "missing"
            group["status_label"] = "Pending"
            group["pending_bucket"] = 8 if pending >= 8 else pending
        group["drawing_pending"] = pending
        group["line_pending_bucket"] = group["pending_bucket"]
        for stage in ("total", "po", "no_po", "no_yard", "yard"):
            stage_pending = int(group.get(f"stage_{stage}_pending") or 0)
            group[f"stage_{stage}_pending_bucket"] = 8 if stage_pending >= 8 else stage_pending
        group["has_po"] = 1 if int(group.get("with_po") or 0) > 0 else 0
        group["without_po_flag"] = 1 if int(group.get("without_po") or 0) > 0 else 0
        group["po_not_arrived_flag"] = 1 if int(group.get("po_not_arrived") or 0) > 0 else 0
        lines = sorted(group["lines"])
        group["line_count"] = len(lines)
        group["lines_full"] = ", ".join(lines) if lines else "-"
        group["line"] = ", ".join(lines[:4]) + (f" +{len(lines) - 4}" if len(lines) > 4 else "") if lines else "-"
        group["families"] = ", ".join(sorted(group["families"])) or "-"
        group["pending_families"] = ", ".join(sorted(group["pending_families"])) or "-"
        group.pop("lines", None)
        output.append(group)

    def _sort_key(item: dict[str, Any]) -> tuple[int, int, float, str, str]:
        priority = item.get("priority")
        priority_missing = 1 if priority in (None, "") else 0
        try:
            priority_value = int(priority)
        except (TypeError, ValueError):
            priority_value = 999999
        return (
            priority_missing,
            priority_value,
            -float(item.get("coverage_pct") or 0),
            str(item.get("drawing_number") or ""),
            str(item.get("line") or ""),
        )

    output.sort(key=_sort_key)
    for index, group in enumerate(output, start=1):
        group["row_no"] = index
    return output[:limit]


def _supply_enrich_drawing_line_families(
    drawing_line_rows: list[dict[str, Any]],
    material_rows: list[dict[str, Any]],
) -> None:
    pending_by_line_scope: dict[tuple[str, str], set[str]] = {}
    all_by_line_scope: dict[tuple[str, str], set[str]] = {}
    for row in material_rows:
        line = str(row.get("line") or "").strip()
        family = str(row.get("family") or "").strip()
        if not line or line == "-" or not family or family == "-":
            continue
        scope = row.get("scope") or _supply_scope_from_table(row.get("table_name"))
        if scope == "other":
            continue
        key = (str(scope), line)
        all_by_line_scope.setdefault(key, set()).add(family)
        try:
            missing = float(row.get("missing_qty") or 0)
        except (TypeError, ValueError):
            missing = 0
        if missing > 0:
            pending_by_line_scope.setdefault(key, set()).add(family)

    for row in drawing_line_rows:
        scope = str(row.get("scope") or "")
        lines = [
            part.strip()
            for part in str(row.get("lines_full") or row.get("line") or "").split(",")
            if part.strip() and part.strip() != "-"
        ]
        pending_families: set[str] = set()
        all_families: set[str] = set()
        for line in lines:
            key = (scope, line)
            pending_families.update(pending_by_line_scope.get(key, set()))
            all_families.update(all_by_line_scope.get(key, set()))
        if pending_families:
            row["pending_families"] = ", ".join(sorted(pending_families))
        if all_families and str(row.get("families") or "-").strip() == "-":
            row["families"] = ", ".join(sorted(all_families))


def _supply_split_text_values(value: Any) -> list[str]:
    if isinstance(value, (list, tuple, set)):
        raw_values = value
    else:
        raw_values = str(value or "").replace("\n", ",").replace(";", ",").split(",")
    return [str(item or "").strip() for item in raw_values if str(item or "").strip()]


def _supply_forecast_delivery_date(row: dict[str, Any]) -> str:
    dates: list[str] = []
    for value in [row.get("po_expected_date"), *_supply_split_text_values(row.get("po_expected_dates"))]:
        iso = _expected_date_iso(value)
        if iso:
            dates.append(iso)
    raw_pairs = row.get("po_delivery_pairs")
    if raw_pairs:
        try:
            pairs = json.loads(raw_pairs) if isinstance(raw_pairs, str) else raw_pairs
        except (TypeError, ValueError):
            pairs = []
            if isinstance(raw_pairs, str):
                pairs = [
                    {"po": part.partition("::")[0], "expected_date": part.partition("::")[2]}
                    for part in raw_pairs.split(";;")
                    if part.strip() and part.partition("::")[2].strip()
                ]
        for pair in pairs or []:
            if not isinstance(pair, dict):
                continue
            iso = _expected_date_iso(pair.get("expected_date"))
            if iso:
                dates.append(iso)
    return sorted(set(dates))[-1] if dates else ""


def _supply_forecast_delivery_pairs(row: dict[str, Any]) -> list[dict[str, str]]:
    pairs: list[dict[str, str]] = []
    raw_pairs = row.get("po_delivery_pairs")
    parsed_pairs: Any = []
    if raw_pairs:
        if isinstance(raw_pairs, str):
            try:
                parsed_pairs = json.loads(raw_pairs)
            except (TypeError, ValueError):
                parsed_pairs = [
                    {"po": part.partition("::")[0], "expected_date": part.partition("::")[2]}
                    for part in raw_pairs.split(";;")
                    if part.strip()
                ]
        else:
            parsed_pairs = raw_pairs
    for pair in parsed_pairs or []:
        if not isinstance(pair, dict):
            continue
        po = str(pair.get("po") or "").strip()
        iso = _expected_date_iso(pair.get("expected_date"))
        if po and iso:
            pairs.append({"po": po, "expected_date": iso})

    if not pairs:
        delivery_date = row.get("po_delivery_date") or _supply_forecast_delivery_date(row)
        for po in _supply_split_text_values(row.get("po_numbers") or row.get("po_covering")):
            if po.upper() in {"NO LINKED PO", "COVERED", "-"}:
                continue
            iso = _expected_date_iso(delivery_date)
            if po and iso:
                pairs.append({"po": po, "expected_date": iso})

    seen: set[tuple[str, str]] = set()
    output: list[dict[str, str]] = []
    for pair in pairs:
        key = (pair["po"].upper(), pair["expected_date"])
        if key in seen:
            continue
        seen.add(key)
        output.append(pair)
    return output


def _supply_forecast_items_from_rows(material_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items = []
    for row in material_rows:
        scope = row.get("scope") or _supply_scope_from_table(row.get("table_name"))
        if scope not in {"fabrication", "erection"}:
            continue
        items.append({
            "document_id": row.get("document_id"),
            "drawing": row.get("drawing_number") or row.get("original_filename") or "-",
            "scope": scope,
            "campaign": _campaign_label(row.get("campaign")),
            "priority": row.get("priority"),
            "material_item_id": row.get("material_item_id"),
            "line": row.get("line") or "-",
            "is_finalized": 1 if _truthy_flag(row.get("is_finalized")) else 0,
            "requested_qty": float(row.get("requested_qty") or 0),
            "allocated_qty": float(row.get("allocated_qty") or 0),
            "missing_qty": float(row.get("missing_qty") or 0),
            "has_po": 1 if _truthy_flag(row.get("has_po")) else 0,
            "yard_actual": 1 if _truthy_flag(row.get("yard_actual")) else 0,
            "po_delivery_date": row.get("po_delivery_date") or _supply_forecast_delivery_date(row),
            "po_numbers": row.get("po_numbers") or row.get("po_covering") or "",
            "po_delivery_pairs": _supply_forecast_delivery_pairs(row),
        })
    return items


def _supply_normalize_operational_payload(payload: dict[str, Any]) -> None:
    material_rows = payload.get("material_rows") or []
    status_counts = _supply_enrich_material_rows(material_rows)
    payload["material_status_counts"] = {
        "covered": status_counts.get("ok", 0),
        "partial": status_counts.get("partial", 0),
        "missing": status_counts.get("missing", 0),
        "unknown": status_counts.get("unknown", 0),
    }
    drawing_source_rows = payload.get("material_scope_items") or material_rows
    payload["drawing_line_rows"] = _supply_build_drawing_line_rows(drawing_source_rows)
    _supply_enrich_drawing_line_families(payload["drawing_line_rows"], material_rows)
    payload["supply_forecast_items_json"] = json.dumps(
        _supply_forecast_items_from_rows(material_rows),
        ensure_ascii=False,
    )
    for view in payload.get("supply_campaign_views") or []:
        totals = view.setdefault("totals", {})
        executive = view.setdefault("executive", {})
        total = int(float(totals.get("total") or 0))
        po = int(float(totals.get("po") or 0))
        no_po = totals.get("no_po")
        if no_po in (None, ""):
            no_po = executive.get("po_gap")
        if no_po in (None, ""):
            no_po = max(total - po, 0)
        no_po = int(float(no_po or 0))
        totals["no_po"] = no_po
        executive["no_po_pct"] = _supply_ratio(no_po, total)
        executive["no_po_pct_css"] = _supply_ratio_css(no_po, total)


SUPPLY_CAMPAIGN_PALETTE = [
    ("#bfdbfe", "#2563eb"),
    ("#fecaca", "#dc2626"),
    ("#d9e7b8", "#4d7c0f"),
    ("#fde68a", "#d97706"),
    ("#ddd6fe", "#7c3aed"),
    ("#bae6fd", "#0284c7"),
    ("#e5e7eb", "#6b7280"),
]


def _campaign_label(value: Any) -> str:
    text = str(value or "").strip()
    return text if text else "-"


def _campaign_sort_key(label: str) -> tuple[int, str]:
    normalized = label.strip().lower()
    aliases = {
        "0": 0,
        "1": 1,
        "1st": 1,
        "1 tbd": 1,
        "2": 2,
        "2nd": 2,
        "3": 3,
        "3rd": 3,
        "4": 4,
        "4th": 4,
        "5": 5,
        "5th": 5,
        "6": 6,
        "6th": 6,
    }
    if normalized in aliases:
        return (aliases[normalized], normalized)
    if normalized in {"-", "tbd", "1 tbd"}:
        return (90, normalized)
    return (50, normalized)


def _supply_campaign_for_label(label: Any, index: int = 0) -> dict[str, Any]:
    text = _campaign_label(label)
    normalized = text.lower()
    fixed_palette = {
        "1": SUPPLY_CAMPAIGN_PALETTE[0],
        "1st": SUPPLY_CAMPAIGN_PALETTE[0],
        "2": SUPPLY_CAMPAIGN_PALETTE[1],
        "2nd": SUPPLY_CAMPAIGN_PALETTE[1],
        "3": SUPPLY_CAMPAIGN_PALETTE[2],
        "3rd": SUPPLY_CAMPAIGN_PALETTE[2],
        "4": SUPPLY_CAMPAIGN_PALETTE[3],
        "4th": SUPPLY_CAMPAIGN_PALETTE[3],
        "tam 20206": SUPPLY_CAMPAIGN_PALETTE[4],
        "-": SUPPLY_CAMPAIGN_PALETTE[-1],
    }
    color, accent = fixed_palette.get(
        normalized,
        SUPPLY_CAMPAIGN_PALETTE[index % len(SUPPLY_CAMPAIGN_PALETTE)],
    )
    slug = "".join(char.lower() if char.isalnum() else "_" for char in text).strip("_") or "sem_campanha"
    return {
        "key": f"campaign_{slug}",
        "label": text,
        "detail": f"Campanha {text}" if text != "-" else "Sem campanha",
        "color": color,
        "accent": accent,
    }


def _campaign_aliases(value: Any) -> list[str]:
    text = _campaign_label(value)
    normalized = text.lower()
    mapping = {
        "1": ["1", "1st"],
        "1st": ["1st", "1"],
        "2": ["2", "2nd"],
        "2nd": ["2nd", "2"],
        "3": ["3", "3rd"],
        "3rd": ["3rd", "3"],
        "4": ["4", "4th"],
        "4th": ["4th", "4"],
        "5": ["5", "5th"],
        "5th": ["5th", "5"],
        "6": ["6", "6th"],
        "6th": ["6th", "6"],
    }
    return mapping.get(normalized, [text])


def _supply_ratio(part: int | float, total: int | float, decimals: int = 1) -> float:
    return round(100 * float(part) / float(total), decimals) if total else 0.0


def _supply_ratio_css(part: int | float, total: int | float) -> str:
    value = _supply_ratio(part, total, 2)
    return f"{max(0, min(100, value)):.2f}"


def _supply_is_finalized_item(item: dict[str, Any]) -> bool:
    return bool(item.get("is_finalized"))


def _supply_campaign_views(
    scoped_items: list[dict[str, Any]],
    po_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    covered_ids = {row["material_item_id"] for row in po_rows}
    yard_ids = {
        row["material_item_id"]
        for row in po_rows
        if _po_row_has_yard_actual(row)
    }
    scopes = [
        ("fabrication", "Fabrication"),
        ("erection", "Erection"),
    ]
    stage_defs = [
        ("total", "Total scope", "All items in the campaign"),
        ("po", "With PO", "Items with a linked purchase order"),
        ("no_po", "Without PO", "Items still without a purchase order"),
        ("no_yard", "Item not arrived", "Items with PO, awaiting yard arrival confirmation"),
        ("yard", "At Yard", "Items with confirmed yard arrival"),
    ]
    bucket_defs = [
        (0, "0 pendencias"),
        (1, "1 pendencia"),
        (2, "2 pendencias"),
        (3, "3 pendencias"),
        (4, "4 pendencias"),
        (5, "5 pendencias"),
        (6, "6 pendencias"),
        (7, "7 pendencias"),
        (8, "8+ pendencias"),
    ]
    views: list[dict[str, Any]] = []

    for scope_key, scope_label in scopes:
        labels = sorted({
            _campaign_label(item.get("campaign"))
            for item in scoped_items
            if item.get("scope") == scope_key
        }, key=_campaign_sort_key)
        campaigns = [
            _supply_campaign_for_label(label, index)
            for index, label in enumerate(labels)
        ]
        campaign_by_label = {
            campaign["label"]: campaign
            for campaign in campaigns
        }
        campaign_rows = {
            campaign["key"]: {
                **campaign,
                "total": 0,
                "po": 0,
                "yard": 0,
                "finalized": 0,
            }
            for campaign in campaigns
        }
        drawing_pending: dict[tuple[Any, str], dict[str, Any]] = {}
        drawing_finalized: dict[tuple[Any, str], dict[str, Any]] = {}
        for item in scoped_items:
            if item.get("scope") != scope_key:
                continue
            campaign = campaign_by_label.get(_campaign_label(item.get("campaign")))
            if campaign is None:
                campaign = _supply_campaign_for_label(item.get("campaign"), len(campaign_by_label))
                campaign_by_label[campaign["label"]] = campaign
                campaign_rows[campaign["key"]] = {
                    **campaign,
                    "total": 0,
                    "po": 0,
                    "yard": 0,
                    "finalized": 0,
                }
            campaign_row = campaign_rows[campaign["key"]]
            material_id = item.get("material_item_id")
            drawing_key = (item.get("document_id"), campaign["key"])
            if _supply_is_finalized_item(item):
                drawing_finalized.setdefault(drawing_key, {
                    "campaign": campaign,
                    "total": 0,
                })["total"] += 1
                continue

            campaign_row["total"] += 1
            if material_id in covered_ids:
                campaign_row["po"] += 1
            if material_id in yard_ids:
                campaign_row["yard"] += 1

            drawing = drawing_pending.setdefault(drawing_key, {
                "campaign": campaign,
                "pending": 0,
                "total": 0,
                "yard": 0,
            })
            drawing["total"] += 1
            if float(item.get("missing_qty") or 0) > 0:
                drawing["pending"] += 1
            if material_id in yard_ids:
                drawing["yard"] += 1

        campaigns = list(campaign_rows.values())
        finalized_counts = {
            row["key"]: 0
            for row in campaigns
        }
        for drawing in drawing_finalized.values():
            finalized_counts[drawing["campaign"]["key"]] += 1
        finalized_drawings = sum(finalized_counts.values())
        for row in campaigns:
            row["finalized"] = finalized_counts.get(row["key"], 0)
            row["no_po"] = max(int(row["total"] or 0) - int(row["po"] or 0), 0)
            row["no_yard"] = max(int(row["po"] or 0) - int(row["yard"] or 0), 0)
        total_items = sum(row["total"] for row in campaigns)
        po_items = sum(row["po"] for row in campaigns)
        yard_items = sum(row["yard"] for row in campaigns)
        no_po_items = sum(row["no_po"] for row in campaigns)
        no_yard_items = sum(row["no_yard"] for row in campaigns)
        max_stage = max(total_items, po_items, no_po_items, no_yard_items, yard_items, 1)
        stages = []
        for key, label, detail in stage_defs:
            stage_total = sum(row[key] for row in campaigns)
            segments = []
            for row in campaigns:
                value = int(row[key])
                segments.append({
                    "key": row["key"],
                    "label": row["label"],
                    "detail": row["detail"],
                    "value": value,
                    "color": row["color"],
                    "accent": row["accent"],
                    "height_css": f"{(100 * value / max_stage) if max_stage else 0:.2f}",
                    "small": (100 * value / max_stage) < 7 if max_stage else True,
                })
            stages.append({
                "key": key,
                "label": label,
                "detail": detail,
                "value": stage_total,
                "height_css": f"{100 * stage_total / max_stage:.2f}",
                "segments": segments,
            })

        pending_counts = {
            pending: {
                row["key"]: 0
                for row in campaigns
            }
            for pending, _label in bucket_defs
        }
        # contagem de desenhos com pelo menos 1 item no yard, por bucket de pendencia
        at_yard_per_bucket = {pending: 0 for pending, _ in bucket_defs}
        for drawing in drawing_pending.values():
            pending = int(drawing["pending"] or 0)
            bucket = 8 if pending >= 8 else pending
            pending_counts[bucket][drawing["campaign"]["key"]] += 1
            if drawing.get("yard", 0) > 0:
                at_yard_per_bucket[bucket] += 1
        pending_max = max(
            [sum(row.values()) for row in pending_counts.values()] + [finalized_drawings, 1]
        )
        pending_rows = []
        finalized_segments = []
        for campaign in campaigns:
            value = int(finalized_counts.get(campaign["key"], 0))
            finalized_segments.append({
                "key": campaign["key"],
                "label": campaign["label"],
                "detail": campaign["detail"],
                "value": value,
                "color": campaign["color"],
                "accent": campaign["accent"],
                "width_css": _supply_ratio_css(value, finalized_drawings),
            })
        pending_rows.append({
            "pending": -1,
            "label": "Finalizados",
            "value": finalized_drawings,
            "at_yard": 0,
            "width_css": f"{100 * finalized_drawings / pending_max:.2f}",
            "segments": finalized_segments,
        })
        for pending, label in bucket_defs:
            campaign_counts = pending_counts[pending]
            total = sum(campaign_counts.values())
            segments = []
            for campaign in campaigns:
                value = int(campaign_counts[campaign["key"]])
                segments.append({
                    "key": campaign["key"],
                    "label": campaign["label"],
                    "detail": campaign["detail"],
                    "value": value,
                    "color": campaign["color"],
                    "accent": campaign["accent"],
                    "width_css": f"{(100 * value / total) if total else 0:.2f}",
                })
            pending_rows.append({
                "pending": pending,
                "label": label,
                "value": total,
                "at_yard": int(at_yard_per_bucket[pending]),
                "width_css": f"{100 * total / pending_max:.2f}",
                "segments": segments,
            })

        clean_drawings = sum(pending_counts[0].values())
        light_drawings = sum(sum(pending_counts[bucket].values()) for bucket in (1, 2))
        medium_drawings = sum(sum(pending_counts[bucket].values()) for bucket in (3, 4))
        high_drawings = sum(sum(pending_counts[bucket].values()) for bucket in (5, 6, 7))
        critical_drawings = sum(pending_counts[8].values())
        total_drawings = finalized_drawings + clean_drawings + light_drawings + medium_drawings + high_drawings + critical_drawings
        pending_groups = []
        pending_group_defs = [
            ("finished", "Finalizados", "desenhos encerrados/NLA", finalized_drawings, [], "#22c55e"),
            ("clean", "0 pend.", "sem falta de material", clean_drawings, [0], "#16a34a"),
            ("light", "1-2", "baixa exposição", light_drawings, [1, 2], "#65a30d"),
            ("medium", "3-4", "atenção no curto prazo", medium_drawings, [3, 4], "#d97706"),
            ("high", "5-7", "risco alto de bloqueio", high_drawings, [5, 6, 7], "#dc2626"),
            ("critical", "8+", "crítico para liberação", critical_drawings, [8], "#991b1b"),
        ]
        for key, label, detail, total, buckets, tone in pending_group_defs:
            segments = []
            for campaign in campaigns:
                if key == "finished":
                    value = finalized_counts.get(campaign["key"], 0)
                else:
                    value = sum(pending_counts[bucket][campaign["key"]] for bucket in buckets)
                segments.append({
                    "key": campaign["key"],
                    "label": campaign["label"],
                    "detail": campaign["detail"],
                    "value": value,
                    "color": campaign["color"],
                    "accent": campaign["accent"],
                    "width_css": _supply_ratio_css(value, total),
                })
            pending_groups.append({
                "key": key,
                "label": label,
                "detail": detail,
                "value": total,
                "tone": tone,
                "width_css": _supply_ratio_css(total, total_drawings),
                "pct": _supply_ratio(total, total_drawings),
                "segments": segments,
            })

        campaign_max = max([int(row["total"]) for row in campaigns] + [1])
        campaign_exec_rows = []
        for row in campaigns:
            total = int(row["total"] or 0)
            po = int(row["po"] or 0)
            yard = int(row["yard"] or 0)
            no_po = int(row.get("no_po") or max(total - po, 0))
            no_yard = int(row.get("no_yard") or max(po - yard, 0))
            campaign_exec_rows.append({
                **row,
                "no_po": no_po,
                "no_yard": no_yard,
                "total_width_css": _supply_ratio_css(total, campaign_max),
                "po_total_width_css": _supply_ratio_css(po, campaign_max),
                "yard_total_width_css": _supply_ratio_css(yard, campaign_max),
                "po_width_css": _supply_ratio_css(po, total),
                "yard_width_css": _supply_ratio_css(yard, total),
                "po_pct": _supply_ratio(po, total),
                "yard_pct": _supply_ratio(yard, total),
                "yard_of_po_pct": _supply_ratio(yard, po),
                "risk_items": no_po + no_yard,
            })
        bottleneck_rows = sorted(
            campaign_exec_rows,
            key=lambda row: (row["risk_items"], row["no_po"], row["no_yard"], row["total"]),
            reverse=True,
        )[:5]

        po_gap = no_po_items
        yard_gap = no_yard_items
        executive = {
            "po_pct": _supply_ratio(po_items, total_items),
            "po_pct_css": _supply_ratio_css(po_items, total_items),
            "no_po_pct": _supply_ratio(no_po_items, total_items),
            "no_po_pct_css": _supply_ratio_css(no_po_items, total_items),
            "yard_pct": _supply_ratio(yard_items, total_items),
            "yard_pct_css": _supply_ratio_css(yard_items, total_items),
            "yard_of_po_pct": _supply_ratio(yard_items, po_items),
            "yard_of_po_pct_css": _supply_ratio_css(yard_items, po_items),
            "po_gap": po_gap,
            "yard_gap": yard_gap,
            "risk_items": po_gap + yard_gap,
            "risk_pct": _supply_ratio(po_gap + yard_gap, total_items + po_items),
            "critical_drawings": critical_drawings,
            "critical_pct": _supply_ratio(critical_drawings, total_drawings),
            "clean_drawings": clean_drawings,
            "clean_pct": _supply_ratio(clean_drawings, total_drawings),
            "campaign_rows": campaign_exec_rows,
            "bottleneck_rows": bottleneck_rows,
            "pending_groups": pending_groups,
            "funnel": [
                {
                    "key": "total",
                    "label": "Scope",
                    "value": total_items,
                    "width_css": "100.00",
                    "pct": 100.0 if total_items else 0.0,
                },
                {
                    "key": "po",
                    "label": "With PO",
                    "value": po_items,
                    "width_css": _supply_ratio_css(po_items, total_items),
                    "pct": _supply_ratio(po_items, total_items),
                },
                {
                    "key": "no_po",
                    "label": "Without PO",
                    "value": no_po_items,
                    "width_css": _supply_ratio_css(no_po_items, total_items),
                    "pct": _supply_ratio(no_po_items, total_items),
                },
                {
                    "key": "no_yard",
                    "label": "Item not arrived",
                    "value": no_yard_items,
                    "width_css": _supply_ratio_css(no_yard_items, total_items),
                    "pct": _supply_ratio(no_yard_items, total_items),
                },
                {
                    "key": "yard",
                    "label": "At Yard",
                    "value": yard_items,
                    "width_css": _supply_ratio_css(yard_items, total_items),
                    "pct": _supply_ratio(yard_items, total_items),
                },
            ],
        }

        chart_campaigns = [
            {
                "key": row["key"],
                "label": row["label"],
                "color": row["color"],
                "accent": row["accent"],
                "total": int(row["total"] or 0),
                "po": int(row["po"] or 0),
                "yard": int(row["yard"] or 0),
                "no_po": int(row["no_po"] or 0),
                "no_yard": int(row["no_yard"] or 0),
                "finalized": int(row["finalized"] or 0),
                "po_pct": float(row["po_pct"] or 0),
                "yard_pct": float(row["yard_pct"] or 0),
            }
            for row in campaign_exec_rows
        ]
        chart_pending = [
            {
                "pending": row["pending"],
                "label": row["label"],
                "total": int(row["value"] or 0),
                "at_yard": int(row.get("at_yard") or 0),
                "segments": [
                    {
                        "key": seg["key"],
                        "label": seg["label"],
                        "value": int(seg["value"] or 0),
                    }
                    for seg in row["segments"]
                ],
            }
            for row in pending_rows
        ]
        chart_payload = {
            "scope": scope_key,
            "label": scope_label,
            "totals": {
                "total": total_items,
                "po": po_items,
                "no_po": no_po_items,
                "no_yard": no_yard_items,
                "finalized": finalized_drawings,
                "yard": yard_items,
                "drawings": total_drawings,
            },
            "campaigns": chart_campaigns,
            "pending_rows": chart_pending,
        }

        views.append({
            "key": scope_key,
            "label": scope_label,
            "campaigns": campaigns,
            "stages": stages,
            "executive": executive,
            "pending_rows": pending_rows,
            "totals": {
                "total": total_items,
                "po": po_items,
                "yard": yard_items,
                "drawings": len({item.get("document_id") for item in scoped_items if item.get("scope") == scope_key}),
            },
            "chart_payload_json": json.dumps(chart_payload, ensure_ascii=False),
        })
    return views


def _supply_filters_hash(filters: dict[str, Any]) -> str:
    return hash_supply_snapshot_filters(filters)


def _rows(cursor, query: str, params: tuple = ()) -> list[dict[str, Any]]:
    return [dict(row) for row in cursor.execute(query, params).fetchall()]


def _scalar(cursor, query: str, params: tuple = (), default: Any = 0) -> Any:
    row = cursor.execute(query, params).fetchone()
    if row is None:
        return default
    value = next(iter(row.values()), default) if isinstance(row, dict) else row[0]
    return default if value is None else value


def _pct(part: int | float, total: int | float) -> int:
    return round(100 * float(part) / float(total)) if total else 0


def _parse_date(value: str | None, default: date) -> date:
    text = (value or "").strip()
    if not text:
        return default
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return default


def _fmt_date(value: Any) -> str:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value or "")


def _cumulative_from_rows(labels: list[str], rows: list[dict[str, Any]], value_key: str) -> list[float]:
    by_label = {str(row["label"]): float(row.get(value_key) or 0) for row in rows}
    total = 0.0
    values = []
    for label in labels:
        total += by_label.get(label, 0.0)
        values.append(round(total, 2))
    return values


def _series_labels(*rowsets: list[dict[str, Any]]) -> list[str]:
    labels = {str(row["label"]) for rows in rowsets for row in rows}
    return sorted(labels)


def _p6_clean_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\ufeff", "").split())


def _p6_float(value: Any) -> float:
    text = str(value or "").strip().replace("%", "").replace("'", "").replace("\xa0", "")
    if not text:
        return 0.0
    text = text.replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _p6_ratio(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    parsed = _p6_float(value)
    if isinstance(value, str) and "%" in value:
        parsed = parsed / 100
    elif abs(parsed) > 1.5:
        parsed = parsed / 100
    return max(0.0, min(1.0, parsed))


def _p6_percent(value: Any) -> float:
    return max(0.0, min(100.0, _p6_float(value)))


def _p6_parse_date(value: Any) -> date | None:
    text = str(value or "").strip().lower()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    parts = text.replace("-", "/").split("/")
    if len(parts) != 3:
        return None
    day_text, month_text, year_text = parts
    try:
        day = int(day_text)
    except ValueError:
        return None
    month = P6_MONTHS.get(month_text[:3], P6_MONTHS.get(month_text))
    if not month:
        return None
    try:
        year = int(year_text)
    except ValueError:
        return None
    if year < 100:
        year += 2000 if year < 70 else 1900
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _p6_date_label(value: date | None) -> str:
    return value.strftime("%d/%m/%y") if value else "-"


def _p6_number_label(value: Any) -> str:
    return f"{float(value or 0):,.0f}".replace(",", ".")


def _p6_percent_label(value: Any) -> str:
    return f"{_p6_ratio(value) * 100:.2f}%"


def _p6_percent_css(value: Any) -> str:
    return f"{_p6_ratio(value) * 100:.2f}"


def _p6_month_floor(value: date) -> date:
    return date(value.year, value.month, 1)


def _p6_add_month(value: date) -> date:
    if value.month == 12:
        return date(value.year + 1, 1, 1)
    return date(value.year, value.month + 1, 1)


def _p6_month_last(value: date) -> date:
    next_month = _p6_add_month(_p6_month_floor(value))
    return date.fromordinal(next_month.toordinal() - 1)


def _p6_month_entries(start: date, finish: date) -> list[dict[str, Any]]:
    if start > finish:
        start, finish = finish, start
    current = _p6_month_floor(start)
    end = _p6_month_floor(finish)
    entries = []
    while current <= end:
        entries.append({
            "iso": current.isoformat(),
            "label": f"{P6_MONTH_LABELS[current.month]}/{str(current.year)[2:]}",
        })
        current = _p6_add_month(current)
    return entries


def _p6_month_segments(start: date, finish: date) -> list[tuple[str, int]]:
    if start > finish:
        start, finish = finish, start
    current = _p6_month_floor(start)
    end = _p6_month_floor(finish)
    segments = []
    while current <= end:
        month_start = current
        month_finish = _p6_month_last(current)
        segment_start = max(start, month_start)
        segment_finish = min(finish, month_finish)
        days = max((segment_finish - segment_start).days + 1, 0)
        if days:
            segments.append((month_start.isoformat(), days))
        current = _p6_add_month(current)
    return segments


def _p6_normalize_label(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _p6_workbook_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return _p6_parse_date(value)


def _p6_find_sheet_row(ws: Any, wanted: str) -> int | None:
    wanted_norm = _p6_normalize_label(wanted)
    for row_idx in range(1, ws.max_row + 1):
        if _p6_normalize_label(ws.cell(row_idx, 1).value) == wanted_norm:
            return row_idx
    return None


def _p6_curve_chart_from_workbook(ws: Any, *, today: date) -> dict[str, Any]:
    dates: list[tuple[int, date]] = []
    for col_idx in range(2, ws.max_column + 1):
        current = _p6_workbook_date(ws.cell(4, col_idx).value)
        if current:
            dates.append((col_idx, current))
    planned_row = _p6_find_sheet_row(ws, "PROJECT ACCO(%)") or _p6_find_sheet_row(ws, "PROJECT ACCO BL (%)")
    actual_row = _p6_find_sheet_row(ws, "PROJECT ACCO Real (%)")
    if not dates or not planned_row or not actual_row:
        return _empty_chart("Sem curva fisica no XLSX")

    labels = [item[1].isoformat() for item in dates]
    planned_values = []
    actual_values = []
    for col_idx, current in dates:
        planned_values.append(round(_p6_ratio(ws.cell(planned_row, col_idx).value) * 100, 2))
        if current <= today:
            actual_values.append(round(_p6_ratio(ws.cell(actual_row, col_idx).value) * 100, 2))
        else:
            actual_values.append(None)

    return {
        "data": [
            {
                "type": "scatter",
                "mode": "lines",
                "name": "Previsto",
                "x": labels,
                "y": planned_values,
                "line": {"color": "#0a0a0a", "width": 2.6},
                "hovertemplate": "%{x|%d/%m/%Y}<br>Previsto: %{y:.2f}%<extra></extra>",
            },
            {
                "type": "scatter",
                "mode": "lines+markers",
                "name": "Avanco",
                "x": labels,
                "y": actual_values,
                "connectgaps": False,
                "line": {"color": "#15803d", "width": 2.6},
                "marker": {"size": 4},
                "hovertemplate": "%{x|%d/%m/%Y}<br>Avanco: %{y:.2f}%<extra></extra>",
            },
        ],
        "layout": {
            "height": 330,
            "margin": {"l": 52, "r": 18, "t": 8, "b": 42},
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "font": {"family": "Inter, sans-serif", "size": 11, "color": "#525252"},
            "xaxis": {"type": "date", "gridcolor": "#ededed", "tickformat": "%b/%y"},
            "yaxis": {
                "ticksuffix": "%",
                "range": [0, 105],
                "gridcolor": "#ededed",
                "zerolinecolor": "#d4d4d4",
            },
            "legend": {"orientation": "h", "y": -0.18},
            "shapes": [{
                "type": "line",
                "xref": "x",
                "yref": "paper",
                "x0": today.isoformat(),
                "x1": today.isoformat(),
                "y0": 0,
                "y1": 1,
                "line": {"color": "#b91c1c", "width": 1, "dash": "dot"},
            }],
            "annotations": [{
                "text": "Hoje",
                "xref": "x",
                "yref": "paper",
                "x": today.isoformat(),
                "y": 1,
                "showarrow": False,
                "font": {"size": 10, "color": "#b91c1c"},
                "xanchor": "left",
            }],
        },
    }


def _p6_curve_chart_from_points(points: list[Any], *, today: date) -> dict[str, Any]:
    if not points:
        return _empty_chart("Sem curva fisica na base importada")

    labels = [point.period.isoformat() for point in points]
    planned_values = [round(float(point.planned_pct or 0), 2) for point in points]
    actual_values = [
        round(float(point.actual_pct or 0), 2)
        if point.period <= today and point.actual_pct is not None
        else None
        for point in points
    ]

    return {
        "data": [
            {
                "type": "scatter",
                "mode": "lines",
                "name": "Previsto",
                "x": labels,
                "y": planned_values,
                "line": {"color": "#0a0a0a", "width": 2.6},
                "hovertemplate": "%{x|%d/%m/%Y}<br>Previsto: %{y:.2f}%<extra></extra>",
            },
            {
                "type": "scatter",
                "mode": "lines+markers",
                "name": "Avanco",
                "x": labels,
                "y": actual_values,
                "connectgaps": False,
                "line": {"color": "#15803d", "width": 2.6},
                "marker": {"size": 4},
                "hovertemplate": "%{x|%d/%m/%Y}<br>Avanco: %{y:.2f}%<extra></extra>",
            },
        ],
        "layout": {
            "height": 330,
            "margin": {"l": 52, "r": 18, "t": 8, "b": 42},
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "font": {"family": "Inter, sans-serif", "size": 11, "color": "#525252"},
            "xaxis": {"type": "date", "gridcolor": "#ededed", "tickformat": "%b/%y"},
            "yaxis": {
                "ticksuffix": "%",
                "range": [0, 105],
                "gridcolor": "#ededed",
                "zerolinecolor": "#d4d4d4",
            },
            "legend": {"orientation": "h", "y": -0.18},
            "shapes": [{
                "type": "line",
                "xref": "x",
                "yref": "paper",
                "x0": today.isoformat(),
                "x1": today.isoformat(),
                "y0": 0,
                "y1": 1,
                "line": {"color": "#b91c1c", "width": 1, "dash": "dot"},
            }],
            "annotations": [{
                "text": "Hoje",
                "xref": "x",
                "yref": "paper",
                "x": today.isoformat(),
                "y": 1,
                "showarrow": False,
                "font": {"size": 10, "color": "#b91c1c"},
                "xanchor": "left",
            }],
        },
    }


def _p6_curves_from_database(*, today: date) -> dict[str, Any]:
    try:
        latest = (
            P6CurveImport.objects
            .filter(is_active=True)
            .order_by("-created_at", "-id")
            .first()
        )
    except (OperationalError, ProgrammingError):
        return {}
    if not latest:
        return {}

    cache_token = hashlib.sha1(
        f"db:{latest.pk}:{latest.updated_at.isoformat()}:{today.isoformat()}".encode("utf-8")
    ).hexdigest()
    cache_key = f"p6-curves-db:{cache_token}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    progress_rows = list(latest.progress_rows.order_by("row_number"))
    curve_points = list(latest.curve_points.order_by("sequence", "period"))

    total_snapshot: dict[str, Any] = {}
    executive_rows: list[dict[str, Any]] = []
    for row in progress_rows:
        planned = float(row.planned_pct or 0)
        actual = float(row.actual_pct or 0)
        baseline = float(row.baseline_pct or 0)
        weight = float(row.weight_pct or 0)
        if row.level == 0 and not total_snapshot:
            total_snapshot = {
                "name": row.name,
                "planned_pct": planned,
                "actual_pct": actual,
                "baseline_pct": baseline,
                "delta_pct": actual - planned,
            }
            continue
        if row.level != 1:
            continue
        delta = actual - planned
        executive_rows.append({
            "level": int(row.level),
            "label": row.name,
            "weight_pct": weight,
            "weight_label": _p6_percent_label(weight),
            "baseline_pct": baseline,
            "baseline_label": _p6_percent_label(baseline),
            "planned_pct": planned,
            "planned_label": _p6_percent_label(planned),
            "planned_pct_css": _p6_percent_css(planned),
            "actual_pct": actual,
            "actual_label": _p6_percent_label(actual),
            "actual_pct_css": _p6_percent_css(actual),
            "delta_pct": delta,
            "delta_label": f"{delta:+.2f} p.p.",
            "status": "ahead" if delta >= 0 else "late",
        })

    result = {
        "curves_source_path": "Base importada",
        "curves_file_name": latest.original_filename,
        "curves_sheet": latest.progress_sheet,
        "curves_imported_at": latest.created_at,
        "curves_import_id": latest.pk,
        "progress_row_count": latest.progress_row_count,
        "curve_point_count": latest.curve_point_count,
        "executive_row_count": latest.executive_row_count,
        "executive_rows": executive_rows,
        "total_snapshot": total_snapshot,
        "charts": {"physical_curve": _p6_curve_chart_from_points(curve_points, today=today)},
    }
    try:
        result["management_snapshot"] = latest.management_snapshot.payload
    except Exception:
        result["management_snapshot"] = {}
    cache.set(cache_key, result, 300)
    return result


def _p6_curves_from_workbook(path: Path, *, today: date) -> dict[str, Any]:
    if not path.exists():
        return {}
    cache_token = hashlib.sha1(f"{path}:{path.stat().st_mtime_ns}:{today.isoformat()}".encode("utf-8")).hexdigest()
    cache_key = f"p6-curves:{cache_token}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    progress_sheet = workbook["Progress PMS (%)"] if "Progress PMS (%)" in workbook.sheetnames else None
    curve_sheet = workbook["DB (%)"] if "DB (%)" in workbook.sheetnames else None

    total_snapshot: dict[str, Any] = {}
    executive_rows: list[dict[str, Any]] = []
    if progress_sheet:
        for row_idx in range(1, progress_sheet.max_row + 1):
            level = progress_sheet.cell(row_idx, 3).value
            name = _p6_clean_text(progress_sheet.cell(row_idx, 4).value)
            if not name:
                continue
            planned = _p6_ratio(progress_sheet.cell(row_idx, 7).value)
            actual = _p6_ratio(progress_sheet.cell(row_idx, 8).value)
            baseline = _p6_ratio(progress_sheet.cell(row_idx, 6).value)
            weight = _p6_ratio(progress_sheet.cell(row_idx, 5).value)
            if level == 0:
                total_snapshot = {
                    "name": name,
                    "planned_pct": planned * 100,
                    "actual_pct": actual * 100,
                    "baseline_pct": baseline * 100,
                    "delta_pct": (actual - planned) * 100,
                }
                continue
            if level != 1:
                continue
            delta = (actual - planned) * 100
            executive_rows.append({
                "level": int(level),
                "label": name,
                "weight_pct": weight * 100,
                "weight_label": _p6_percent_label(progress_sheet.cell(row_idx, 5).value),
                "baseline_pct": baseline * 100,
                "baseline_label": _p6_percent_label(progress_sheet.cell(row_idx, 6).value),
                "planned_pct": planned * 100,
                "planned_label": _p6_percent_label(progress_sheet.cell(row_idx, 7).value),
                "planned_pct_css": _p6_percent_css(progress_sheet.cell(row_idx, 7).value),
                "actual_pct": actual * 100,
                "actual_label": _p6_percent_label(progress_sheet.cell(row_idx, 8).value),
                "actual_pct_css": _p6_percent_css(progress_sheet.cell(row_idx, 8).value),
                "delta_pct": delta,
                "delta_label": f"{delta:+.2f} p.p.",
                "status": "ahead" if delta >= 0 else "late",
            })

    chart = _p6_curve_chart_from_workbook(curve_sheet, today=today) if curve_sheet else _empty_chart("Sem curva fisica no XLSX")
    result = {
        "curves_source_path": str(path),
        "curves_file_name": path.name,
        "curves_sheet": "Progress PMS (%)",
        "executive_rows": executive_rows,
        "total_snapshot": total_snapshot,
        "charts": {"physical_curve": chart},
    }
    cache.set(cache_key, result, 300)
    return result


def _p6_physical_curve_chart(
    month_entries: list[dict[str, Any]],
    planned_by_month: dict[str, float],
    actual_by_month: dict[str, float],
    total_budget: float,
    today: date,
    min_start: date,
    max_finish: date,
) -> dict[str, Any]:
    if not month_entries or total_budget <= 0:
        return _empty_chart("Sem curva fisica P6")

    labels = [row["iso"] for row in month_entries]
    current_month = _p6_month_floor(today).isoformat()
    planned_values = []
    actual_values = []
    planned_units = 0.0
    actual_units = 0.0
    for label in labels:
        planned_units += planned_by_month.get(label, 0.0)
        planned_values.append(round(100 * planned_units / total_budget, 2))
        if label <= current_month:
            actual_units += actual_by_month.get(label, 0.0)
            actual_values.append(round(100 * actual_units / total_budget, 2))
        else:
            actual_values.append(None)

    shapes = []
    annotations = []
    if min_start <= today <= max_finish:
        shapes.append({
            "type": "line",
            "xref": "x",
            "yref": "paper",
            "x0": today.isoformat(),
            "x1": today.isoformat(),
            "y0": 0,
            "y1": 1,
            "line": {"color": "#b91c1c", "width": 1, "dash": "dot"},
        })
        annotations.append({
            "text": "Hoje",
            "xref": "x",
            "yref": "paper",
            "x": today.isoformat(),
            "y": 1,
            "showarrow": False,
            "font": {"size": 10, "color": "#b91c1c"},
            "xanchor": "left",
        })

    return {
        "data": [
            {
                "type": "scatter",
                "mode": "lines",
                "name": "Planejado fisico",
                "x": labels,
                "y": planned_values,
                "line": {"color": "#0a0a0a", "width": 2.6},
                "hovertemplate": "%{x|%b/%Y}<br>Planejado: %{y:.2f}%<extra></extra>",
            },
            {
                "type": "scatter",
                "mode": "lines+markers",
                "name": "Real / earned",
                "x": labels,
                "y": actual_values,
                "connectgaps": False,
                "line": {"color": "#15803d", "width": 2.6},
                "marker": {"size": 4},
                "hovertemplate": "%{x|%b/%Y}<br>Real: %{y:.2f}%<extra></extra>",
            },
        ],
        "layout": {
            "height": 330,
            "margin": {"l": 52, "r": 18, "t": 8, "b": 42},
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "font": {"family": "Inter, sans-serif", "size": 11, "color": "#525252"},
            "xaxis": {"type": "date", "gridcolor": "#ededed", "tickformat": "%b/%y"},
            "yaxis": {
                "ticksuffix": "%",
                "range": [0, 105],
                "gridcolor": "#ededed",
                "zerolinecolor": "#d4d4d4",
            },
            "legend": {"orientation": "h", "y": -0.18},
            "shapes": shapes,
            "annotations": annotations,
        },
    }


def _p6_area_performance_chart(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return _empty_chart("Sem areas executivas P6")
    chart_rows = sorted(rows, key=lambda item: float(item.get("delta_pct") or 0))
    labels = [str(row.get("label") or "-") for row in chart_rows]
    planned = [round(float(row.get("planned_pct") or 0), 2) for row in chart_rows]
    actual = [round(float(row.get("actual_pct") or 0), 2) for row in chart_rows]
    delta = [round(float(row.get("delta_pct") or 0), 2) for row in chart_rows]
    return {
        "data": [
            {
                "type": "bar",
                "orientation": "h",
                "name": "Previsto",
                "x": planned,
                "y": labels,
                "marker": {"color": "#d4d4d8"},
                "hovertemplate": "%{y}<br>Previsto: %{x:.2f}%<extra></extra>",
            },
            {
                "type": "bar",
                "orientation": "h",
                "name": "Avanco",
                "x": actual,
                "y": labels,
                "customdata": delta,
                "marker": {"color": "#dc2626"},
                "hovertemplate": "%{y}<br>Avanco: %{x:.2f}%<br>Delta: %{customdata:+.2f} p.p.<extra></extra>",
            },
        ],
        "layout": {
            "height": 300,
            "margin": {"l": 120, "r": 18, "t": 8, "b": 36},
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "font": {"family": "Inter, sans-serif", "size": 11, "color": "#525252"},
            "barmode": "group",
            "xaxis": {"ticksuffix": "%", "range": [0, 105], "gridcolor": "#ededed", "zeroline": False},
            "yaxis": {"automargin": True},
            "legend": {"orientation": "h", "y": -0.16},
        },
    }


def _p6_monthly_units_chart(month_entries: list[dict[str, Any]], planned_by_month: dict[str, float], actual_by_month: dict[str, float]) -> dict[str, Any]:
    if not month_entries:
        return _empty_chart("Sem distribuicao mensal P6")
    labels = [row["label"] for row in month_entries]
    planned = [round(planned_by_month.get(row["iso"], 0.0), 0) for row in month_entries]
    actual = [round(actual_by_month.get(row["iso"], 0.0), 0) for row in month_entries]
    return {
        "data": [
            {
                "type": "bar",
                "name": "Planejado mensal",
                "x": labels,
                "y": planned,
                "marker": {"color": "#cbd5e1"},
                "hovertemplate": "%{x}<br>Planejado: %{y:,.0f}<extra></extra>",
            },
            {
                "type": "scatter",
                "mode": "lines+markers",
                "name": "Avanco distribuido",
                "x": labels,
                "y": actual,
                "line": {"color": "#dc2626", "width": 2.2},
                "marker": {"size": 4, "color": "#dc2626"},
                "hovertemplate": "%{x}<br>Avanco: %{y:,.0f}<extra></extra>",
            },
        ],
        "layout": {
            "height": 300,
            "margin": {"l": 54, "r": 18, "t": 8, "b": 42},
            "paper_bgcolor": "transparent",
            "plot_bgcolor": "transparent",
            "font": {"family": "Inter, sans-serif", "size": 11, "color": "#525252"},
            "xaxis": {"gridcolor": "#ededed", "tickangle": 0},
            "yaxis": {"gridcolor": "#ededed", "zerolinecolor": "#d4d4d4"},
            "legend": {"orientation": "h", "y": -0.18},
        },
    }


def _p6_dashboard_from_curve_snapshot(curves: dict[str, Any], *, today: date) -> dict[str, Any]:
    management_snapshot = curves.get("management_snapshot") or {}
    snapshot_kpis = management_snapshot.get("management_kpis") or {}
    total_snapshot = curves.get("total_snapshot") or {}
    executive_rows = curves.get("executive_rows") or []

    physical_pct = round(float(snapshot_kpis.get("physical_pct") or total_snapshot.get("actual_pct") or 0), 2)
    planned_today_pct = round(float(snapshot_kpis.get("planned_today_pct") or total_snapshot.get("planned_pct") or 0), 2)
    remaining_units = float(snapshot_kpis.get("remaining_units") or 0)
    total_budget = float(snapshot_kpis.get("total_budget_units") or 0)
    if total_budget <= 0 and remaining_units and physical_pct < 100:
        total_budget = remaining_units / max(1 - (physical_pct / 100), 0.0001)
    actual_total = float(snapshot_kpis.get("actual_units") or 0)
    if actual_total <= 0 and total_budget:
        actual_total = total_budget * physical_pct / 100

    area_rows = management_snapshot.get("area_rows") or []
    if not area_rows:
        for row in executive_rows:
            delta = float(row.get("delta_pct") or 0)
            area_rows.append({
                "label": row.get("label", "-"),
                "weight_label": row.get("weight_label", "0.00%"),
                "planned_pct": float(row.get("planned_pct") or 0),
                "planned_pct_css": row.get("planned_pct_css", "0.00"),
                "planned_label": row.get("planned_label", "0.00%"),
                "actual_pct": float(row.get("actual_pct") or 0),
                "actual_label": row.get("actual_label", "0.00%"),
                "actual_pct_css": row.get("actual_pct_css", "0.00"),
                "delta_label": row.get("delta_label", "+0.00 p.p."),
                "delta_pct": delta,
                "delta_abs_css": f"{min(abs(delta) * 8, 100):.2f}",
                "status": row.get("status", "ahead"),
                "finish_label": "-",
                "budget_label": "-",
                "remaining_label": "-",
            })

    timeline_payload = management_snapshot.get("timeline") or {}
    month_entries = timeline_payload.get("months") or management_snapshot.get("monthly_rows") or []
    monthly_planned = {
        str(key): float(value or 0)
        for key, value in (management_snapshot.get("monthly_planned") or {}).items()
    }
    monthly_actual = {
        str(key): float(value or 0)
        for key, value in (management_snapshot.get("monthly_actual") or {}).items()
    }
    if not timeline_payload:
        timeline_payload = {
            "start": "",
            "finish": "",
            "start_label": "-",
            "finish_label": "-",
            "months": month_entries,
            "month_count": len(month_entries),
        }
    else:
        timeline_payload = {
            **timeline_payload,
            "months": month_entries,
            "month_count": timeline_payload.get("month_count") or len(month_entries),
        }

    consult_tree = management_snapshot.get("consult_tree") or []
    level_counts = Counter(int(row.get("level") or 0) for row in consult_tree)
    total_rows = int(curves.get("progress_row_count") or len(consult_tree))
    curve_point_count = int(curves.get("curve_point_count") or 0)
    level_summary = [
        {
            "level": level,
            "count": count,
            "pct": round(100 * count / total_rows, 2) if total_rows else 0,
            "pct_css": f"{(round(100 * count / total_rows, 2) if total_rows else 0):.2f}",
        }
        for level, count in sorted(level_counts.items())
    ]
    late_area_rows = [row for row in area_rows if float(row.get("delta_pct") or 0) < 0]
    worst_area = min(area_rows, key=lambda row: float(row.get("delta_pct") or 0), default={})
    management_kpis = {
        "late_areas": len(late_area_rows),
        "remaining_units": remaining_units,
        "remaining_units_label": _p6_number_label(remaining_units),
        "remaining_pct": round(max(100 - physical_pct, 0), 2),
        "active_packages": 0,
        "next_90_packages": 0,
        "worst_area_label": worst_area.get("label", "-"),
        "worst_area_delta": worst_area.get("delta_label", "-"),
    }
    management_kpis.update(snapshot_kpis)

    return {
        "source_path": curves.get("curves_source_path") or "Base importada",
        "file_name": curves.get("curves_file_name") or "Annex III importado",
        "curves_source_path": curves.get("curves_source_path", ""),
        "curves_file_name": curves.get("curves_file_name", ""),
        "curves_imported_at": curves.get("curves_imported_at"),
        "curves_import_id": curves.get("curves_import_id"),
        "rows": [],
        "render_rows": [],
        "level_summary": level_summary,
        "executive_level_summary": [row for row in level_summary if int(row["level"]) <= 2],
        "executive_rows": executive_rows,
        "management": {
            "areas": area_rows,
            "critical_packages": [],
            "timeline_rows": [],
            "consult_tree": consult_tree,
            "source": management_snapshot.get("source", "Annex III XLSX"),
            "source_sheet": management_snapshot.get("source_sheet", ""),
            "kpis": management_kpis,
        },
        "timeline": timeline_payload,
        "kpis": {
            "total_rows": total_rows,
            "activities": total_rows,
            "wbs": total_rows,
            "curve_points": curve_point_count,
            "level_2_rows": int(level_counts.get(2, 0)),
            "executive_rows": len(executive_rows) or len(area_rows),
            "budget_units": total_budget,
            "actual_units": actual_total,
            "physical_pct": physical_pct,
            "planned_today_pct": planned_today_pct,
            "delta_pct": round(physical_pct - planned_today_pct, 2),
            "start": timeline_payload.get("start", ""),
            "finish": timeline_payload.get("finish", ""),
            "start_label": timeline_payload.get("start_label", "-"),
            "finish_label": timeline_payload.get("finish_label", "-"),
            "duration_days": int(timeline_payload.get("month_count") or len(month_entries) or 0) * 30,
        },
        "charts": {
            "physical_curve": curves.get("charts", {}).get("physical_curve") or _empty_chart("Sem curva fisica P6"),
            "area_performance": _p6_area_performance_chart(area_rows),
            "monthly_units": _p6_monthly_units_chart(month_entries, monthly_planned, monthly_actual),
        },
    }


def _p6_dashboard(filters: dict) -> dict:
    today = date.today()
    curves = _p6_curves_from_database(today=today)
    if not curves:
        raise FileNotFoundError("Nenhum import P6 ativo na base do sistema. Importe o Annex III XLSX.")
    return _p6_dashboard_from_curve_snapshot(curves, today=today)

    raise RuntimeError("Caminho legado de arquivo local P6 desativado. Use a base do sistema.")

    nodes: list[dict[str, Any]] = []
    stack: list[dict[str, Any]] = []
    for index, raw in enumerate(raw_rows):
        try:
            level = int(_p6_float(raw.get("Level")))
        except (TypeError, ValueError):
            level = 0
        while stack and int(stack[-1]["level"]) >= level:
            stack.pop()
        parent_id = stack[-1]["node_id"] if stack else ""
        node_id = f"p6-{index}"

        activity_id = _p6_clean_text(raw.get("Activity ID"))
        activity_name = _p6_clean_text(raw.get("Activity Name"))
        display_name = activity_name or activity_id or "Sem nome"
        display_id = activity_id if activity_name else ""
        start = _p6_parse_date(raw.get("Start"))
        finish = _p6_parse_date(raw.get("Finish"))
        bar_start = start or finish
        bar_finish = finish or start
        if bar_start and bar_finish and bar_start > bar_finish:
            bar_start, bar_finish = bar_finish, bar_start

        budget_nonlabor = _p6_float(raw.get("Budgeted Nonlabor Units"))
        pct_complete = _p6_percent(raw.get("Nonlabor Units % Complete"))
        actual_nonlabor = _p6_float(raw.get("Actual Nonlabor Units"))
        if actual_nonlabor <= 0 and budget_nonlabor > 0 and pct_complete > 0:
            actual_nonlabor = budget_nonlabor * pct_complete / 100

        node = {
            "node_id": node_id,
            "parent_id": parent_id,
            "level": level,
            "activity_id": activity_id,
            "display_id": display_id,
            "display_name": display_name,
            "budget_labor": _p6_float(raw.get("Budgeted Labor Units")),
            "budget_nonlabor": budget_nonlabor,
            "actual_nonlabor": actual_nonlabor,
            "pct_complete": pct_complete,
            "pct_complete_css": f"{pct_complete:.2f}",
            "start": start,
            "finish": finish,
            "bar_start": bar_start,
            "bar_finish": bar_finish,
            "start_label": _p6_date_label(start),
            "finish_label": _p6_date_label(finish),
            "duration": int(_p6_float(raw.get("Original Duration"))),
        }
        nodes.append(node)
        stack.append({"node_id": node_id, "level": level})

    child_counts = Counter(node["parent_id"] for node in nodes if node["parent_id"])
    for node in nodes:
        node["child_count"] = int(child_counts.get(node["node_id"], 0))
        node["has_children"] = node["child_count"] > 0
        node["is_summary"] = node["has_children"]
        node["initial_expanded"] = node["has_children"] and node["level"] < 3
        node["initial_visible"] = node["level"] <= 3

    min_start = min((node["bar_start"] for node in nodes if node.get("bar_start")), default=date.today())
    max_finish = max((node["bar_finish"] for node in nodes if node.get("bar_finish")), default=min_start)
    if max_finish < min_start:
        max_finish = min_start
    total_days = max((max_finish - min_start).days + 1, 1)
    month_entries = _p6_month_entries(min_start, max_finish)

    for node in nodes:
        bar_start = node.get("bar_start")
        bar_finish = node.get("bar_finish")
        has_bar = bool(bar_start and bar_finish)
        node["has_bar"] = has_bar
        if not has_bar:
            node["bar_left"] = "0.0000"
            node["bar_width"] = "0.0000"
            node["is_milestone"] = False
            continue
        duration_days = max((bar_finish - bar_start).days + 1, 1)
        node["bar_left"] = f"{100 * (bar_start - min_start).days / total_days:.4f}"
        node["bar_width"] = f"{max(100 * duration_days / total_days, 0.22):.4f}"
        node["is_milestone"] = duration_days <= 1 or node.get("duration") == 0

    leaves = [node for node in nodes if not node.get("has_children")]
    curve_activities = [
        node for node in leaves
        if node.get("budget_nonlabor", 0) > 0 and node.get("bar_start") and node.get("bar_finish")
    ]
    root = nodes[0] if nodes else {}
    total_budget = float(root.get("budget_nonlabor") or 0) or sum(float(node.get("budget_nonlabor") or 0) for node in curve_activities)
    actual_total = float(root.get("actual_nonlabor") or 0) or sum(float(node.get("actual_nonlabor") or 0) for node in curve_activities)
    physical_pct = _p6_percent(root.get("pct_complete")) if root else (100 * actual_total / total_budget if total_budget else 0)

    planned_by_month = {row["iso"]: 0.0 for row in month_entries}
    actual_by_month = {row["iso"]: 0.0 for row in month_entries}
    planned_until_today = 0.0
    for node in curve_activities:
        start = node["bar_start"]
        finish = node["bar_finish"]
        budget = float(node.get("budget_nonlabor") or 0)
        actual = float(node.get("actual_nonlabor") or 0)
        duration_days = max((finish - start).days + 1, 1)

        for month_key, days in _p6_month_segments(start, finish):
            planned_by_month[month_key] = planned_by_month.get(month_key, 0.0) + budget * days / duration_days

        if today >= start:
            elapsed_finish = min(today, finish)
            elapsed_days = max((elapsed_finish - start).days + 1, 0)
            planned_until_today += budget * min(elapsed_days / duration_days, 1)

        if actual > 0:
            earned_finish = min(max(today, start), finish)
            earned_days = max((earned_finish - start).days + 1, 1)
            for month_key, days in _p6_month_segments(start, earned_finish):
                actual_by_month[month_key] = actual_by_month.get(month_key, 0.0) + actual * days / earned_days

    actual_curve_total = sum(actual_by_month.values())
    if actual_total > 0 and actual_curve_total > 0:
        scale = actual_total / actual_curve_total
        actual_by_month = {key: value * scale for key, value in actual_by_month.items()}

    planned_today_pct = round(100 * planned_until_today / total_budget, 2) if total_budget else 0.0
    total_snapshot = curves.get("total_snapshot") or {}
    if total_snapshot:
        physical_pct = round(float(total_snapshot.get("actual_pct") or 0), 2)
        planned_today_pct = round(float(total_snapshot.get("planned_pct") or 0), 2)
        actual_total = round(total_budget * physical_pct / 100, 2) if total_budget else actual_total
    physical_curve_chart = curves.get("charts", {}).get("physical_curve") or _p6_physical_curve_chart(
        month_entries,
        planned_by_month,
        actual_by_month,
        total_budget,
        today,
        min_start,
        max_finish,
    )
    executive_rows = curves.get("executive_rows") or []
    management_snapshot = curves.get("management_snapshot") or {}

    level2_nodes = [node for node in nodes if int(node.get("level") or 0) == 2]
    wbs_nodes = [node for node in nodes if node.get("has_children")]
    active_packages = [
        node for node in leaves
        if node.get("bar_start") and node.get("bar_finish")
        and node["bar_start"] <= today <= node["bar_finish"]
        and float(node.get("pct_complete") or 0) < 99.9
    ]
    next_90_packages = [
        node for node in leaves
        if node.get("bar_finish")
        and 0 <= (node["bar_finish"] - today).days <= 90
        and float(node.get("pct_complete") or 0) < 99.9
    ]

    def _p6_match_level2(label: str) -> dict[str, Any]:
        normalized = _p6_normalize_label(label)
        for candidate in level2_nodes:
            name = _p6_normalize_label(candidate.get("display_name"))
            if name and (name in normalized or normalized in name):
                return candidate
        return {}

    area_rows = []
    if executive_rows:
        for row in executive_rows:
            matched = _p6_match_level2(str(row.get("label") or ""))
            budget = float(matched.get("budget_nonlabor") or 0)
            actual = float(matched.get("actual_nonlabor") or 0)
            remaining = max(budget - actual, 0.0)
            delta = float(row.get("delta_pct") or 0)
            area_rows.append({
                "label": row.get("label", "-"),
                "weight_label": row.get("weight_label", "0.00%"),
                "planned_pct": float(row.get("planned_pct") or 0),
                "planned_pct_css": row.get("planned_pct_css", "0.00"),
                "planned_label": row.get("planned_label", "0.00%"),
                "actual_pct": float(row.get("actual_pct") or 0),
                "actual_label": row.get("actual_label", "0.00%"),
                "actual_pct_css": row.get("actual_pct_css", "0.00"),
                "delta_label": row.get("delta_label", "+0.00 p.p."),
                "delta_pct": delta,
                "delta_abs_css": f"{min(abs(delta) * 8, 100):.2f}",
                "status": row.get("status", "ahead"),
                "finish_label": matched.get("finish_label", "-"),
                "budget_label": _p6_number_label(budget),
                "remaining_label": _p6_number_label(remaining),
            })
    else:
        for node in level2_nodes:
            budget = float(node.get("budget_nonlabor") or 0)
            actual = float(node.get("actual_nonlabor") or 0)
            remaining = max(budget - actual, 0.0)
            pct = float(node.get("pct_complete") or 0)
            area_rows.append({
                "label": node.get("display_name", "-"),
                "weight_label": "-",
                "planned_pct": pct,
                "planned_pct_css": f"{pct:.2f}",
                "planned_label": "-",
                "actual_pct": pct,
                "actual_label": f"{pct:.2f}%",
                "actual_pct_css": f"{pct:.2f}",
                "delta_label": "-",
                "delta_pct": 0,
                "delta_abs_css": "0.00",
                "status": "ahead",
                "finish_label": node.get("finish_label", "-"),
                "budget_label": _p6_number_label(budget),
                "remaining_label": _p6_number_label(remaining),
            })

    candidate_packages = [
        node for node in nodes
        if 3 <= int(node.get("level") or 0) <= 5
        and float(node.get("budget_nonlabor") or 0) > 0
    ]
    critical_rows = []
    for node in sorted(
        candidate_packages,
        key=lambda item: max(float(item.get("budget_nonlabor") or 0) - float(item.get("actual_nonlabor") or 0), 0),
        reverse=True,
    )[:8]:
        budget = float(node.get("budget_nonlabor") or 0)
        actual = float(node.get("actual_nonlabor") or 0)
        remaining = max(budget - actual, 0.0)
        finish = node.get("bar_finish")
        days_to_finish = (finish - today).days if finish else None
        if days_to_finish is None:
            timing = "-"
            timing_status = "neutral"
        elif days_to_finish < 0 and float(node.get("pct_complete") or 0) < 99.9:
            timing = f"{abs(days_to_finish)}d vencido"
            timing_status = "late"
        elif days_to_finish <= 90:
            timing = f"{days_to_finish}d para finish"
            timing_status = "watch"
        else:
            timing = f"{days_to_finish}d para finish"
            timing_status = "neutral"
        critical_rows.append({
            "label": node.get("display_name", "-"),
            "activity_id": node.get("display_id", "-"),
            "pct_label": f"{float(node.get('pct_complete') or 0):.2f}%",
            "pct_css": f"{float(node.get('pct_complete') or 0):.2f}",
            "remaining_label": _p6_number_label(remaining),
            "finish_label": node.get("finish_label", "-"),
            "timing": timing,
            "timing_status": timing_status,
        })

    timeline_rows = [
        {
            "label": node.get("display_name", "-"),
            "start_label": node.get("start_label", "-"),
            "finish_label": node.get("finish_label", "-"),
            "bar_left": node.get("bar_left", "0"),
            "bar_width": node.get("bar_width", "0"),
            "progress_css": node.get("pct_complete_css", "0"),
        }
        for node in level2_nodes
        if node.get("has_bar")
    ]
    consult_source_nodes = [node for node in nodes if int(node.get("level") or 0) <= 4]
    consult_seen = {node["node_id"] for node in consult_source_nodes}
    for row in critical_rows:
        activity_id = str(row.get("activity_id") or "")
        for node in nodes:
            if node["node_id"] in consult_seen:
                continue
            if activity_id and str(node.get("display_id") or "") == activity_id:
                consult_source_nodes.append(node)
                consult_seen.add(node["node_id"])
                break
    consult_tree = []
    consult_ids = {node["node_id"] for node in consult_source_nodes}
    for node in consult_source_nodes[:900]:
        finish = node.get("bar_finish")
        pct = float(node.get("pct_complete") or 0)
        if finish and finish < today and pct < 99.9:
            status = "late"
            status_label = "Atrasado"
        elif finish and 0 <= (finish - today).days <= 90 and pct < 99.9:
            status = "watch"
            status_label = "90 dias"
        elif pct >= 99.9:
            status = "done"
            status_label = "Concluido"
        else:
            status = "neutral"
            status_label = "Planejado"
        parent_id = node.get("parent_id") if node.get("parent_id") in consult_ids else ""
        consult_tree.append({
            "id": node.get("node_id"),
            "parent": parent_id,
            "level": int(node.get("level") or 0),
            "label": node.get("display_name") or "-",
            "code": node.get("display_id") or "-",
            "start": node.get("start_label") or "-",
            "finish": node.get("finish_label") or "-",
            "pct": f"{pct:.2f}%",
            "pct_css": f"{pct:.2f}",
            "status": status,
            "status_label": status_label,
        })
    late_area_rows = [row for row in area_rows if float(row.get("delta_pct") or 0) < 0]
    worst_area = min(area_rows, key=lambda row: float(row.get("delta_pct") or 0), default={})
    remaining_units = max(total_budget - actual_total, 0.0)

    if management_snapshot:
        area_rows = management_snapshot.get("area_rows") or area_rows
        consult_tree = management_snapshot.get("consult_tree") or consult_tree
        snapshot_timeline = management_snapshot.get("timeline") or {}
        if snapshot_timeline.get("months"):
            month_entries = snapshot_timeline["months"]
        monthly_planned = management_snapshot.get("monthly_planned") or {}
        monthly_actual = management_snapshot.get("monthly_actual") or {}
        if monthly_planned:
            planned_by_month = {str(key): float(value or 0) for key, value in monthly_planned.items()}
        if monthly_actual:
            actual_by_month = {str(key): float(value or 0) for key, value in monthly_actual.items()}
        snapshot_kpis = management_snapshot.get("management_kpis") or {}
        if snapshot_kpis:
            total_budget = float(snapshot_kpis.get("total_budget_units") or total_budget)
            actual_total = float(snapshot_kpis.get("actual_units") or actual_total)
            physical_pct = round(float(snapshot_kpis.get("physical_pct") or physical_pct), 2)
            planned_today_pct = round(float(snapshot_kpis.get("planned_today_pct") or planned_today_pct), 2)
        late_area_rows = [row for row in area_rows if float(row.get("delta_pct") or 0) < 0]
        worst_area = min(area_rows, key=lambda row: float(row.get("delta_pct") or 0), default={})
        remaining_units = float(snapshot_kpis.get("remaining_units") or remaining_units)
    else:
        snapshot_timeline = {}
        snapshot_kpis = {}

    level_counts = Counter(int(node.get("level") or 0) for node in nodes)
    total_rows = len(nodes) or 1
    level_summary = [
        {
            "level": level,
            "count": count,
            "pct": round(100 * count / total_rows, 2),
            "pct_css": f"{round(100 * count / total_rows, 2):.2f}",
        }
        for level, count in sorted(level_counts.items())
    ]
    executive_level_summary = [row for row in level_summary if int(row["level"]) <= 2]
    render_rows = [
        [
            node["node_id"],
            node["parent_id"],
            node["level"],
            node["display_id"] or "-",
            node["display_name"],
            node["start_label"],
            node["finish_label"],
            f"{float(node.get('pct_complete') or 0):.2f}%",
            node["pct_complete_css"],
            _p6_number_label(node.get("budget_nonlabor")),
            _p6_number_label(node.get("actual_nonlabor")),
            node["bar_left"],
            node["bar_width"],
            bool(node["has_children"]),
            bool(node["initial_expanded"]),
            bool(node["initial_visible"]),
            bool(node["has_bar"]),
            bool(node["is_milestone"]),
        ]
        for node in nodes
    ]

    management_kpis = {
        "late_areas": len(late_area_rows),
        "remaining_units": remaining_units,
        "remaining_units_label": _p6_number_label(remaining_units),
        "remaining_pct": round(max(100 - float(physical_pct or 0), 0), 2),
        "active_packages": len(active_packages),
        "next_90_packages": len(next_90_packages),
        "worst_area_label": worst_area.get("label", "-"),
        "worst_area_delta": worst_area.get("delta_label", "-"),
    }
    if snapshot_kpis:
        management_kpis.update(snapshot_kpis)
    area_performance_chart = _p6_area_performance_chart(area_rows)
    monthly_units_chart = _p6_monthly_units_chart(month_entries, planned_by_month, actual_by_month)

    timeline_payload = {
        "start": min_start,
        "finish": max_finish,
        "start_label": _p6_date_label(min_start),
        "finish_label": _p6_date_label(max_finish),
        "months": month_entries,
        "month_count": len(month_entries),
        "total_days": total_days,
    }
    if snapshot_timeline:
        timeline_payload.update(snapshot_timeline)

    return {
        "source_path": str(path),
        "file_name": path.name,
        "curves_source_path": curves.get("curves_source_path", ""),
        "curves_file_name": curves.get("curves_file_name", ""),
        "curves_imported_at": curves.get("curves_imported_at"),
        "curves_import_id": curves.get("curves_import_id"),
        "rows": nodes,
        "render_rows": render_rows,
        "level_summary": level_summary,
        "executive_level_summary": executive_level_summary,
        "executive_rows": executive_rows,
        "management": {
            "areas": area_rows,
            "critical_packages": critical_rows,
            "timeline_rows": timeline_rows,
            "consult_tree": consult_tree,
            "source": management_snapshot.get("source", "") if management_snapshot else "",
            "source_sheet": management_snapshot.get("source_sheet", "") if management_snapshot else "",
            "kpis": management_kpis,
        },
        "timeline": timeline_payload,
        "kpis": {
            "total_rows": len(nodes),
            "activities": len(leaves),
            "wbs": sum(1 for node in nodes if node.get("has_children")),
            "level_2_rows": int(level_counts.get(2, 0)),
            "executive_rows": len(executive_rows),
            "budget_units": total_budget,
            "actual_units": actual_total,
            "physical_pct": physical_pct,
            "planned_today_pct": planned_today_pct,
            "delta_pct": round(physical_pct - planned_today_pct, 2),
            "start": min_start,
            "finish": max_finish,
            "start_label": _p6_date_label(min_start),
            "finish_label": _p6_date_label(max_finish),
            "duration_days": total_days,
        },
        "charts": {
            "physical_curve": physical_curve_chart,
            "area_performance": area_performance_chart,
            "monthly_units": monthly_units_chart,
        },
    }


@contextmanager
def _spdm_conn():
    db_path = Path(settings.SPDM_DB_PATH)
    if not db_path.exists():
        raise FileNotFoundError(f"SPDM database not found: {db_path}")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()


class _PostgresCompatCursor:
    def __init__(self, cursor):
        self.cursor = cursor

    def execute(self, query: str, params: tuple = ()):
        pg_query = query.replace("%", "%%").replace("?", "%s")
        self.cursor.execute(pg_query, params)
        return self

    def fetchall(self):
        return self.cursor.fetchall()

    def fetchone(self):
        return self.cursor.fetchone()


class _PostgresCompatConnection:
    def __init__(self, conn):
        self.conn = conn

    def cursor(self):
        import psycopg2.extras

        return _PostgresCompatCursor(
            self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        )

    def close(self):
        self.conn.close()


@contextmanager
def _datafy_conn():
    import psycopg2

    conn = psycopg2.connect(
        dbname=settings.DATAFY_DB_NAME,
        user=settings.DATAFY_DB_USER,
        password=settings.DATAFY_DB_PASSWORD,
        host=settings.DATAFY_DB_HOST,
        port=settings.DATAFY_DB_PORT,
        connect_timeout=5,
    )
    compat = _PostgresCompatConnection(conn)
    try:
        yield compat
    finally:
        compat.close()


@contextmanager
def _taskfy_conn():
    import psycopg2
    import psycopg2.extras

    conn = psycopg2.connect(
        dbname=settings.TASKFY_DB_NAME,
        user=settings.TASKFY_DB_USER,
        password=settings.TASKFY_DB_PASSWORD,
        host=settings.TASKFY_DB_HOST,
        port=settings.TASKFY_DB_PORT,
        connect_timeout=5,
        cursor_factory=psycopg2.extras.RealDictCursor,
    )
    try:
        yield conn
    finally:
        conn.close()


def _safe_source(name: str, fn):
    try:
        data = fn()
        data["available"] = True
        data["source_name"] = name
        return data
    except Exception as exc:
        return {
            "available": False,
            "source_name": name,
            "error": str(exc),
            "kpis": {},
            "charts": {},
            "rows": [],
        }


def _sqlite_only_source(name: str, message: str = "External source disabled; using local SQLite snapshots.") -> dict:
    return {
        "available": True,
        "source_name": f"{name} SQLite-only",
        "source_mode": "sqlite_only",
        "external_disabled": True,
        "error": "",
        "message": message,
        "base_url": "",
        "kpis": {},
        "charts": {},
        "rows": [],
    }


def datafy_dashboard() -> dict:
    if settings.DASHFY_SQLITE_ONLY:
        return _sqlite_only_source("DATAFY")
    return _safe_source("DATAFY PostgreSQL", _datafy_dashboard)


def _datafy_dashboard() -> dict:
    with _datafy_conn() as conn:
        cur = conn.cursor()
        total_materials = int(_scalar(cur, "select count(*) from core_materialitem"))
        covered_materials = int(_scalar(cur, "select count(distinct material_item_id) from catalog_allocation"))
        missing_materials = max(total_materials - covered_materials, 0)
        documents = int(_scalar(cur, "select count(*) from core_document"))
        drawings_ok = int(_scalar(cur, "select count(*) from core_document where status='success'"))
        pos = int(_scalar(cur, "select count(*) from core_purchaseorder"))
        po_items = int(_scalar(cur, "select count(*) from core_purchaseorderitem"))
        stock_pieces = int(_scalar(cur, "select count(*) from catalog_stockpiece"))
        allocations = int(_scalar(cur, "select count(*) from catalog_allocation"))
        eclic_docs = int(_scalar(cur, "select count(*) from eclic_eclicdocument"))
        eclic_alerts = int(_scalar(
            cur,
            "select count(*) from eclic_eclicdocument where has_update = true or needs_processing = true",
        ))

        status_rows = _rows(
            cur,
            "select status as label, count(*) as value from core_document group by status order by value desc",
        )
        discipline_rows = _rows(
            cur,
            "select discipline as label, count(*) as value from core_document group by discipline order by value desc",
        )
        po_status_rows = _rows(
            cur,
            "select status as label, count(*) as value from core_purchaseorder group by status order by value desc",
        )
        recent_documents = _rows(
            cur,
            """
            select id, drawing_number, original_filename, title, revision, revision_detail,
                   discipline, status, priority, uploaded_at, processed_at
            from core_document
            order by priority asc, uploaded_at desc
            limit 12
            """,
        )
        recent_pos = _rows(
            cur,
            """
            select id, po_number, supplier_name, status, validation_score,
                   requires_manual_review, uploaded_at, processed_at
            from core_purchaseorder
            order by uploaded_at desc
            limit 8
            """,
        )

    coverage_pct = round(100 * covered_materials / total_materials) if total_materials else 0
    return {
        "base_url": settings.DATAFY_BASE_URL.rstrip("/"),
        "kpis": {
            "documents": documents,
            "drawings_ok": drawings_ok,
            "coverage_pct": coverage_pct,
            "materials": total_materials,
            "covered_materials": covered_materials,
            "missing_materials": missing_materials,
            "pos": pos,
            "po_items": po_items,
            "stock_pieces": stock_pieces,
            "allocations": allocations,
            "eclic_docs": eclic_docs,
            "eclic_alerts": eclic_alerts,
        },
        "charts": {
            "documents_status": _donut_chart(
                [r["label"] or "sem status" for r in status_rows],
                [int(r["value"]) for r in status_rows],
            ),
            "disciplines": _bar_chart(
                [r["label"] or "-" for r in discipline_rows],
                [int(r["value"]) for r in discipline_rows],
            ),
            "po_status": _bar_chart(
                [r["label"] or "sem status" for r in po_status_rows],
                [int(r["value"]) for r in po_status_rows],
                color="#525252",
            ),
        },
        "recent_documents": recent_documents,
        "recent_pos": recent_pos,
    }


def datafy_documents(
    q: str = "",
    status: str = "",
    discipline: str = "",
    revision: str = "",
    priority: str = "",
    limit: int = 500,
) -> dict:
    if settings.DASHFY_SQLITE_ONLY:
        return _sqlite_only_source("DATAFY documents")
    return _safe_source(
        "DATAFY PostgreSQL",
        lambda: _datafy_documents(q, status, discipline, revision, priority, limit),
    )


def _datafy_documents(q: str, status: str, discipline: str, revision: str, priority: str, limit: int) -> dict:
    where = []
    params: list[Any] = []
    if q:
        where.append("(drawing_number ilike ? or original_filename ilike ? or title ilike ?)")
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
    if status:
        where.append("status = ?")
        params.append(status)
    if discipline:
        where.append("discipline = ?")
        params.append(discipline)
    if revision:
        where.append("coalesce(nullif(upper(trim(revision)), ''), '-') = ?")
        params.append(revision.upper())
    if priority:
        where.append("cast(priority as text) = ?")
        params.append(priority)
    clause = " where " + " and ".join(where) if where else ""
    with _datafy_conn() as conn:
        cur = conn.cursor()
        choices = {
            "statuses": _rows(
                cur,
                """
                select coalesce(nullif(status, ''), '-') as value,
                       coalesce(nullif(status, ''), '-') as label,
                       count(*) as total
                from core_document
                group by coalesce(nullif(status, ''), '-')
                order by count(*) desc, value
                """,
            ),
            "disciplines": _rows(
                cur,
                """
                select coalesce(nullif(discipline, ''), '-') as value,
                       coalesce(nullif(discipline, ''), '-') as label,
                       count(*) as total
                from core_document
                group by coalesce(nullif(discipline, ''), '-')
                order by count(*) desc, value
                """,
            ),
            "revisions": _rows(
                cur,
                """
                select coalesce(nullif(upper(trim(revision)), ''), '-') as value,
                       coalesce(nullif(upper(trim(revision)), ''), '-') as label,
                       count(*) as total
                from core_document
                group by coalesce(nullif(upper(trim(revision)), ''), '-')
                order by value
                """,
            ),
            "priorities": _rows(
                cur,
                """
                select cast(priority as text) as value,
                       cast(priority as text) as label,
                       count(*) as total
                from core_document
                where priority is not null
                group by priority
                order by priority asc
                """,
            ),
        }
        total = int(_scalar(cur, f"select count(*) from core_document{clause}", tuple(params)))
        rows = _rows(
            cur,
            f"""
            select d.id, d.drawing_number, d.original_filename, d.title, d.revision,
                   d.revision_detail, d.discipline, d.status, d.priority,
                   d.uploaded_at, d.processed_at,
                   count(distinct t.id) as tables_count,
                   count(mi.id) as material_count
            from core_document d
            left join core_extractedtable t on t.document_id = d.id
            left join core_materialitem mi on mi.table_id = t.id
            {clause}
            group by d.id
            order by d.priority asc, d.uploaded_at desc
            limit ?
            """,
            tuple(params + [limit]),
        )
    return {
        "base_url": settings.DATAFY_BASE_URL.rstrip("/"),
        "total": total,
        "rows": rows,
        "filters": {
            "q": q,
            "status": status,
            "discipline": discipline,
            "revision": revision,
            "priority": priority,
        },
        "choices": choices,
    }


def taskfy_dashboard() -> dict:
    if settings.DASHFY_SQLITE_ONLY:
        return _sqlite_only_source("Taskfy")
    return _safe_source("Taskfy", _taskfy_dashboard)


def _taskfy_dashboard() -> dict:
    today = date.today()
    with _taskfy_conn() as conn:
        cur = conn.cursor()
        cur.execute("select count(*) as n from jobcards_jobcard")
        total = int(cur.fetchone()["n"])
        cur.execute("select count(*) as n from jobcards_jobcard where completed='YES'")
        completed = int(cur.fetchone()["n"])
        cur.execute("select count(*) as n from jobcards_jobcard where jobcard_status='CANCELED'")
        canceled = int(cur.fetchone()["n"])
        cur.execute(
            """
            select count(*) as n
            from jobcards_jobcard
            where finish >= date '2020-01-01'
              and finish < %s
              and completed <> 'YES'
              and jobcard_status <> 'CANCELED'
            """,
            (today,),
        )
        overdue = int(cur.fetchone()["n"])
        cur.execute("select count(*) as n from jobcards_allocatedtask")
        allocated_tasks = int(cur.fetchone()["n"])
        cur.execute("select count(*) as n from jobcards_scheduleactivity")
        schedule_activities = int(cur.fetchone()["n"])
        cur.execute("select count(*) as n from jobcards_jobcardprogrammingpack")
        packs = int(cur.fetchone()["n"])
        cur.execute("select count(*) as n from jobcards_dailyfieldreportheader")
        dfrs = int(cur.fetchone()["n"])
        cur.execute("select count(*) as n from trackfy_shipment")
        shipments = int(cur.fetchone()["n"])

        cur.execute(
            """
            select coalesce(jobcard_status, 'Sem status') as label, count(*) as value
            from jobcards_jobcard
            group by jobcard_status
            order by count(*) desc
            limit 9
            """
        )
        status_rows = [dict(r) for r in cur.fetchall()]
        cur.execute(
            """
            select coalesce(discipline_code, '-') as label, count(*) as value
            from jobcards_jobcard
            group by discipline_code
            order by count(*) desc
            limit 10
            """
        )
        discipline_rows = [dict(r) for r in cur.fetchall()]
        cur.execute(
            """
            select id, job_card_number, discipline_code, discipline, activity_id,
                   jobcard_status, completed, start, finish, working_code, tag
            from jobcards_jobcard
            order by id desc
            limit 12
            """
        )
        recent_jobcards = [dict(r) for r in cur.fetchall()]

    active = max(total - completed - canceled, 0)
    completion_pct = round(100 * completed / total) if total else 0
    return {
        "base_url": settings.TASKFY_BASE_URL.rstrip("/"),
        "kpis": {
            "total": total,
            "active": active,
            "completed": completed,
            "completion_pct": completion_pct,
            "canceled": canceled,
            "overdue": overdue,
            "allocated_tasks": allocated_tasks,
            "schedule_activities": schedule_activities,
            "packs": packs,
            "dfrs": dfrs,
            "shipments": shipments,
        },
        "charts": {
            "status": _bar_chart(
                [str(r["label"]) for r in status_rows],
                [int(r["value"]) for r in status_rows],
            ),
            "disciplines": _bar_chart(
                [str(r["label"]) for r in discipline_rows],
                [int(r["value"]) for r in discipline_rows],
                color="#525252",
            ),
            "completion": _donut_chart(["Concluidas", "Ativas", "Canceladas"], [completed, active, canceled]),
        },
        "recent_jobcards": recent_jobcards,
    }


def taskfy_jobcards(q: str = "", status: str = "", discipline: str = "", limit: int = 500) -> dict:
    if settings.DASHFY_SQLITE_ONLY:
        return _sqlite_only_source("Taskfy jobcards")
    return _safe_source("Taskfy", lambda: _taskfy_jobcards(q, status, discipline, limit))


def _taskfy_jobcards(q: str, status: str, discipline: str, limit: int) -> dict:
    where = []
    params: list[Any] = []
    if q:
        where.append("(job_card_number ilike %s or activity_id ilike %s or tag ilike %s)")
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
    if status:
        where.append("jobcard_status = %s")
        params.append(status)
    if discipline:
        where.append("discipline_code = %s")
        params.append(discipline)
    clause = " where " + " and ".join(where) if where else ""
    with _taskfy_conn() as conn:
        cur = conn.cursor()
        cur.execute(f"select count(*) as n from jobcards_jobcard{clause}", params)
        total = int(cur.fetchone()["n"])
        cur.execute(
            f"""
            select id, job_card_number, discipline_code, discipline, activity_id,
                   jobcard_status, completed, start, finish, working_code, tag,
                   total_man_hours, total_weight, campaign
            from jobcards_jobcard
            {clause}
            order by id desc
            limit %s
            """,
            params + [limit],
        )
        rows = [dict(r) for r in cur.fetchall()]
    return {
        "base_url": settings.TASKFY_BASE_URL.rstrip("/"),
        "total": total,
        "rows": rows,
        "filters": {"q": q, "status": status, "discipline": discipline},
    }


def taskfy_kanban() -> dict:
    if settings.DASHFY_SQLITE_ONLY:
        return _sqlite_only_source("Taskfy kanban")
    def _load():
        with _taskfy_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                """
                select jobcard_status as status, count(*) as total
                from jobcards_jobcard
                group by jobcard_status
                order by count(*) desc
                limit 8
                """
            )
            statuses = [dict(r) for r in cur.fetchall()]
            columns = []
            for row in statuses:
                status = row["status"] or "Sem status"
                cur.execute(
                    """
                    select id, job_card_number, discipline_code, activity_id, completed,
                           start, finish, working_code, tag
                    from jobcards_jobcard
                    where coalesce(jobcard_status, 'Sem status') = %s
                    order by id desc
                    limit 30
                    """,
                    (status,),
                )
                columns.append({"label": status, "total": int(row["total"]), "cards": [dict(r) for r in cur.fetchall()]})
        return {"base_url": settings.TASKFY_BASE_URL.rstrip("/"), "columns": columns}

    return _safe_source("Taskfy", _load)


def _contract_week_start(contract_week: str) -> int | str:
    try:
        number = int(str(contract_week).strip().upper().replace("W", ""))
    except Exception:
        return str(contract_week)
    anchor = date(2024, 5, 25)
    return (anchor.toordinal() + ((number - 1) * 7))


def _contract_week_label(contract_week: str) -> str:
    ordinal = _contract_week_start(contract_week)
    if isinstance(ordinal, int):
        return date.fromordinal(ordinal).isoformat()
    return str(ordinal)


def _construction_filters(raw: dict | None) -> dict:
    raw = raw or {}
    today = date.today()
    contract_week = (raw.get("contract_week") or "").strip().upper()
    discipline = (raw.get("discipline") or "").strip()
    supply_discipline = (raw.get("supply_discipline") or discipline).strip()

    def coverage_threshold(name: str) -> int:
        try:
            value = int(raw.get(name) or 0)
        except (TypeError, ValueError):
            value = 0
        return value if value in {0, 25, 50, 80, 100} else 0

    try:
        min_readiness = int(raw.get("min_readiness") or 0)
    except (TypeError, ValueError):
        min_readiness = 0
    if min_readiness not in {0, 25, 50, 80, 100}:
        min_readiness = 0
    if contract_week and not (raw.get("date_from") or raw.get("date_to")):
        ordinal = _contract_week_start(contract_week)
        if isinstance(ordinal, int):
            start = date.fromordinal(ordinal)
            end = date.fromordinal(ordinal + 6)
        else:
            start = date(today.year, 1, 1)
            end = today
    else:
        start = _parse_date(raw.get("date_from"), date(today.year, 1, 1))
        end = _parse_date(raw.get("date_to"), today)
    if start > end:
        start, end = end, start
    return {
        "start": start,
        "end": end,
        "date_from": start.isoformat(),
        "date_to": end.isoformat(),
        "discipline": discipline,
        "engineering_discipline": (raw.get("engineering_discipline") or "").strip(),
        "engineering_status": (raw.get("engineering_status") or "").strip(),
        "engineering_issue_status": (raw.get("engineering_issue_status") or "").strip(),
        "engineering_revision": (raw.get("engineering_revision") or "").strip().upper(),
        "engineering_responsible": (raw.get("engineering_responsible") or "").strip(),
        "engineering_q": (raw.get("engineering_q") or "").strip(),
        "supply_priority": (raw.get("supply_priority") or "").strip(),
        "supply_drawing_q": (raw.get("supply_drawing_q") or "").strip(),
        "supply_revision": (raw.get("supply_revision") or "").strip().upper(),
        "supply_discipline": supply_discipline,
        "supply_line": (raw.get("supply_line") or "").strip(),
        "supply_table": (raw.get("supply_table") or "").strip(),
        "supply_page": (raw.get("supply_page") or "").strip(),
        "supply_item": (raw.get("supply_item") or "").strip(),
        "supply_family": (raw.get("supply_family") or "").strip(),
        "supply_code_q": (raw.get("supply_code_q") or "").strip(),
        "supply_description_q": (raw.get("supply_description_q") or "").strip(),
        "supply_fab_min": coverage_threshold("supply_fab_min"),
        "supply_erection_min": coverage_threshold("supply_erection_min"),
        "campaign": (raw.get("campaign") or "").strip(),
        "contract_week": contract_week,
        "min_readiness": min_readiness,
    }


def _append_taskfy_filters(filters: dict, alias: str, where: list[str], params: list[Any]) -> None:
    if filters["discipline"]:
        where.append(f"{alias}.discipline_code = %s")
        params.append(filters["discipline"])
    if filters["campaign"]:
        where.append(f"coalesce(nullif({alias}.campaign, ''), '-') = %s")
        params.append(filters["campaign"])


def _taskfy_choices() -> dict:
    with _taskfy_conn() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            select discipline_code as value,
                   coalesce(max(nullif(discipline, '')), discipline_code) as label,
                   count(*) as total
            from jobcards_jobcard
            where coalesce(discipline_code, '') <> ''
            group by discipline_code
            order by count(*) desc, discipline_code
            """
        )
        disciplines = [dict(r) for r in cur.fetchall()]
        cur.execute(
            """
            select coalesce(nullif(campaign, ''), '-') as value, count(*) as total
            from jobcards_jobcard
            group by coalesce(nullif(campaign, ''), '-')
            order by count(*) desc
            limit 20
            """
        )
        campaigns = [dict(r) for r in cur.fetchall()]
        cur.execute(
            """
            select contract_week as value, count(*) as total
            from jobcards_jobcardprogrammingpack
            group by contract_week
            order by contract_week desc
            """
        )
        weeks = [dict(r) for r in cur.fetchall()]
    return {
        "disciplines": disciplines,
        "campaigns": campaigns,
        "weeks": weeks,
        "engineering_disciplines": [],
        "engineering_statuses": [],
        "engineering_issue_statuses": [],
        "engineering_responsibles": [],
        "engineering_revisions": [
            {"value": "R", "label": "REV. R"},
            {"value": "A", "label": "REV. A"},
            {"value": "C", "label": "REV. C"},
        ],
    }


def _eclic_value(item: dict, *keys: str, default: str = "") -> str:
    for key in keys:
        value = item.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return default


def _eclic_documents_from_api() -> list[dict]:
    client = EclicClient()
    cache_key = f"eclic-api-documents:{client.base_url}:{client.client_id}:{client.project_id}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached
    documents = list(client.list_documents())
    cache.set(cache_key, documents, 300)
    return documents


def _engineering_revision_bucket(revision: str) -> str:
    value = (revision or "").strip().upper()
    if value.startswith("R"):
        return "REV. R"
    if value.startswith("A"):
        return "REV. A"
    if value.startswith("C"):
        return "REV. C"
    return ""


def _engineering_control_summary(docs: list[dict[str, Any]], discipline_rows: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(docs)

    def pct(value: int) -> float:
        return round(100 * value / total, 2) if total else 0.0

    issued = sum(
        1 for doc in docs
        if str(doc.get("status_documento") or "").strip().upper() in {"ISSUED", "EMITIDO"}
    )
    in_engineering = sum(
        1 for doc in docs
        if str(doc.get("status_documento") or "").strip().upper().startswith("ENGINEERING")
    )
    pending = max(total - issued - in_engineering, 0)

    rev_r = sum(1 for doc in docs if str(doc.get("revisao") or "").strip().upper().startswith("R"))
    rev_a = sum(1 for doc in docs if str(doc.get("revisao") or "").strip().upper().startswith("A"))
    rev_c = sum(1 for doc in docs if str(doc.get("revisao") or "").strip().upper().startswith("C"))
    rev_other = max(total - rev_r - rev_a - rev_c, 0)
    afc = sum(
        1 for doc in docs
        if str(doc.get("doc_status_group") or "").strip().upper() == "AFC"
        or str(doc.get("doc_status") or "").strip().upper().startswith("AFC")
    )
    if afc == 0:
        afc = rev_c
    non_afc = max(total - afc, 0)

    def item(key: str, label: str, value: int, color: str, icon: str, detail: str) -> dict[str, Any]:
        item_pct = pct(value)
        return {
            "key": key,
            "label": label,
            "value": value,
            "pct": item_pct,
            "pct_css": f"{item_pct:.2f}",
            "color": color,
            "icon": icon,
            "detail": detail,
        }

    status_items = [
        item("issued", "Emitidos", issued, "#059669", "bi-check2-circle", "Documentos liberados para a cadeia."),
        item("engineering", "Engenharia", in_engineering, "#2563eb", "bi-pencil-square", "Carteira ainda em elaboracao tecnica."),
        item("pending", "Outros", pending, "#d97706", "bi-hourglass-split", "Pendencias fora do fluxo emitido."),
    ]
    revision_items = [
        item("rev_r", "Rev. R", rev_r, "#2563eb", "bi-arrow-repeat", "Revisoes R na carteira filtrada."),
        item("rev_a", "Rev. A", rev_a, "#0891b2", "bi-layers", "Revisoes A na carteira filtrada."),
        item("rev_c", "AFC / Rev. C", rev_c, "#059669", "bi-patch-check", "Documentos em AFC, considerados como revisao C."),
        item("rev_other", "Outras", rev_other, "#737373", "bi-three-dots", "Revisoes fora dos buckets R/A/C."),
    ]
    max_revision_value = max((int(row["value"] or 0) for row in revision_items), default=0)
    for row in revision_items:
        relative_pct = round(100 * int(row["value"] or 0) / max_revision_value, 2) if max_revision_value else 0.0
        row["relative_pct"] = relative_pct
        row["relative_pct_css"] = f"{relative_pct:.2f}"

    top_discipline_source = discipline_rows[:5]
    max_discipline_value = max((int(row.get("total") or 0) for row in top_discipline_source), default=0)
    top_disciplines = []
    for row in top_discipline_source:
        value = int(row.get("total") or 0)
        discipline_pct = pct(value)
        issued_for_discipline = int(row.get("issued") or 0)
        issued_pct = round(100 * issued_for_discipline / value, 2) if value else 0.0
        relative_pct = round(100 * value / max_discipline_value, 2) if max_discipline_value else 0.0
        top_disciplines.append({
            "label": row.get("discipline") or "-",
            "value": value,
            "pct": discipline_pct,
            "pct_css": f"{discipline_pct:.2f}",
            "relative_pct": relative_pct,
            "relative_pct_css": f"{relative_pct:.2f}",
            "issued": issued_for_discipline,
            "issued_pct": issued_pct,
            "issued_pct_css": f"{issued_pct:.2f}",
            "in_engineering": int(row.get("in_engineering") or 0),
        })

    return {
        "total": total,
        "afc": afc,
        "afc_pct": pct(afc),
        "afc_pct_css": f"{pct(afc):.2f}",
        "non_afc": non_afc,
        "non_afc_pct": pct(non_afc),
        "non_afc_pct_css": f"{pct(non_afc):.2f}",
        "issued": issued,
        "issued_pct": pct(issued),
        "issued_pct_css": f"{pct(issued):.2f}",
        "in_engineering": in_engineering,
        "in_engineering_pct": pct(in_engineering),
        "pending": pending,
        "pending_pct": pct(pending),
        "status_items": status_items,
        "revision_items": revision_items,
        "top_disciplines": top_disciplines,
        "top_discipline": top_disciplines[0] if top_disciplines else {},
        "rev_r": rev_r,
        "rev_a": rev_a,
        "rev_c": rev_c,
        "rev_other": rev_other,
    }


def _engineering_empty(error: str = "") -> dict:
    return {
        "source": "ECLIC API",
        "error": error,
        "engineering_docs": 0,
        "engineering_counts": {"disciplines": 0, "revisions": 0, "issued": 0, "in_engineering": 0},
        "engineering_flow": _engineering_control_summary([], []),
        "engineering_summary": [],
        "engineering_discipline_groups": [],
        "engineering_revision_rows": [],
        "engineering_status_rows": [],
        "engineering_documents": [],
        "choices": {
            "engineering_disciplines": [],
            "engineering_statuses": [],
            "engineering_issue_statuses": [],
            "engineering_responsibles": [],
            "engineering_revisions": [
                {"value": "R", "label": "REV. R"},
                {"value": "A", "label": "REV. A"},
                {"value": "C", "label": "REV. C"},
            ],
        },
    }


_ENGINEERING_MONITOR_STATUS_ORDER = (
    "NI",
    "IFR",
    "IFA",
    "IFI",
    "AFC 1",
    "AFC 3",
    "AFC CODE 3A",
    "UNDER REVIEW",
)

_ENGINEERING_MONITOR_AFC_STATUSES = {"AFC 1", "AFC 3", "AFC CODE 3A", "AFC 3A"}
_ENGINEERING_MONITOR_AFC_3A_STATUSES = {"AFC CODE 3A", "AFC 3A"}


def _engineering_monitor_empty(error: str = "") -> dict[str, Any]:
    return {
        "source": "Engineering base monitor",
        "source_mode": "sqlite_monitor_missing",
        "error": error,
        "import_id": None,
        "imported_at": None,
        "source_file_name": "",
        "detail_sheet": "",
        "document_count": 0,
        "raw_document_count": 0,
        "excluded_count": 0,
        "flow": {
            "total": 0,
            "afc": 0,
            "afc_pct": 0,
            "in_engineering": 0,
            "in_engineering_pct": 0,
            "issued": 0,
            "issued_pct": 0,
            "rev_r": 0,
            "rev_a": 0,
            "rev_c": 0,
            "rev_other": 0,
        },
        "summary": [],
        "status_rows": [],
        "revision_rows": [],
        "documents": [],
        "metadata": {},
    }


def _engineering_monitor_pct(value: int | float, total: int | float) -> float:
    return round(100 * float(value or 0) / float(total or 0), 2) if total else 0.0


def _engineering_monitor_flow(docs: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(docs)
    ifr = sum(1 for doc in docs if str(doc.get("status_bucket") or "").upper() == "IFR")
    ifa = sum(1 for doc in docs if str(doc.get("status_bucket") or "").upper() == "IFA")
    afc = sum(1 for doc in docs if str(doc.get("status_bucket") or "").upper() in _ENGINEERING_MONITOR_AFC_STATUSES)
    under_review = sum(1 for doc in docs if str(doc.get("status_bucket") or "").upper() == "UNDER REVIEW")
    issued = sum(
        1 for doc in docs
        if str(doc.get("document_status_effective") or "").upper() in {"ISSUED", "EMITIDO"}
        or str(doc.get("status_bucket") or "").upper() in {*_ENGINEERING_MONITOR_AFC_STATUSES, "IFI"}
    )
    rev_r = sum(1 for doc in docs if str(doc.get("revision_family") or "").upper() == "R")
    rev_a = sum(1 for doc in docs if str(doc.get("revision_family") or "").upper() == "A")
    rev_c = sum(1 for doc in docs if str(doc.get("revision_family") or "").upper() == "C")
    rev_other = max(total - rev_r - rev_a - rev_c, 0)
    return {
        "total": total,
        "afc": afc,
        "afc_pct": _engineering_monitor_pct(afc, total),
        "ifr": ifr,
        "ifr_pct": _engineering_monitor_pct(ifr, total),
        "ifa": ifa,
        "ifa_pct": _engineering_monitor_pct(ifa, total),
        "in_engineering": under_review,
        "in_engineering_pct": _engineering_monitor_pct(under_review, total),
        "issued": issued,
        "issued_pct": _engineering_monitor_pct(issued, total),
        "rev_r": rev_r,
        "rev_a": rev_a,
        "rev_c": rev_c,
        "rev_other": rev_other,
    }


def _engineering_monitor_normalized_doc(doc: dict[str, Any]) -> dict[str, Any]:
    discipline = (
        normalize_monitor_discipline(doc.get("source_discipline") or doc.get("discipline"), doc.get("title"))
        or normalize_monitor_discipline(doc.get("discipline"), doc.get("title"))
    )
    if discipline and discipline != doc.get("discipline"):
        return {**doc, "discipline": discipline}
    return doc


def _engineering_monitor_from_snapshot(filters: dict) -> dict[str, Any]:
    try:
        latest = EngineeringMonitorImport.objects.filter(is_active=True).first()
    except (OperationalError, ProgrammingError):
        return _engineering_monitor_empty("Engineering monitor table is not migrated.")
    if latest is None:
        return _engineering_monitor_empty("No active engineering base import.")

    payload = latest.payload or {}
    raw_docs = [_engineering_monitor_normalized_doc(dict(doc)) for doc in list(payload.get("documents") or [])]
    monitored_docs = [doc for doc in raw_docs if doc.get("is_monitored")]
    countable_docs = [doc for doc in monitored_docs if doc.get("is_countable")]

    filtered_docs = list(countable_docs)
    if filters.get("engineering_revision") in {"R", "A", "C"}:
        filtered_docs = [doc for doc in filtered_docs if doc.get("revision_family") == filters["engineering_revision"]]
    if filters.get("engineering_q"):
        query = str(filters["engineering_q"]).lower()
        filtered_docs = [
            doc for doc in filtered_docs
            if query in " ".join([
                str(doc.get("document_number") or ""),
                str(doc.get("title") or ""),
                str(doc.get("discipline") or ""),
                str(doc.get("source_discipline") or ""),
                str(doc.get("revision") or ""),
                str(doc.get("status_bucket") or ""),
                str(doc.get("issue_status") or ""),
                str(doc.get("last_transmittal_purpose") or ""),
                str(doc.get("fabrication_ref") or ""),
            ]).lower()
        ]

    total = len(filtered_docs)
    discipline_order = monitor_discipline_order({str(doc.get("discipline") or "-") for doc in monitored_docs})

    status_counts = Counter(str(doc.get("status_bucket") or "") for doc in filtered_docs)
    excluded_by_discipline: dict[str, Counter[str]] = {}
    for doc in monitored_docs:
        if doc.get("is_countable"):
            continue
        discipline = str(doc.get("discipline") or "-")
        excluded_by_discipline.setdefault(discipline, Counter())[str(doc.get("excluded_reason") or "Excluded")] += 1

    summary = []
    revision_counts: Counter[tuple[str, str]] = Counter()
    revision_distinct = set()
    for doc in filtered_docs:
        discipline = str(doc.get("discipline") or "-")
        revision = str(doc.get("revision") or "")
        family = str(doc.get("revision_family") or "")
        if revision:
            revision_distinct.add(revision)
        if family in {"R", "A", "C"}:
            revision_counts[(discipline, f"REV. {family}")] += 1

    for discipline in discipline_order:
        docs_for_discipline = [doc for doc in filtered_docs if doc.get("discipline") == discipline]
        row_total = len(docs_for_discipline)
        excluded_counts = excluded_by_discipline.get(discipline, Counter())
        remarks = []
        if excluded_counts:
            remarks = [f"{reason}: {count}" for reason, count in sorted(excluded_counts.items())]
        row = {
            "discipline": discipline,
            "total": row_total,
            "pct_total": round(100 * row_total / total, 1) if total else 0,
            "rev_a": sum(1 for doc in docs_for_discipline if doc.get("revision_family") == "A"),
            "rev_r": sum(1 for doc in docs_for_discipline if doc.get("revision_family") == "R"),
            "rev_c": sum(1 for doc in docs_for_discipline if doc.get("revision_family") == "C"),
            "rev_other": sum(1 for doc in docs_for_discipline if doc.get("revision_family") not in {"R", "A", "C"}),
            "not_issued": sum(1 for doc in docs_for_discipline if doc.get("status_bucket") == "NI"),
            "ifr": sum(1 for doc in docs_for_discipline if doc.get("status_bucket") == "IFR"),
            "ifa": sum(1 for doc in docs_for_discipline if doc.get("status_bucket") == "IFA"),
            "ifi": sum(1 for doc in docs_for_discipline if doc.get("status_bucket") == "IFI"),
            "afc_code1": sum(1 for doc in docs_for_discipline if doc.get("status_bucket") == "AFC 1"),
            "afc_code3": sum(1 for doc in docs_for_discipline if doc.get("status_bucket") == "AFC 3"),
            "afc_code3a": sum(1 for doc in docs_for_discipline if doc.get("status_bucket") in _ENGINEERING_MONITOR_AFC_3A_STATUSES),
            "under_review": sum(1 for doc in docs_for_discipline if doc.get("status_bucket") == "UNDER REVIEW"),
            "issued": sum(1 for doc in docs_for_discipline if doc.get("status_bucket") in {*_ENGINEERING_MONITOR_AFC_STATUSES, "IFI"}),
            "in_engineering": sum(1 for doc in docs_for_discipline if doc.get("status_bucket") == "UNDER REVIEW"),
            "excluded": sum(excluded_counts.values()),
            "remarks": "; ".join(remarks),
        }
        summary.append(row)

    discipline_rank = {discipline: index for index, discipline in enumerate(discipline_order)}
    summary.sort(key=lambda row: (discipline_rank.get(str(row.get("discipline") or ""), 999), str(row.get("discipline") or "")))
    status_order = payload.get("status_order") or list(_ENGINEERING_MONITOR_STATUS_ORDER)
    status_rows = [
        {"label": label, "value": int(status_counts.get(label, 0))}
        for label in status_order
        if status_counts.get(label, 0)
    ]
    revision_rank = {"REV. R": 0, "REV. A": 1, "REV. C": 2}
    revision_rows = [
        {"label": discipline, "revision": revision, "total": count}
        for (discipline, revision), count in sorted(
            revision_counts.items(),
            key=lambda item: (
                discipline_rank.get(str(item[0][0] or ""), 999),
                revision_rank.get(str(item[0][1] or ""), 99),
                str(item[0][0] or ""),
                str(item[0][1] or ""),
            ),
        )
    ]

    status_rank = {label: index for index, label in enumerate(status_order)}
    documents = sorted(
        filtered_docs,
        key=lambda doc: (
            status_rank.get(str(doc.get("status_bucket") or ""), 99),
            discipline_rank.get(str(doc.get("discipline") or ""), 999),
            str(doc.get("discipline") or ""),
            str(doc.get("document_number") or ""),
        ),
    )[:24]

    return {
        "source": "Engineering base monitor",
        "source_mode": "sqlite_monitor",
        "error": "",
        "import_id": latest.pk,
        "imported_at": latest.created_at,
        "source_file_name": latest.original_filename,
        "detail_sheet": latest.detail_sheet,
        "document_count": total,
        "raw_document_count": latest.document_count,
        "excluded_count": latest.excluded_count,
        "flow": _engineering_monitor_flow(filtered_docs),
        "summary": summary,
        "status_rows": status_rows,
        "revision_rows": revision_rows,
        "documents": documents,
        "metadata": {
            **(latest.metadata or {}),
            "revision_count": len(revision_distinct),
        },
    }


def _engineering_is_under_review_code(value: Any) -> bool:
    text = str(value or "").strip().upper()
    return text == "MABU" or "UNDER REVIEW" in text


def _engineering_ded_status_label(value: Any) -> str:
    text = " ".join(str(value or "").strip().split())
    if not text:
        return "-"
    if text.upper() == "MABU":
        return "UNDER REVIEW"
    if "MABU" in text.upper() and "UNDER REVIEW" in text.upper():
        return " ".join(part for part in text.split() if part.upper() != "MABU")
    return text


def _engineering_from_ded_snapshot(filters: dict) -> dict:
    try:
        latest = (
            EngineeringStatusImport.objects
            .filter(is_active=True)
            .prefetch_related("documents", "discipline_summaries")
            .first()
        )
    except (OperationalError, ProgrammingError):
        return {}
    if latest is None:
        return {}

    raw_docs = list(latest.documents.all().values(
        "document_number",
        "title",
        "discipline",
        "revision",
        "revision_family",
        "revision_number",
        "doc_status",
        "doc_status_group",
        "afc_code",
        "document_status",
        "workflow_start",
        "workflow_end",
        "responsible",
        "issue_status",
        "fabrication_ref",
    ))
    summary_remarks = {
        row.discipline: row.remarks
        for row in latest.discipline_summaries.all()
    }

    base_docs = []
    for row in raw_docs:
        workflow_start = row.get("workflow_start")
        workflow_end = row.get("workflow_end")
        afc_code = row.get("afc_code") or ""
        base_docs.append({
            "codigo": row.get("document_number") or "-",
            "codigo_secundario": row.get("fabrication_ref") or "",
            "titulo": row.get("title") or "-",
            "disciplina": row.get("discipline") or "-",
            "revisao": row.get("revision") or "-",
            "revision_family": row.get("revision_family") or "",
            "revision_number": int(row.get("revision_number") or 0),
            "doc_status": _engineering_ded_status_label(row.get("doc_status")),
            "doc_status_group": row.get("doc_status_group") or "",
            "afc_code": "UNDER REVIEW" if _engineering_is_under_review_code(afc_code) else afc_code,
            "status_documento": row.get("document_status") or "-",
            "status_emissao": row.get("issue_status") or "-",
            "responsavel_atividade": row.get("responsible") or "-",
            "workflow_start": workflow_start,
            "workflow_end": workflow_end,
            "workflow_start_label": workflow_start.strftime("%d/%m/%y") if workflow_start else "-",
            "workflow_end_label": workflow_end.strftime("%d/%m/%y") if workflow_end else "-",
            "fabrication_ref": row.get("fabrication_ref") or "",
            "remarks": summary_remarks.get(row.get("discipline") or "-", ""),
        })

    discipline_counts = Counter(doc["disciplina"] for doc in base_docs)
    doc_status_counts = Counter(doc["doc_status"] for doc in base_docs)
    issue_counts = Counter(doc["status_emissao"] for doc in base_docs if doc["status_emissao"] != "-")
    responsible_counts = Counter(doc["responsavel_atividade"] for doc in base_docs if doc["responsavel_atividade"] != "-")
    revision_families = Counter(doc["revision_family"] for doc in base_docs if doc["revision_family"])
    choices = {
        "engineering_disciplines": [
            {"value": value, "total": total}
            for value, total in sorted(discipline_counts.items(), key=lambda pair: (-pair[1], pair[0]))
        ],
        "engineering_statuses": [
            {"value": value, "total": total}
            for value, total in sorted(doc_status_counts.items(), key=lambda pair: (-pair[1], pair[0]))
        ],
        "engineering_issue_statuses": [
            {"value": value, "total": total}
            for value, total in sorted(issue_counts.items(), key=lambda pair: (-pair[1], pair[0]))
        ],
        "engineering_responsibles": [
            {"value": value, "total": total}
            for value, total in sorted(responsible_counts.items(), key=lambda pair: (-pair[1], pair[0]))
        ],
        "engineering_revisions": [
            {"value": value, "label": f"REV. {value}", "total": revision_families[value]}
            for value in ("R", "A", "C")
            if revision_families.get(value)
        ],
    }

    filtered_docs = list(base_docs)
    discipline_filter = filters.get("engineering_discipline") or filters.get("discipline")
    if discipline_filter:
        filtered_docs = [doc for doc in filtered_docs if doc["disciplina"] == discipline_filter]
    if filters.get("engineering_status"):
        status_filter = filters["engineering_status"].upper()
        filtered_docs = [
            doc for doc in filtered_docs
            if doc["doc_status"].upper() == status_filter
            or doc["status_documento"].upper() == status_filter
        ]
    if filters.get("engineering_issue_status"):
        issue_filter = filters["engineering_issue_status"].upper()
        filtered_docs = [doc for doc in filtered_docs if doc["status_emissao"].upper() == issue_filter]
    if filters.get("engineering_responsible"):
        responsible_filter = filters["engineering_responsible"].upper()
        filtered_docs = [doc for doc in filtered_docs if doc["responsavel_atividade"].upper() == responsible_filter]
    if filters.get("engineering_revision") in {"R", "A", "C"}:
        filtered_docs = [doc for doc in filtered_docs if doc["revision_family"] == filters["engineering_revision"]]
    if filters.get("engineering_q"):
        query = filters["engineering_q"].lower()
        filtered_docs = [
            doc for doc in filtered_docs
            if query in " ".join([
                doc["codigo"],
                doc["titulo"],
                doc["disciplina"],
                doc["revisao"],
                doc["doc_status"],
                doc["status_documento"],
                doc["status_emissao"],
            ]).lower()
        ]

    engineering_docs = len(filtered_docs)
    filtered_discipline_counts = Counter(doc["disciplina"] for doc in filtered_docs)
    filtered_status_counts = Counter(doc["doc_status"] for doc in filtered_docs)
    revision_counts: Counter[tuple[str, str]] = Counter()
    revision_distinct = set()
    engineering_summary = []

    for doc in filtered_docs:
        if doc["revisao"] and doc["revisao"] != "-":
            revision_distinct.add(doc["revisao"])
        if doc["revision_family"] in {"R", "A", "C"}:
            revision_counts[(doc["disciplina"], f"REV. {doc['revision_family']}")] += 1

    for discipline, total in sorted(filtered_discipline_counts.items(), key=lambda pair: (-pair[1], pair[0]))[:14]:
        docs_for_discipline = [doc for doc in filtered_docs if doc["disciplina"] == discipline]
        not_issued = sum(1 for doc in docs_for_discipline if doc["doc_status_group"] == "NOT ISSUED")
        ifr = sum(1 for doc in docs_for_discipline if doc["doc_status_group"] == "IFR")
        ifa = sum(1 for doc in docs_for_discipline if doc["doc_status_group"] == "IFA")
        afc_code1 = sum(1 for doc in docs_for_discipline if doc["doc_status_group"] == "AFC" and doc["afc_code"] == "1")
        afc_code3 = sum(1 for doc in docs_for_discipline if doc["doc_status_group"] == "AFC" and doc["afc_code"] == "3")
        under_review = sum(1 for doc in docs_for_discipline if doc["doc_status_group"] == "AFC" and _engineering_is_under_review_code(doc["afc_code"]))
        afc_total = afc_code1 + afc_code3 + under_review
        engineering_summary.append({
            "discipline": discipline,
            "total": total,
            "pct_total": round(100 * total / engineering_docs, 1) if engineering_docs else 0,
            "rev_a": sum(1 for doc in docs_for_discipline if doc["revision_family"] == "A"),
            "rev_r": sum(1 for doc in docs_for_discipline if doc["revision_family"] == "R"),
            "rev_c": sum(1 for doc in docs_for_discipline if doc["revision_family"] == "C"),
            "rev_other": sum(1 for doc in docs_for_discipline if doc["revision_family"] not in {"R", "A", "C"}),
            "not_issued": not_issued,
            "ifr": ifr,
            "ifa": ifa,
            "afc": afc_total,
            "afc_code1": afc_code1,
            "afc_code3": afc_code3,
            "under_review": under_review,
            "mabu_under_review": under_review,
            "issued": sum(1 for doc in docs_for_discipline if doc["status_documento"].upper() in {"ISSUED", "EMITIDO"}),
            "in_engineering": sum(1 for doc in docs_for_discipline if doc["status_documento"].upper().startswith("ENGINEERING")),
            "remarks": summary_remarks.get(discipline, ""),
        })

    engineering_discipline_groups = []
    for discipline, total in sorted(filtered_discipline_counts.items(), key=lambda pair: (-pair[1], pair[0]))[:18]:
        docs_for_discipline = [doc for doc in filtered_docs if doc["disciplina"] == discipline]
        docs_for_discipline.sort(key=lambda doc: (doc["doc_status"], doc["revisao"], doc["codigo"]))
        engineering_discipline_groups.append({
            "label": discipline,
            "value": total,
            "pct": round(100 * total / engineering_docs, 1) if engineering_docs else 0,
            "documents": docs_for_discipline,
            "remarks": summary_remarks.get(discipline, ""),
        })

    engineering_revision_rows = [
        {"label": discipline, "revision": revision, "total": total}
        for (discipline, revision), total in sorted(revision_counts.items())
    ]
    engineering_status_rows = [
        {"label": label, "value": total}
        for label, total in sorted(filtered_status_counts.items(), key=lambda pair: (-pair[1], pair[0]))[:8]
    ]
    _status_total = sum(row["value"] for row in engineering_status_rows) or 1
    _status_max = max((row["value"] for row in engineering_status_rows), default=0) or 1
    for _row in engineering_status_rows:
        _pct = round(100 * _row["value"] / _status_total, 2)
        _rel = round(100 * _row["value"] / _status_max, 2)
        _row["pct"] = _pct
        _row["pct_css"] = f"{_pct:.2f}"
        _row["relative_pct"] = _rel
        _row["relative_pct_css"] = f"{_rel:.2f}"

    def _doc_rank(doc: dict[str, Any]) -> tuple[int, str, str]:
        group = doc.get("doc_status_group") or ""
        code = doc.get("afc_code") or ""
        rank = {
            "NOT ISSUED": 0,
            "IFR": 1,
            "IFA": 2,
        }.get(group, 4)
        if group == "AFC" and _engineering_is_under_review_code(code):
            rank = 3
        return (rank, doc.get("disciplina") or "", doc.get("codigo") or "")

    engineering_documents = sorted(filtered_docs, key=_doc_rank)[:24]
    engineering_flow = _engineering_control_summary(filtered_docs, engineering_summary)
    return {
        "source": "DED XLSX / Base importada",
        "source_mode": "sqlite_snapshot",
        "error": "",
        "import_id": latest.pk,
        "imported_at": latest.created_at,
        "source_file_name": latest.original_filename,
        "summary_sheet": latest.summary_sheet,
        "detail_sheet": latest.detail_sheet,
        "engineering_docs": engineering_docs,
        "engineering_counts": {
            "disciplines": len(filtered_discipline_counts),
            "revisions": len(revision_distinct),
            "issued": sum(1 for doc in filtered_docs if doc["status_documento"].upper() in {"ISSUED", "EMITIDO"}),
            "in_engineering": sum(1 for doc in filtered_docs if doc["status_documento"].upper().startswith("ENGINEERING")),
        },
        "engineering_flow": engineering_flow,
        "engineering_summary": engineering_summary,
        "engineering_discipline_groups": engineering_discipline_groups,
        "engineering_revision_rows": engineering_revision_rows,
        "engineering_status_rows": engineering_status_rows,
        "engineering_documents": engineering_documents,
        "choices": choices,
    }


def _engineering_from_eclic_api(filters: dict) -> dict:
    try:
        raw_documents = _eclic_documents_from_api()
    except EclicAPIError as exc:
        return _engineering_empty(str(exc))

    base_docs = []
    cancelled = {"CANCELADO", "CANCELED", "CANCELLED"}
    managerial = set(ECLIC_MANAGERIAL_DISCIPLINES)
    for item in raw_documents:
        discipline = _eclic_value(item, "Discipline", "discipline", "disciplina", default="Sem disciplina")
        status = _eclic_value(item, "DocumentStatus", "status_documento", "status", default="Sem status")
        if discipline not in managerial:
            continue
        if status.upper() in cancelled:
            continue
        code = _eclic_value(item, "Code", "code", "codigo", default="-")
        revision = _eclic_value(item, "Revision", "revision", "revisao", default="-")
        title = _eclic_value(item, "Title", "title", "titulo", "name", default="-")
        base_docs.append({
            "codigo": code,
            "codigo_secundario": "",
            "titulo": title,
            "disciplina": discipline,
            "revisao": revision,
            "status_documento": status,
            "status_emissao": _eclic_value(item, "IssueStatus", "status_emissao", default="-"),
            "responsavel_atividade": _eclic_value(item, "Responsible", "responsavel_atividade", default="-"),
            "eclic_id": item.get("Id") or item.get("id"),
        })

    discipline_counts = Counter(doc["disciplina"] for doc in base_docs)
    status_counts = Counter(doc["status_documento"] for doc in base_docs)
    choices = {
        "engineering_disciplines": [
            {"value": value, "total": total}
            for value, total in sorted(discipline_counts.items(), key=lambda pair: (-pair[1], pair[0]))
        ],
        "engineering_statuses": [
            {"value": value, "total": total}
            for value, total in sorted(status_counts.items(), key=lambda pair: (-pair[1], pair[0]))
        ],
        "engineering_issue_statuses": [],
        "engineering_responsibles": [],
        "engineering_revisions": [
            {"value": "R", "label": "REV. R"},
            {"value": "A", "label": "REV. A"},
            {"value": "C", "label": "REV. C"},
        ],
    }

    filtered_docs = list(base_docs)
    if filters["engineering_discipline"]:
        filtered_docs = [doc for doc in filtered_docs if doc["disciplina"] == filters["engineering_discipline"]]
    if filters["engineering_status"]:
        filtered_docs = [doc for doc in filtered_docs if doc["status_documento"] == filters["engineering_status"]]
    if filters["engineering_revision"] in {"R", "A", "C"}:
        filtered_docs = [
            doc for doc in filtered_docs
            if doc["revisao"].upper().startswith(filters["engineering_revision"])
        ]
    if filters["engineering_q"]:
        query = filters["engineering_q"].lower()
        filtered_docs = [
            doc for doc in filtered_docs
            if query in " ".join([
                doc["codigo"],
                doc["titulo"],
                doc["disciplina"],
                doc["status_documento"],
            ]).lower()
        ]

    engineering_docs = len(filtered_docs)
    filtered_discipline_counts = Counter(doc["disciplina"] for doc in filtered_docs)
    filtered_status_counts = Counter(doc["status_documento"] for doc in filtered_docs)
    revision_counts: Counter[tuple[str, str]] = Counter()
    revision_distinct = set()
    for doc in filtered_docs:
        revision = doc["revisao"]
        if revision and revision != "-":
            revision_distinct.add(revision)
        bucket = _engineering_revision_bucket(revision)
        if bucket:
            revision_counts[(doc["disciplina"], bucket)] += 1

    engineering_summary = []
    for discipline, total in sorted(filtered_discipline_counts.items(), key=lambda pair: (-pair[1], pair[0]))[:14]:
        docs_for_discipline = [doc for doc in filtered_docs if doc["disciplina"] == discipline]
        engineering_summary.append({
            "discipline": discipline,
            "total": total,
            "pct_total": round(100 * total / engineering_docs, 1) if engineering_docs else 0,
            "rev_a": sum(1 for doc in docs_for_discipline if doc["revisao"].upper().startswith("A")),
            "rev_r": sum(1 for doc in docs_for_discipline if doc["revisao"].upper().startswith("R")),
            "rev_c": sum(1 for doc in docs_for_discipline if doc["revisao"].upper().startswith("C")),
            "rev_other": sum(
                1 for doc in docs_for_discipline
                if doc["revisao"] and doc["revisao"] != "-"
                and not doc["revisao"].upper().startswith(("A", "R", "C"))
            ),
            "issued": sum(1 for doc in docs_for_discipline if doc["status_documento"].upper() in {"ISSUED", "EMITIDO"}),
            "in_engineering": sum(1 for doc in docs_for_discipline if doc["status_documento"].upper().startswith("ENGINEERING")),
        })

    engineering_discipline_groups = []
    for discipline, total in sorted(filtered_discipline_counts.items(), key=lambda pair: (-pair[1], pair[0]))[:18]:
        docs_for_discipline = [doc for doc in filtered_docs if doc["disciplina"] == discipline]
        docs_for_discipline.sort(key=lambda doc: (doc["status_documento"], doc["revisao"], doc["codigo"]))
        engineering_discipline_groups.append({
            "label": discipline,
            "value": total,
            "pct": round(100 * total / engineering_docs, 1) if engineering_docs else 0,
            "documents": docs_for_discipline,
        })

    engineering_revision_rows = [
        {"label": discipline, "revision": revision, "total": total}
        for (discipline, revision), total in sorted(revision_counts.items())
    ]
    engineering_status_rows = [
        {"label": label, "value": total}
        for label, total in sorted(filtered_status_counts.items(), key=lambda pair: (-pair[1], pair[0]))[:8]
    ]
    _status_total = sum(row["value"] for row in engineering_status_rows) or 1
    _status_max = max((row["value"] for row in engineering_status_rows), default=0) or 1
    for _row in engineering_status_rows:
        _pct = round(100 * _row["value"] / _status_total, 2)
        _rel = round(100 * _row["value"] / _status_max, 2)
        _row["pct"] = _pct
        _row["pct_css"] = f"{_pct:.2f}"
        _row["relative_pct"] = _rel
        _row["relative_pct_css"] = f"{_rel:.2f}"
    engineering_flow = _engineering_control_summary(filtered_docs, engineering_summary)
    return {
        "source": "ECLIC API",
        "error": "",
        "engineering_docs": engineering_docs,
        "engineering_counts": {
            "disciplines": len(filtered_discipline_counts),
            "revisions": len(revision_distinct),
            "issued": sum(1 for doc in filtered_docs if doc["status_documento"].upper() in {"ISSUED", "EMITIDO"}),
            "in_engineering": sum(1 for doc in filtered_docs if doc["status_documento"].upper().startswith("ENGINEERING")),
        },
        "engineering_flow": engineering_flow,
        "engineering_summary": engineering_summary,
        "engineering_discipline_groups": engineering_discipline_groups,
        "engineering_revision_rows": engineering_revision_rows,
        "engineering_status_rows": engineering_status_rows,
        "engineering_documents": filtered_docs[:24],
        "choices": choices,
    }


def _construction_taskfy(filters: dict) -> dict:
    with _taskfy_conn() as conn:
        cur = conn.cursor()

        sched_where = ["sa.finish between %s and %s"]
        sched_params: list[Any] = [filters["start"], filters["end"]]
        _append_taskfy_filters(filters, "jc", sched_where, sched_params)
        sched_clause = " and ".join(sched_where)

        cur.execute(
            f"""
            select count(distinct sa.id) as activities,
                   count(distinct nullif(sa.jobcard_number, '')) as jobcards,
                   coalesce(sum(sa.hh), 0) as hh,
                   coalesce(sum(sa.points), 0) as points,
                   coalesce(avg(sa.percent_complete), 0) as avg_progress
            from jobcards_scheduleactivity sa
            left join jobcards_jobcard jc on jc.job_card_number = sa.jobcard_number
            where {sched_clause}
            """,
            sched_params,
        )
        planned = dict(cur.fetchone())

        cur.execute(
            f"""
            select to_char(date_trunc('week', sa.finish), 'YYYY-MM-DD') as label,
                   coalesce(sum(sa.hh), 0) as planned_hh,
                   count(distinct nullif(sa.jobcard_number, '')) as jobcards
            from jobcards_scheduleactivity sa
            left join jobcards_jobcard jc on jc.job_card_number = sa.jobcard_number
            where {sched_clause}
            group by date_trunc('week', sa.finish)
            order by date_trunc('week', sa.finish)
            """,
            sched_params,
        )
        planned_week = [dict(r) for r in cur.fetchall()]

        cur.execute(
            f"""
            select sa.activity_id, sa.name, sa.start, sa.finish,
                   coalesce(sa.percent_complete, 0) as percent_complete,
                   coalesce(sa.hh, 0) as hh,
                   sa.jobcard_number
            from jobcards_scheduleactivity sa
            left join jobcards_jobcard jc on jc.job_card_number = sa.jobcard_number
            where {sched_clause}
              and sa.start is not null
              and sa.finish is not null
            order by sa.start, sa.finish, sa.activity_id
            limit 22
            """,
            sched_params,
        )
        gantt_rows = [dict(r) for r in cur.fetchall()]

        prog_where = ["1=1"]
        prog_params: list[Any] = []
        if filters["contract_week"]:
            prog_where.append("p.contract_week = %s")
            prog_params.append(filters["contract_week"])
        if filters["discipline"]:
            prog_where.append("p.discipline_code = %s")
            prog_params.append(filters["discipline"])
        if filters["campaign"]:
            prog_where.append(
                """
                exists (
                    select 1
                    from jobcards_jobcardprogrammingpackitem xi
                    join jobcards_jobcard xj on xj.job_card_number = xi.jobcard_number
                    where xi.pack_id = p.id
                      and coalesce(nullif(xj.campaign, ''), '-') = %s
                )
                """
            )
            prog_params.append(filters["campaign"])
        prog_clause = " and ".join(prog_where)

        cur.execute(
            f"""
            with pack_base as (
                select distinct p.id, p.contract_week, p.discipline_code, p.discipline,
                       p.program_code, p.updated_at
                from jobcards_jobcardprogrammingpack p
                where {prog_clause}
            ),
            pack_rollup as (
                select pb.*,
                       (select count(distinct i.jobcard_number)
                        from jobcards_jobcardprogrammingpackitem i
                        where i.pack_id = pb.id) as jobcards,
                       (select coalesce(sum(t.hh), 0)
                        from jobcards_jobcardprogrammingteam t
                        where t.pack_id = pb.id) as hh
                from pack_base pb
            )
            select count(*) as packs,
                   coalesce(sum(jobcards), 0) as jobcards,
                   coalesce(sum(hh), 0) as hh
            from pack_rollup
            """,
            prog_params,
        )
        programmed = dict(cur.fetchone())

        cur.execute(
            f"""
            with pack_base as (
                select distinct p.id, p.contract_week
                from jobcards_jobcardprogrammingpack p
                where {prog_clause}
            ),
            pack_rollup as (
                select pb.contract_week,
                       (select count(distinct i.jobcard_number)
                        from jobcards_jobcardprogrammingpackitem i
                        where i.pack_id = pb.id) as jobcards,
                       (select coalesce(sum(t.hh), 0)
                        from jobcards_jobcardprogrammingteam t
                        where t.pack_id = pb.id) as hh
                from pack_base pb
            )
            select contract_week as week, coalesce(sum(jobcards), 0) as jobcards,
                   coalesce(sum(hh), 0) as programmed_hh
            from pack_rollup
            group by contract_week
            order by contract_week
            """,
            prog_params,
        )
        programmed_week = [
            {
                "label": _contract_week_label(row["week"]),
                "week": row["week"],
                "programmed_hh": float(row["programmed_hh"] or 0),
                "jobcards": int(row["jobcards"] or 0),
            }
            for row in cur.fetchall()
        ]

        cur.execute(
            f"""
            with pack_base as (
                select distinct p.id, p.contract_week, p.discipline_code, p.discipline,
                       p.program_code, p.updated_at
                from jobcards_jobcardprogrammingpack p
                where {prog_clause}
            )
            select pb.id, pb.program_code, pb.contract_week, pb.discipline_code, pb.discipline,
                   pb.updated_at,
                   (select count(distinct i.jobcard_number)
                    from jobcards_jobcardprogrammingpackitem i
                    where i.pack_id = pb.id) as jobcards,
                   (select coalesce(sum(t.hh), 0)
                    from jobcards_jobcardprogrammingteam t
                    where t.pack_id = pb.id) as hh
            from pack_base pb
            order by pb.contract_week desc, pb.discipline_code, pb.program_code
            limit 12
            """,
            prog_params,
        )
        packs = [dict(r) for r in cur.fetchall()]

        dfr_where = ["h.report_date between %s and %s"]
        dfr_params: list[Any] = [filters["start"], filters["end"]]
        _append_taskfy_filters(filters, "jc", dfr_where, dfr_params)
        dfr_clause = " and ".join(dfr_where)

        cur.execute(
            f"""
            select count(distinct h.dfr_number) as dfrs,
                   count(distinct nullif(h.jobcard_number, '')) as jobcards,
                   coalesce(sum(h.total_hours), 0) as hours,
                   coalesce(sum(h.total_lines), 0) as lines
            from jobcards_dailyfieldreportheader h
            left join jobcards_jobcard jc on jc.job_card_number = h.jobcard_number
            where {dfr_clause}
            """,
            dfr_params,
        )
        dfr = dict(cur.fetchone())

        cur.execute(
            f"""
            select to_char(date_trunc('week', h.report_date), 'YYYY-MM-DD') as label,
                   coalesce(sum(h.total_hours), 0) as actual_hh,
                   count(distinct h.dfr_number) as dfrs
            from jobcards_dailyfieldreportheader h
            left join jobcards_jobcard jc on jc.job_card_number = h.jobcard_number
            where {dfr_clause}
            group by date_trunc('week', h.report_date)
            order by date_trunc('week', h.report_date)
            """,
            dfr_params,
        )
        actual_week = [dict(r) for r in cur.fetchall()]

        cur.execute(
            f"""
            select coalesce(nullif(h.discipline, ''), 'Sem disciplina') as label,
                   count(distinct h.dfr_number) as dfrs,
                   coalesce(sum(h.total_hours), 0) as hours
            from jobcards_dailyfieldreportheader h
            left join jobcards_jobcard jc on jc.job_card_number = h.jobcard_number
            where {dfr_clause}
            group by coalesce(nullif(h.discipline, ''), 'Sem disciplina')
            order by hours desc
            limit 10
            """,
            dfr_params,
        )
        dfr_by_discipline = [dict(r) for r in cur.fetchall()]

        cur.execute(
            f"""
            select h.dfr_number, h.report_date, h.jobcard_number, h.discipline,
                   h.working_code, h.total_hours, h.total_lines, h.notes
            from jobcards_dailyfieldreportheader h
            left join jobcards_jobcard jc on jc.job_card_number = h.jobcard_number
            where {dfr_clause}
            order by h.report_date desc, h.dfr_number desc
            limit 10
            """,
            dfr_params,
        )
        recent_dfrs = [dict(r) for r in cur.fetchall()]

        jc_where = ["jc.jobcard_status <> 'CANCELED'"]
        jc_params: list[Any] = []
        _append_taskfy_filters(filters, "jc", jc_where, jc_params)
        jc_clause = " and ".join(jc_where)
        cur.execute(
            f"""
            select count(*) as total,
                   count(*) filter(where jc.completed = 'YES') as completed,
                   count(*) filter(where jc.completed <> 'YES') as active,
                   count(*) filter(where jc.jobcard_status in ('AFC JOBCARD CHECKED', 'RELEASED FOR EXECUTION', 'JOBCARD FINALIZED')) as released_pool,
                   count(*) filter(where jc.finish between %s and %s and jc.completed <> 'YES') as due_period
            from jobcards_jobcard jc
            where {jc_clause}
            """,
            [filters["start"], filters["end"]] + jc_params,
        )
        jobcards = dict(cur.fetchone())

    engineering = _engineering_from_ded_snapshot(filters) or _engineering_from_eclic_api(filters)
    engineering_monitor = _engineering_monitor_from_snapshot(filters)
    engineering_docs = engineering["engineering_docs"]
    engineering_counts = engineering["engineering_counts"]
    engineering_summary = engineering["engineering_summary"]
    engineering_discipline_groups = engineering["engineering_discipline_groups"]
    engineering_revision_rows = engineering["engineering_revision_rows"]
    engineering_status_rows = engineering["engineering_status_rows"]
    engineering_documents = engineering.get("engineering_documents", [])
    choices = _taskfy_choices()
    choices.update(engineering["choices"])

    labels = _series_labels(planned_week, actual_week, programmed_week)
    planned_curve = _cumulative_from_rows(labels, planned_week, "planned_hh")
    actual_curve = _cumulative_from_rows(labels, actual_week, "actual_hh")
    planned_by_label = {str(r["label"]): float(r.get("planned_hh") or 0) for r in planned_week}
    actual_by_label = {str(r["label"]): float(r.get("actual_hh") or 0) for r in actual_week}
    programmed_by_label = {str(r["label"]): float(r.get("programmed_hh") or 0) for r in programmed_week}

    revision_labels = []
    revision_names = []
    for row in engineering_revision_rows:
        label = str(row["label"])
        revision = str(row["revision"])
        if label not in revision_labels:
            revision_labels.append(label)
        if revision not in revision_names:
            revision_names.append(revision)
    revision_names = [name for name in ("REV. R", "REV. A", "REV. C") if name in revision_names]
    revision_series = []
    for revision in revision_names:
        revision_series.append({
            "name": revision,
            "values": [
                int(next((r["total"] for r in engineering_revision_rows if r["label"] == label and r["revision"] == revision), 0))
                for label in revision_labels
            ],
        })

    programmed_hh = float(programmed.get("hh") or 0)
    actual_hh = float(dfr.get("hours") or 0)
    planned_hh = float(planned.get("hh") or 0)
    programmed_jobcards = int(programmed.get("jobcards") or 0)
    dfr_jobcards = int(dfr.get("jobcards") or 0)

    return {
        "available": True,
        "filters": filters,
        "choices": choices,
        "engineering_source": engineering["source"],
        "engineering_source_mode": engineering.get("source_mode", "live_api"),
        "engineering_import_id": engineering.get("import_id"),
        "engineering_imported_at": engineering.get("imported_at"),
        "engineering_source_file_name": engineering.get("source_file_name", ""),
        "engineering_summary_sheet": engineering.get("summary_sheet", ""),
        "engineering_detail_sheet": engineering.get("detail_sheet", ""),
        "engineering_error": engineering["error"],
        "kpis": {
            "p6_activities": int(planned.get("activities") or 0),
            "p6_jobcards": int(planned.get("jobcards") or 0),
            "p6_hh": planned_hh,
            "programmed_packs": int(programmed.get("packs") or 0),
            "programmed_jobcards": programmed_jobcards,
            "programmed_hh": programmed_hh,
            "dfrs": int(dfr.get("dfrs") or 0),
            "dfr_jobcards": dfr_jobcards,
            "actual_hh": actual_hh,
            "released_pool": int(jobcards.get("released_pool") or 0),
            "completed": int(jobcards.get("completed") or 0),
            "active": int(jobcards.get("active") or 0),
            "due_period": int(jobcards.get("due_period") or 0),
            "engineering_docs": engineering_docs,
            "program_execution_pct": _pct(dfr_jobcards, programmed_jobcards),
            "hh_realized_pct": _pct(actual_hh, programmed_hh),
            "p6_realized_pct": _pct(actual_hh, planned_hh),
        },
        "charts": {
            "s_curve": _line_chart(
                labels,
                [
                    {"name": "P6 planejado HH acumulado", "values": planned_curve, "color": "#737373"},
                    {"name": "DFR realizado HH acumulado", "values": actual_curve, "color": "#0a0a0a"},
                ],
                height=340,
            ),
            "weekly_histogram": _grouped_bar_chart(
                labels,
                [
                    {"name": "P6 HH", "values": [planned_by_label.get(label, 0) for label in labels], "color": "#a3a3a3"},
                    {"name": "Programado HH", "values": [programmed_by_label.get(label, 0) for label in labels], "color": "#0a0a0a"},
                    {"name": "DFR HH", "values": [actual_by_label.get(label, 0) for label in labels], "color": "#b91c1c"},
                ],
                height=300,
            ),
            "gantt": _gantt_chart(gantt_rows),
            "dfr_by_discipline": _bar_chart(
                [str(r["label"]) for r in dfr_by_discipline],
                [float(r["hours"] or 0) for r in dfr_by_discipline],
                color="#0a0a0a",
            ),
            "engineering_revisions": _stacked_horizontal_chart(revision_labels, revision_series, height=360),
            "engineering_status": _bar_chart(
                [str(r["label"]) for r in engineering_status_rows],
                [int(r["value"]) for r in engineering_status_rows],
                color="#525252",
            ),
        },
        "planned_week": planned_week,
        "programmed_week": programmed_week,
        "actual_week": actual_week,
        "gantt_rows": gantt_rows,
        "packs": packs,
        "recent_dfrs": recent_dfrs,
        "dfr_by_discipline": dfr_by_discipline,
        "engineering_counts": engineering_counts,
        "engineering_flow": engineering["engineering_flow"],
        "engineering_summary": engineering_summary,
        "engineering_discipline_groups": engineering_discipline_groups,
        "engineering_revision_rows": engineering_revision_rows,
        "engineering_status_rows": engineering_status_rows,
        "engineering_documents": engineering_documents,
        "engineering_monitor": engineering_monitor,
    }


def _construction_datafy(filters: dict) -> dict:
    with _datafy_conn() as conn:
        cur = conn.cursor()
        min_readiness = int(filters.get("min_readiness") or 0)
        supply_priority = str(filters.get("supply_priority") or "").strip()
        supply_drawing_q = str(filters.get("supply_drawing_q") or "").strip()
        supply_revision = str(filters.get("supply_revision") or "").strip().upper()
        supply_discipline = str(filters.get("supply_discipline") or "").strip()
        supply_line = str(filters.get("supply_line") or "").strip()
        supply_table = str(filters.get("supply_table") or "").strip()
        supply_page = str(filters.get("supply_page") or "").strip()
        supply_item = str(filters.get("supply_item") or "").strip()
        supply_family = str(filters.get("supply_family") or "").strip()
        supply_code_q = str(filters.get("supply_code_q") or "").strip()
        supply_description_q = str(filters.get("supply_description_q") or "").strip()
        campaign_filter = str(filters.get("campaign") or "").strip()
        campaign_aliases = _campaign_aliases(campaign_filter) if campaign_filter else []
        supply_fab_min = int(filters.get("supply_fab_min") or 0)
        supply_erection_min = int(filters.get("supply_erection_min") or 0)
        if supply_discipline:
            cur.execute(
                "select 1 from core_document d where upper(coalesce(nullif(d.discipline, ''), '-')) = upper(?) limit 1",
                (supply_discipline,),
            )
            if cur.fetchone() is None:
                supply_discipline = ""
        drawing_where = []
        drawing_params: list[Any] = []
        if supply_drawing_q:
            query = f"%{supply_drawing_q}%"
            drawing_where.append(
                "(d.drawing_number ilike ? or d.original_filename ilike ? or d.title ilike ?)"
            )
            drawing_params.extend([query, query, query])
        if supply_priority:
            drawing_where.append("cast(d.priority as text) = ?")
            drawing_params.append(supply_priority)
        if supply_revision:
            drawing_where.append("coalesce(nullif(upper(trim(d.revision)), ''), '-') = ?")
            drawing_params.append(supply_revision)
        if supply_discipline:
            drawing_where.append("upper(coalesce(nullif(d.discipline, ''), '-')) = upper(?)")
            drawing_params.append(supply_discipline)
        if campaign_aliases:
            drawing_where.append(
                "coalesce(nullif(d.campaign, ''), '-') in ("
                + ", ".join(["?"] * len(campaign_aliases))
                + ")"
            )
            drawing_params.extend(campaign_aliases)
        drawing_where_sql = f"where {' and '.join(drawing_where)}" if drawing_where else ""

        material_where = []
        material_params: list[Any] = []
        if supply_priority:
            material_where.append("cast(d.priority as text) = ?")
            material_params.append(supply_priority)
        if supply_drawing_q:
            query = f"%{supply_drawing_q}%"
            material_where.append("(d.drawing_number ilike ? or d.original_filename ilike ? or d.title ilike ?)")
            material_params.extend([query, query, query])
        if supply_revision:
            material_where.append("coalesce(nullif(upper(trim(d.revision)), ''), '-') = ?")
            material_params.append(supply_revision)
        if supply_discipline:
            material_where.append("upper(coalesce(nullif(d.discipline, ''), '-')) = upper(?)")
            material_params.append(supply_discipline)
        if supply_line:
            material_where.append("coalesce(nullif(d.piping_line_number, ''), '-') = ?")
            material_params.append(supply_line)
        if supply_table:
            material_where.append("coalesce(nullif(t.name, ''), '-') = ?")
            material_params.append(supply_table)
        if supply_page:
            material_where.append("cast(t.page_number as text) = ?")
            material_params.append(supply_page)
        if supply_item:
            query = f"%{supply_item}%"
            material_where.append("(mi.item_number ilike ? or cast(mi.row_order as text) ilike ?)")
            material_params.extend([query, query])
        if supply_family:
            material_where.append("coalesce(nullif(mf.name_en, ''), nullif(mi.category, ''), '-') = ?")
            material_params.append(supply_family)
        if supply_code_q:
            query = f"%{supply_code_q}%"
            material_where.append("(ci.pmto_code ilike ? or mi.material_code ilike ?)")
            material_params.extend([query, query])
        if supply_description_q:
            query = f"%{supply_description_q}%"
            material_where.append("mi.description ilike ?")
            material_params.append(query)
        if campaign_aliases:
            material_where.append(
                "coalesce(nullif(d.campaign, ''), '-') in ("
                + ", ".join(["?"] * len(campaign_aliases))
                + ")"
            )
            material_params.extend(campaign_aliases)
        material_where_sql = f"where {' and '.join(material_where)}" if material_where else ""
        finalized_doc_sql = """
            (
                coalesce(d.field_complete, false)
                or coalesce(d.allocation_complete, false)
                or upper(
                    coalesce(d.status, '') || ' ' ||
                    coalesce(d.status_message, '') || ' ' ||
                    coalesce(d.revision_detail, '') || ' ' ||
                    coalesce(d.title, '')
                ) like '%NLA%'
                or upper(
                    coalesce(d.status, '') || ' ' ||
                    coalesce(d.status_message, '') || ' ' ||
                    coalesce(d.revision_detail, '') || ' ' ||
                    coalesce(d.title, '')
                ) like '%FINALIZ%'
                or upper(
                    coalesce(d.status, '') || ' ' ||
                    coalesce(d.status_message, '') || ' ' ||
                    coalesce(d.revision_detail, '') || ' ' ||
                    coalesce(d.title, '')
                ) like '%FINALIZED%'
                or upper(
                    coalesce(d.status, '') || ' ' ||
                    coalesce(d.status_message, '') || ' ' ||
                    coalesce(d.revision_detail, '') || ' ' ||
                    coalesce(d.title, '')
                ) like '%NO LONGER APPLICABLE%'
                or upper(
                    coalesce(d.status, '') || ' ' ||
                    coalesce(d.status_message, '') || ' ' ||
                    coalesce(d.revision_detail, '') || ' ' ||
                    coalesce(d.title, '')
                ) like '%NOT APPLICABLE%'
                or upper(
                    coalesce(d.status, '') || ' ' ||
                    coalesce(d.status_message, '') || ' ' ||
                    coalesce(d.revision_detail, '') || ' ' ||
                    coalesce(d.title, '')
                ) like '%NAO APLICA%'
                or upper(
                    coalesce(d.status, '') || ' ' ||
                    coalesce(d.status_message, '') || ' ' ||
                    coalesce(d.revision_detail, '') || ' ' ||
                    coalesce(d.title, '')
                ) like '%NÃO APLICA%'
            )
        """
        active_drawing_where = [*drawing_where, f"not {finalized_doc_sql}"]
        active_material_where = [*material_where, f"not {finalized_doc_sql}"]
        active_drawing_where_sql = f"where {' and '.join(active_drawing_where)}"
        active_material_where_sql = f"where {' and '.join(active_material_where)}"

        line_rows = _rows(
            cur,
            """
            select d.piping_line_number as line,
                   count(distinct mi.id) as materials,
                   count(distinct a.material_item_id) as covered,
                   round(100.0 * count(distinct a.material_item_id) / count(distinct mi.id), 1) as coverage_pct,
                   max(d.revision) as revision,
                   string_agg(distinct d.drawing_number, ', ') as drawings
            from core_document d
            join core_extractedtable t on t.document_id = d.id
            join core_materialitem mi on mi.table_id = t.id
            left join catalog_allocation a on a.material_item_id = mi.id
            where d.piping_line_number <> ''
            group by d.piping_line_number
            having count(distinct mi.id) >= 5
               and (100.0 * count(distinct a.material_item_id) / count(distinct mi.id)) >= ?
            order by coverage_pct desc, materials desc
            limit 14
            """,
            (min_readiness,),
        )
        readiness = _rows(
            cur,
            """
            select count(*) as lines,
                   sum(case when coverage_pct = 100 then 1 else 0 end) as ready_100,
                   sum(case when coverage_pct >= 80 then 1 else 0 end) as ready_80,
                   sum(case when coverage_pct >= 50 then 1 else 0 end) as ready_50
            from (
                select d.piping_line_number,
                       100.0 * count(distinct a.material_item_id) / count(distinct mi.id) as coverage_pct
                from core_document d
                join core_extractedtable t on t.document_id = d.id
                join core_materialitem mi on mi.table_id = t.id
                left join catalog_allocation a on a.material_item_id = mi.id
                where d.piping_line_number <> ''
                group by d.piping_line_number
            ) line_coverage
            """,
        )[0]
        histogram = _rows(
            cur,
            """
            select bucket as label, count(*) as value
            from (
                select case
                    when coverage_pct = 100 then '100%'
                    when coverage_pct >= 80 then '80-99%'
                    when coverage_pct >= 50 then '50-79%'
                    when coverage_pct >= 25 then '25-49%'
                    when coverage_pct > 0 then '1-24%'
                    else '0%'
                end as bucket
                from (
                    select d.piping_line_number,
                           100.0 * count(distinct a.material_item_id) / count(distinct mi.id) as coverage_pct
                    from core_document d
                    join core_extractedtable t on t.document_id = d.id
                    join core_materialitem mi on mi.table_id = t.id
                    left join catalog_allocation a on a.material_item_id = mi.id
                    where d.piping_line_number <> ''
                    group by d.piping_line_number
                ) line_coverage
            ) readiness_buckets
            group by bucket
            order by case bucket
                when '100%' then 6
                when '80-99%' then 5
                when '50-79%' then 4
                when '25-49%' then 3
                when '1-24%' then 2
                else 1
            end desc
            """,
        )
        docs_by_revision = _rows(
            cur,
            """
            select coalesce(nullif(discipline, ''), 'Sem disciplina') as discipline,
                   coalesce(nullif(revision, ''), 'Sem rev') as revision,
                   count(*) as total
            from core_document
            group by coalesce(nullif(discipline, ''), 'Sem disciplina'),
                     coalesce(nullif(revision, ''), 'Sem rev')
            order by discipline, revision
            """,
        )
        drawing_revisions = _rows(
            cur,
            """
            select coalesce(nullif(upper(trim(d.revision)), ''), '-') as value,
                   coalesce(nullif(upper(trim(d.revision)), ''), '-') as label,
                   count(distinct d.id) as total
            from core_document d
            join core_extractedtable t on t.document_id = d.id
            join core_materialitem mi on mi.table_id = t.id
            group by coalesce(nullif(upper(trim(d.revision)), ''), '-')
            order by value
            """,
        )
        drawing_priorities = _rows(
            cur,
            """
            select cast(d.priority as text) as value,
                   cast(d.priority as text) as label,
                   count(distinct d.id) as total
            from core_document d
            join core_extractedtable t on t.document_id = d.id
            join core_materialitem mi on mi.table_id = t.id
            where d.priority is not null
            group by d.priority
            order by d.priority asc
            """,
        )
        material_disciplines = _rows(
            cur,
            """
            select coalesce(nullif(d.discipline, ''), '-') as value,
                   coalesce(nullif(d.discipline, ''), '-') as label,
                   count(*) as total
            from core_document d
            join core_extractedtable t on t.document_id = d.id
            join core_materialitem mi on mi.table_id = t.id
            group by coalesce(nullif(d.discipline, ''), '-')
            order by count(*) desc, value
            """,
        )
        material_lines = _rows(
            cur,
            """
            select coalesce(nullif(d.piping_line_number, ''), '-') as value,
                   coalesce(nullif(d.piping_line_number, ''), '-') as label,
                   count(*) as total
            from core_document d
            join core_extractedtable t on t.document_id = d.id
            join core_materialitem mi on mi.table_id = t.id
            group by coalesce(nullif(d.piping_line_number, ''), '-')
            order by count(*) desc, value
            limit 160
            """,
        )
        material_tables = _rows(
            cur,
            """
            select coalesce(nullif(t.name, ''), '-') as value,
                   coalesce(nullif(t.name, ''), '-') as label,
                   count(*) as total
            from core_extractedtable t
            join core_materialitem mi on mi.table_id = t.id
            group by coalesce(nullif(t.name, ''), '-')
            order by count(*) desc, value
            limit 80
            """,
        )
        material_families = _rows(
            cur,
            """
            with match_one as (
                select material_item_id, min(catalog_item_id) as catalog_item_id
                from catalog_catalogmatch
                where material_item_id is not null
                group by material_item_id
            )
            select coalesce(nullif(mf.name_en, ''), nullif(mi.category, ''), '-') as value,
                   coalesce(nullif(mf.name_en, ''), nullif(mi.category, ''), '-') as label,
                   count(*) as total
            from core_materialitem mi
            left join match_one cm on cm.material_item_id = mi.id
            left join catalog_catalogitem ci on ci.id = cm.catalog_item_id
            left join catalog_materialfamily mf on mf.id = ci.family_id
            group by coalesce(nullif(mf.name_en, ''), nullif(mi.category, ''), '-')
            order by count(*) desc, value
            limit 120
            """,
        )
        drawing_readiness = _rows(
            cur,
            f"""
            with doc_base as (
                select d.id, d.drawing_number, d.original_filename, d.title,
                       d.revision, d.discipline, d.status, d.priority, d.uploaded_at
                from core_document d
                {active_drawing_where_sql}
            ),
            doc_tables as (
                select db.id as document_id,
                       max(case
                         when (upper(coalesce(t.name, '')) like '%FABRICATION%'
                               or upper(coalesce(t.name, '')) like '%FABRICAC%')
                          and not (upper(coalesce(t.name, '')) like '%DEMOLISH%'
                                   or upper(coalesce(t.name, '')) like '%DEMOLI%'
                                   or upper(coalesce(t.name, '')) like '%REMOV%')
                         then 1 else 0 end) as has_explicit_fab
                from doc_base db
                left join core_extractedtable t on t.document_id = db.id
                group by db.id
            ),
            scoped_items as (
                select db.id, db.drawing_number, db.original_filename, db.title,
                       db.revision, db.discipline, db.status, db.priority,
                       db.uploaded_at, mi.id as material_item_id,
                       case
                         when upper(coalesce(t.name, '')) like '%DEMOLISH%'
                           or upper(coalesce(t.name, '')) like '%DEMOLI%'
                           or upper(coalesce(t.name, '')) like '%REMOV%'
                         then 'other'
                         when upper(coalesce(t.name, '')) like '%ERECTION%'
                           or upper(coalesce(t.name, '')) like '%INSTALLATION%'
                           or upper(coalesce(t.name, '')) like '%ONBOARD%'
                         then 'erection'
                         when upper(coalesce(t.name, '')) like '%FABRICATION%'
                           or upper(coalesce(t.name, '')) like '%FABRICAC%'
                         then 'fabrication'
                         when coalesce(dt.has_explicit_fab, 0) = 0
                         then 'fabrication'
                         else 'other'
                       end as scope
                from doc_base db
                join core_extractedtable t on t.document_id = db.id
                join core_materialitem mi on mi.table_id = t.id
                left join doc_tables dt on dt.document_id = db.id
            ),
            rollup as (
                select s.id, s.drawing_number, s.original_filename, s.title,
                       s.revision, s.discipline, s.status, s.priority,
                       count(distinct case when s.scope = 'fabrication' then s.material_item_id end) as fab_total,
                       count(distinct case when s.scope = 'fabrication' then a.material_item_id end) as fab_covered,
                       case
                         when count(distinct case when s.scope = 'fabrication' then s.material_item_id end) > 0
                         then round(
                           100.0 * count(distinct case when s.scope = 'fabrication' then a.material_item_id end)
                           / count(distinct case when s.scope = 'fabrication' then s.material_item_id end),
                           1
                         )
                         else 0
                       end as fab_pct,
                       count(distinct case when s.scope = 'erection' then s.material_item_id end) as erection_total,
                       count(distinct case when s.scope = 'erection' then a.material_item_id end) as erection_covered,
                       case
                         when count(distinct case when s.scope = 'erection' then s.material_item_id end) > 0
                         then round(
                           100.0 * count(distinct case when s.scope = 'erection' then a.material_item_id end)
                           / count(distinct case when s.scope = 'erection' then s.material_item_id end),
                           1
                         )
                         else 0
                       end as erection_pct
                from scoped_items s
                left join catalog_allocation a on a.material_item_id = s.material_item_id
                group by s.id, s.drawing_number, s.original_filename, s.title,
                         s.revision, s.discipline, s.status, s.priority
            )
            select *
            from rollup
            where (fab_total > 0 or erection_total > 0)
              and (case when fab_pct >= erection_pct then fab_pct else erection_pct end) >= ?
              and fab_pct >= ?
              and erection_pct >= ?
            order by case when priority is null then 1 else 0 end,
                     priority asc,
                     case when fab_pct >= erection_pct then fab_pct else erection_pct end desc,
                     drawing_number
            """,
            tuple(drawing_params + [min_readiness, supply_fab_min, supply_erection_min]),
        )
        material_rows = _rows(
            cur,
            f"""
            with match_one as (
                select material_item_id, min(catalog_item_id) as catalog_item_id
                from catalog_catalogmatch
                where material_item_id is not null
                group by material_item_id
            ),
            doc_tables as (
                select d.id as document_id,
                       max(case
                         when (upper(coalesce(t.name, '')) like '%FABRICATION%'
                               or upper(coalesce(t.name, '')) like '%FABRICAC%')
                          and not (upper(coalesce(t.name, '')) like '%DEMOLISH%'
                                   or upper(coalesce(t.name, '')) like '%DEMOLI%'
                                   or upper(coalesce(t.name, '')) like '%REMOV%')
                         then 1 else 0 end) as has_explicit_fab
                from core_document d
                left join core_extractedtable t on t.document_id = d.id
                group by d.id
            ),
            alloc as (
                select a.material_item_id,
                       sum(a.qty_allocated) as allocated_qty,
                       string_agg(distinct nullif(po.po_number, ''), ', ') as po_covering,
                       min(po.procurement_plan_date) filter (where po.procurement_plan_date is not null) as po_expected_date,
                       string_agg(distinct to_char(po.procurement_plan_date, 'YYYY-MM-DD'), ', ')
                         filter (where po.procurement_plan_date is not null) as po_expected_dates,
                       string_agg(
                         distinct concat(nullif(po.po_number, ''), '::', coalesce(to_char(po.procurement_plan_date, 'YYYY-MM-DD'), '')),
                         ';;'
                       ) filter (where nullif(po.po_number, '') is not null) as po_delivery_pairs
                from catalog_allocation a
                left join catalog_stockpiece sp on sp.id = a.stock_piece_id
                left join core_purchaseorderitem poi on poi.id = sp.po_item_id
                left join core_purchaseorder po on po.id = poi.purchase_order_id
                group by a.material_item_id
            ),
            stock as (
                select catalog_item_id, sum(remaining_qty) as stock_free_qty
                from catalog_stockpiece
                group by catalog_item_id
            )
            select d.priority,
                   d.id as document_id,
                   mi.id as material_item_id,
                   d.drawing_number,
                   d.original_filename,
                   d.revision,
                   coalesce(nullif(d.campaign, ''), '-') as campaign,
                   coalesce(nullif(d.revision_detail, ''), '-') as revision_detail,
                   coalesce(nullif(d.discipline, ''), '-') as discipline,
                   coalesce(nullif(d.piping_line_number, ''), '-') as line,
                   coalesce(nullif(t.name, ''), '-') as table_name,
                   case
                     when upper(coalesce(t.name, '')) like '%DEMOLISH%'
                       or upper(coalesce(t.name, '')) like '%DEMOLI%'
                       or upper(coalesce(t.name, '')) like '%REMOV%'
                     then 'other'
                     when upper(coalesce(t.name, '')) like '%ERECTION%'
                       or upper(coalesce(t.name, '')) like '%INSTALLATION%'
                       or upper(coalesce(t.name, '')) like '%ONBOARD%'
                     then 'erection'
                     when upper(coalesce(t.name, '')) like '%FABRICATION%'
                       or upper(coalesce(t.name, '')) like '%FABRICAC%'
                     then 'fabrication'
                     when coalesce(dt.has_explicit_fab, 0) = 0
                     then 'fabrication'
                     else 'other'
                   end as scope,
                   t.page_number,
                   coalesce(nullif(mi.item_number, ''), cast(mi.row_order as text)) as item_number,
                   coalesce(nullif(mf.name_en, ''), nullif(mi.category, ''), '-') as family,
                   coalesce(nullif(ci.pmto_code, ''), nullif(mi.material_code, ''), '-') as cpmtocode,
                   coalesce(nullif(mi.description, ''), nullif(ci.canonical_description, ''), '-') as description,
                   coalesce(mi.quantity, 0) as requested_qty,
                   coalesce(nullif(mi.unit, ''), nullif(ci.canonical_unit, ''), '') as unit,
                   coalesce(alloc.allocated_qty, 0) as allocated_qty,
                   greatest(coalesce(mi.quantity, 0) - coalesce(alloc.allocated_qty, 0), 0) as missing_qty,
                   case when ci.id is null then 1 else 0 end as stock_free_na,
                   case
                       when ci.id is null then null
                       else coalesce(stock.stock_free_qty, 0)
                   end as stock_free_qty,
                   coalesce(nullif(alloc.po_covering, ''), '') as po_covering,
                   alloc.po_expected_date,
                   coalesce(nullif(alloc.po_expected_dates, ''), '') as po_expected_dates,
                   coalesce(nullif(alloc.po_delivery_pairs, ''), '') as po_delivery_pairs
            from core_document d
            join core_extractedtable t on t.document_id = d.id
            join core_materialitem mi on mi.table_id = t.id
            left join match_one cm on cm.material_item_id = mi.id
            left join catalog_catalogitem ci on ci.id = cm.catalog_item_id
            left join catalog_materialfamily mf on mf.id = ci.family_id
            left join alloc on alloc.material_item_id = mi.id
            left join stock on stock.catalog_item_id = ci.id
            left join doc_tables dt on dt.document_id = d.id
            {active_material_where_sql}
            order by case when d.priority is null then 1 else 0 end,
                     d.priority asc,
                     d.uploaded_at desc,
                     d.drawing_number,
                     t.page_number,
                     t."order",
                     mi.row_order
            """,
            tuple(material_params),
        )
        material_filter_cte = f"""
            with match_one as (
                select material_item_id, min(catalog_item_id) as catalog_item_id
                from catalog_catalogmatch
                where material_item_id is not null
                group by material_item_id
            ),
            filtered_materials as (
                select distinct mi.id as material_item_id
                from core_document d
                join core_extractedtable t on t.document_id = d.id
                join core_materialitem mi on mi.table_id = t.id
                left join match_one cm on cm.material_item_id = mi.id
                left join catalog_catalogitem ci on ci.id = cm.catalog_item_id
                left join catalog_materialfamily mf on mf.id = ci.family_id
                {active_material_where_sql}
            )
        """
        material_flow_base = _rows(
            cur,
            f"""
            {material_filter_cte}
            select count(distinct fm.material_item_id) as total,
                   count(distinct a.material_item_id) as covered
            from filtered_materials fm
            left join catalog_allocation a on a.material_item_id = fm.material_item_id
            """,
            tuple(material_params),
        )[0]
        material_flow_po_rows = _rows(
            cur,
            f"""
            {material_filter_cte}
            select distinct fm.material_item_id,
                   po.procurement_plan_stage,
                   po.procurement_plan_kind,
                   po.procurement_plan_date,
                   po.procurement_plan_payload
            from filtered_materials fm
            join catalog_allocation a on a.material_item_id = fm.material_item_id
            join catalog_stockpiece sp on sp.id = a.stock_piece_id
            join core_purchaseorderitem poi on poi.id = sp.po_item_id
            join core_purchaseorder po on po.id = poi.purchase_order_id
            """,
            tuple(material_params),
        )
        material_at_aveon = len({
            row["material_item_id"]
            for row in material_flow_po_rows
            if _po_row_has_yard_actual(row)
        })
        material_flow = _material_flow_summary(
            int(material_flow_base.get("total") or 0),
            int(material_flow_base.get("covered") or 0),
            material_at_aveon,
        )
        material_scope_cte = f"""
            with match_one as (
                select material_item_id, min(catalog_item_id) as catalog_item_id
                from catalog_catalogmatch
                where material_item_id is not null
                group by material_item_id
            ),
            alloc as (
                select a.material_item_id,
                       sum(a.qty_allocated) as allocated_qty
                from catalog_allocation a
                group by a.material_item_id
            ),
            doc_tables as (
                select d.id as document_id,
                       max(case
                         when (upper(coalesce(t.name, '')) like '%FABRICATION%'
                               or upper(coalesce(t.name, '')) like '%FABRICAC%')
                          and not (upper(coalesce(t.name, '')) like '%DEMOLISH%'
                                   or upper(coalesce(t.name, '')) like '%DEMOLI%'
                                   or upper(coalesce(t.name, '')) like '%REMOV%')
                         then 1 else 0 end) as has_explicit_fab
                from core_document d
                left join core_extractedtable t on t.document_id = d.id
                group by d.id
            ),
            scoped_items as (
                select distinct mi.id as material_item_id,
                       d.id as document_id,
                       d.priority,
                       d.drawing_number,
                       d.original_filename,
                       coalesce(nullif(d.revision, ''), '-') as revision,
                       coalesce(nullif(d.revision_detail, ''), '-') as revision_detail,
                       coalesce(nullif(d.discipline, ''), '-') as discipline,
                       coalesce(nullif(d.piping_line_number, ''), '-') as line,
                       coalesce(nullif(d.campaign, ''), '-') as campaign,
                       coalesce(mi.quantity, 0) as requested_qty,
                       coalesce(alloc.allocated_qty, 0) as allocated_qty,
                       greatest(coalesce(mi.quantity, 0) - coalesce(alloc.allocated_qty, 0), 0) as missing_qty,
                       case when {finalized_doc_sql} then 1 else 0 end as is_finalized,
                       case
                         when upper(coalesce(t.name, '')) like '%DEMOLISH%'
                           or upper(coalesce(t.name, '')) like '%DEMOLI%'
                           or upper(coalesce(t.name, '')) like '%REMOV%'
                         then 'other'
                         when upper(coalesce(t.name, '')) like '%ERECTION%'
                           or upper(coalesce(t.name, '')) like '%INSTALLATION%'
                           or upper(coalesce(t.name, '')) like '%ONBOARD%'
                         then 'erection'
                         when upper(coalesce(t.name, '')) like '%FABRICATION%'
                           or upper(coalesce(t.name, '')) like '%FABRICAC%'
                         then 'fabrication'
                         when coalesce(dt.has_explicit_fab, 0) = 0
                         then 'fabrication'
                         else 'other'
                       end as scope
                from core_document d
                join core_extractedtable t on t.document_id = d.id
                join core_materialitem mi on mi.table_id = t.id
                left join match_one cm on cm.material_item_id = mi.id
                left join catalog_catalogitem ci on ci.id = cm.catalog_item_id
                left join catalog_materialfamily mf on mf.id = ci.family_id
                left join alloc on alloc.material_item_id = mi.id
                left join doc_tables dt on dt.document_id = d.id
                {material_where_sql}
            )
        """
        material_scope_items = _rows(
            cur,
            f"""
            {material_scope_cte}
            select material_item_id, document_id, priority, drawing_number, original_filename,
                   revision, revision_detail, discipline, line, campaign, scope, is_finalized,
                   requested_qty, allocated_qty, missing_qty
            from scoped_items
            where scope in ('fabrication', 'erection')
            """,
            tuple(material_params),
        )
        material_scope_po_rows = _rows(
            cur,
            f"""
            {material_scope_cte}
            select distinct si.material_item_id,
                   po.po_number,
                   po.procurement_plan_stage,
                   po.procurement_plan_kind,
                   po.procurement_plan_date,
                   po.procurement_plan_payload
            from scoped_items si
            join catalog_allocation a on a.material_item_id = si.material_item_id
            join catalog_stockpiece sp on sp.id = a.stock_piece_id
            join core_purchaseorderitem poi on poi.id = sp.po_item_id
            join core_purchaseorder po on po.id = poi.purchase_order_id
            where si.scope in ('fabrication', 'erection')
            """,
            tuple(material_params),
        )
        material_yard_ids = {
            row["material_item_id"]
            for row in material_scope_po_rows
            if _po_row_has_yard_actual(row)
        }
        material_po_ids = {row["material_item_id"] for row in material_scope_po_rows}
        material_delivery_by_id: dict[Any, str] = {}
        material_po_numbers_by_id: dict[Any, set[str]] = {}
        material_delivery_pairs_by_id: dict[Any, list[dict[str, str]]] = {}
        for po_row in material_scope_po_rows:
            material_id = po_row.get("material_item_id")
            if material_id is None:
                continue
            po_number = str(po_row.get("po_number") or "").strip()
            if po_number:
                material_po_numbers_by_id.setdefault(material_id, set()).add(po_number)
            delivery_date = _po_delivery_date_iso(po_row)
            if delivery_date and delivery_date > material_delivery_by_id.get(material_id, ""):
                material_delivery_by_id[material_id] = delivery_date
            if po_number and delivery_date:
                material_delivery_pairs_by_id.setdefault(material_id, []).append({
                    "po": po_number,
                    "expected_date": delivery_date,
                })
        for row in material_scope_items:
            material_id = row.get("material_item_id")
            row["has_po"] = 1 if material_id in material_po_ids else 0
            row["yard_actual"] = 1 if material_id in material_yard_ids else 0
            row["po_delivery_date"] = material_delivery_by_id.get(material_id, "")
            row["po_numbers"] = ", ".join(sorted(material_po_numbers_by_id.get(material_id, set())))
            row["po_delivery_pairs"] = material_delivery_pairs_by_id.get(material_id, [])
        material_status_counts = _supply_enrich_material_rows(material_rows, material_yard_ids, material_po_ids)
        drawing_line_rows = _supply_build_drawing_line_rows(material_scope_items)
        _supply_enrich_drawing_line_families(drawing_line_rows, material_rows)
        supply_campaign_views = _supply_campaign_views(
            material_scope_items,
            material_scope_po_rows,
        )
        supply_forecast_items = [
            {
                "document_id": item.get("document_id"),
                "drawing": item.get("drawing_number") or item.get("original_filename") or "-",
                "scope": item.get("scope") or "",
                "campaign": _campaign_label(item.get("campaign")),
                "priority": item.get("priority"),
                "material_item_id": item.get("material_item_id"),
                "line": item.get("line") or "-",
                "is_finalized": 1 if _truthy_flag(item.get("is_finalized")) else 0,
                "requested_qty": float(item.get("requested_qty") or 0),
                "allocated_qty": float(item.get("allocated_qty") or 0),
                "missing_qty": float(item.get("missing_qty") or 0),
                "has_po": 1 if _truthy_flag(item.get("has_po")) else 0,
                "yard_actual": 1 if _truthy_flag(item.get("yard_actual")) else 0,
                "po_delivery_date": item.get("po_delivery_date") or "",
                "po_numbers": item.get("po_numbers") or "",
                "po_delivery_pairs": item.get("po_delivery_pairs") or [],
            }
            for item in material_scope_items
            if item.get("scope") in {"fabrication", "erection"}
        ]
        supply_documents = _rows(
            cur,
            """
            select d.id, d.drawing_number, d.original_filename, d.title, d.revision,
                   d.revision_detail, d.discipline, d.status, d.priority,
                   d.piping_line_number,
                   count(distinct t.id) as tables_count,
                   count(distinct mi.id) as material_count,
                   count(distinct a.material_item_id) as covered_materials,
                   case
                     when count(distinct mi.id) > 0
                     then round(100.0 * count(distinct a.material_item_id) / count(distinct mi.id), 1)
                     else 0
                   end as coverage_pct
            from core_document d
            left join core_extractedtable t on t.document_id = d.id
            left join core_materialitem mi on mi.table_id = t.id
            left join catalog_allocation a on a.material_item_id = mi.id
            group by d.id
            order by d.priority asc, d.uploaded_at desc
            limit 18
            """,
        )

    return {
        "available": True,
        "kpis": {
            "piping_lines": int(readiness["lines"] or 0),
            "ready_100": int(readiness["ready_100"] or 0),
            "ready_80": int(readiness["ready_80"] or 0),
            "ready_50": int(readiness["ready_50"] or 0),
        },
        "charts": {
            "line_readiness": _bar_chart(
                [{
                    "100%": "100% cobertas",
                    "80-99%": "80-99% cobertas",
                    "50-79%": "50-79% cobertas",
                    "25-49%": "25-49% cobertas",
                    "1-24%": "1-24% cobertas",
                    "0%": "0% cobertas",
                }.get(str(r["label"]), str(r["label"])) for r in histogram],
                [int(r["value"]) for r in histogram],
                color="#0a0a0a",
            ),
        },
        "line_candidates": line_rows,
        "drawing_readiness": drawing_readiness,
        "drawing_priorities": drawing_priorities,
        "drawing_revisions": drawing_revisions,
        "material_flow": material_flow,
        "supply_campaign_views": supply_campaign_views,
        "supply_forecast_items_json": json.dumps(supply_forecast_items, ensure_ascii=False),
        "material_rows": material_rows,
        "material_scope_items": material_scope_items,
        "material_status_counts": {
            "covered": material_status_counts.get("ok", 0),
            "partial": material_status_counts.get("partial", 0),
            "missing": material_status_counts.get("missing", 0),
            "unknown": material_status_counts.get("unknown", 0),
        },
        "drawing_line_rows": drawing_line_rows,
        "material_disciplines": material_disciplines,
        "material_lines": material_lines,
        "material_tables": material_tables,
        "material_families": material_families,
        "supply_documents": supply_documents,
        "line_histogram": histogram,
        "docs_by_revision": docs_by_revision,
    }


def _construction_datafy_empty(message: str = "No active DATAFY SQLite snapshot.") -> dict:
    return {
        "available": True,
        "source_mode": "sqlite_snapshot_missing",
        "snapshot_missing": True,
        "snapshot_error": message,
        "kpis": {
            "piping_lines": 0,
            "ready_100": 0,
            "ready_80": 0,
            "ready_50": 0,
        },
        "charts": {
            "line_readiness": _empty_chart(message),
        },
        "line_candidates": [],
        "drawing_readiness": [],
        "drawing_priorities": [],
        "drawing_revisions": [],
        "material_flow": _material_flow_summary(0, 0, 0),
        "supply_campaign_views": [],
        "supply_forecast_items_json": "[]",
        "material_rows": [],
        "material_scope_items": [],
        "material_status_counts": {
            "covered": 0,
            "partial": 0,
            "missing": 0,
            "unknown": 0,
        },
        "drawing_line_rows": [],
        "material_disciplines": [],
        "material_lines": [],
        "material_tables": [],
        "material_families": [],
        "supply_documents": [],
        "line_histogram": [],
        "docs_by_revision": [],
    }


def _construction_datafy_snapshot(filters: dict, *, allow_live: bool | None = None) -> dict:
    if allow_live is None:
        allow_live = not settings.DASHFY_SQLITE_ONLY
    requested_snapshot_filters = normalized_supply_snapshot_filters(filters)
    try:
        snapshot = (
            DatafySupplySnapshot.objects
            .filter(is_active=True, filters_hash=_supply_filters_hash(filters))
            .order_by("-created_at", "-id")
            .first()
        )
        if snapshot is None:
            snapshots = (
                DatafySupplySnapshot.objects
                .filter(is_active=True)
                .order_by("-created_at", "-id")
            )
            snapshot = next(
                (
                    candidate
                    for candidate in snapshots
                    if normalized_supply_snapshot_filters(candidate.filters) == requested_snapshot_filters
                ),
                None,
            )
    except (OperationalError, ProgrammingError):
        snapshot = None

    if snapshot:
        payload = json.loads(json.dumps(snapshot.payload, default=json_default, ensure_ascii=False))
        _supply_normalize_operational_payload(payload)
        payload["source_mode"] = "sqlite_snapshot"
        payload["snapshot_id"] = snapshot.pk
        payload["snapshot_refreshed_at"] = snapshot.created_at
        payload["source_database"] = snapshot.source_database
        payload["source_host"] = snapshot.source_host
        return payload

    if not allow_live:
        return _construction_datafy_empty()

    payload = _construction_datafy(filters)
    payload["source_mode"] = "postgres_live"
    payload["snapshot_missing"] = True
    return payload


def _construction_sqlite_snapshot(filters: dict) -> dict:
    engineering = _engineering_from_ded_snapshot(filters)
    if not engineering:
        engineering = _engineering_empty("No active DED SQLite snapshot.")
        engineering["source"] = "SQLite DED snapshot"
        engineering["source_mode"] = "sqlite_snapshot_missing"
    engineering_monitor = _engineering_monitor_from_snapshot(filters)
    choices = {
        "disciplines": [],
        "campaigns": [],
        "weeks": [],
        "engineering_disciplines": [],
        "engineering_statuses": [],
        "engineering_issue_statuses": [],
        "engineering_responsibles": [],
        "engineering_revisions": [
            {"value": "R", "label": "REV. R"},
            {"value": "A", "label": "REV. A"},
            {"value": "C", "label": "REV. C"},
        ],
    }
    choices.update(engineering.get("choices") or {})
    return {
        "available": True,
        "source_mode": "sqlite_only",
        "filters": filters,
        "choices": choices,
        "engineering_source": engineering.get("source", "SQLite DED snapshot"),
        "engineering_source_mode": engineering.get("source_mode", "sqlite_snapshot"),
        "engineering_import_id": engineering.get("import_id"),
        "engineering_imported_at": engineering.get("imported_at"),
        "engineering_source_file_name": engineering.get("source_file_name", ""),
        "engineering_summary_sheet": engineering.get("summary_sheet", ""),
        "engineering_detail_sheet": engineering.get("detail_sheet", ""),
        "engineering_error": engineering.get("error", ""),
        "kpis": {
            "p6_activities": 0,
            "p6_jobcards": 0,
            "p6_hh": 0,
            "programmed_packs": 0,
            "programmed_jobcards": 0,
            "programmed_hh": 0,
            "dfrs": 0,
            "dfr_jobcards": 0,
            "actual_hh": 0,
            "released_pool": 0,
            "completed": 0,
            "active": 0,
            "due_period": 0,
            "engineering_docs": engineering.get("engineering_docs", 0),
            "program_execution_pct": 0,
            "hh_realized_pct": 0,
            "p6_realized_pct": 0,
        },
        "charts": {
            "s_curve": _empty_chart("Taskfy disabled in SQLite-only mode"),
            "weekly_histogram": _empty_chart("Taskfy disabled in SQLite-only mode"),
            "gantt": _empty_chart("Taskfy disabled in SQLite-only mode"),
            "dfr_by_discipline": _empty_chart("Taskfy disabled in SQLite-only mode"),
            "engineering_revisions": _empty_chart(),
            "engineering_status": _empty_chart(),
        },
        "planned_week": [],
        "actual_week": [],
        "gantt_rows": [],
        "packs": [],
        "recent_dfrs": [],
        "dfr_by_discipline": [],
        "engineering_counts": engineering.get("engineering_counts", {}),
        "engineering_flow": engineering.get("engineering_flow", {}),
        "engineering_summary": engineering.get("engineering_summary", []),
        "engineering_discipline_groups": engineering.get("engineering_discipline_groups", []),
        "engineering_revision_rows": engineering.get("engineering_revision_rows", []),
        "engineering_status_rows": engineering.get("engineering_status_rows", []),
        "engineering_documents": engineering.get("engineering_documents", []),
        "engineering_monitor": engineering_monitor,
    }


def management_dashboard(filters: dict | None = None) -> dict:
    parsed_filters = _construction_filters(filters)
    cache_token = hashlib.sha1(
        json.dumps(
            {
                "filters": parsed_filters,
                "sqlite_only": settings.DASHFY_SQLITE_ONLY,
            },
            sort_keys=True,
            default=json_default,
            ensure_ascii=False,
        ).encode("utf-8")
    ).hexdigest()
    cache_key = f"management-dashboard:{cache_token}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    p6 = _safe_source("P6 base importada", lambda: _p6_dashboard(parsed_filters))
    if settings.DASHFY_SQLITE_ONLY:
        material = _safe_source(
            "SQLite DATAFY snapshot",
            lambda: _construction_datafy_snapshot(parsed_filters, allow_live=False),
        )
        construction = _safe_source("SQLite DED snapshot", lambda: _construction_sqlite_snapshot(parsed_filters))
        datafy = _sqlite_only_source("DATAFY")
        taskfy = _sqlite_only_source("Taskfy")
    else:
        datafy = datafy_dashboard()
        taskfy = taskfy_dashboard()
        construction = _safe_source("Taskfy obra", lambda: _construction_taskfy(parsed_filters))
        material = _safe_source("DATAFY suprimentos", lambda: _construction_datafy_snapshot(parsed_filters))

    k_task = construction.get("kpis", {})
    k_mat = material.get("kpis", {})
    k_p6 = p6.get("kpis", {})

    workstream = [
        {
            "area": "P6",
            "title": "Cronograma importado",
            "value": k_p6.get("physical_pct", 0),
            "unit": "% fisico",
            "detail": f"{int(k_p6.get('total_rows', 0)):,}".replace(",", ".") + " linhas PMS importadas",
            "url_name": "taskfy:task_list",
        },
        {
            "area": "Programacao",
            "title": "Pacotes programados",
            "value": k_task.get("programmed_packs", 0),
            "unit": "packs",
            "detail": f"{int(k_task.get('programmed_jobcards', 0)):,}".replace(",", ".") + " JobCards em programacao",
            "url_name": "taskfy:task_list",
        },
        {
            "area": "Campo",
            "title": "Executado via DFR",
            "value": k_task.get("actual_hh", 0),
            "unit": "HH",
            "detail": f"{int(k_task.get('dfrs', 0))} DFRs no periodo",
            "url_name": "taskfy:task_list",
        },
        {
            "area": "Workface",
            "title": "Pool liberado/AFC",
            "value": k_task.get("released_pool", 0),
            "unit": "JobCards",
            "detail": "Carteira disponivel para puxar para campo",
            "url_name": "taskfy:task_list",
        },
        {
            "area": "Materiais",
            "title": "Linhas acima de 50%",
            "value": k_mat.get("ready_50", 0),
            "unit": "linhas",
            "detail": f"{int(k_mat.get('ready_100', 0))} linhas 100% cobertas",
            "url_name": "datafy:home",
        },
        {
            "area": "Engenharia",
            "title": "Documentos DED",
            "value": k_task.get("engineering_docs", 0),
            "unit": "docs",
            "detail": "Distribuicao por disciplina e revisao abaixo",
            "url_name": "datafy:entries",
        },
    ]

    result = {
        "available": bool(datafy.get("available") or taskfy.get("available") or p6.get("available")),
        "sources_online": int(bool(datafy.get("available"))) + int(bool(taskfy.get("available"))) + int(bool(p6.get("available"))),
        "sources_total": 3,
        "sqlite_only": settings.DASHFY_SQLITE_ONLY,
        "filters": parsed_filters,
        "datafy": datafy,
        "taskfy": taskfy,
        "p6": p6,
        "construction": construction,
        "material": material,
        "kpis": {
            "planned_hh": k_task.get("p6_hh", 0),
            "programmed_hh": k_task.get("programmed_hh", 0),
            "actual_hh": k_task.get("actual_hh", 0),
            "program_execution_pct": k_task.get("program_execution_pct", 0),
            "hh_realized_pct": k_task.get("hh_realized_pct", 0),
            "p6_realized_pct": k_task.get("p6_realized_pct", 0),
            "piping_lines": k_mat.get("piping_lines", 0),
            "ready_100": k_mat.get("ready_100", 0),
            "ready_50": k_mat.get("ready_50", 0),
        },
        "workstream": workstream,
        "charts": {
            **construction.get("charts", {}),
            "p6_physical_curve": p6.get("charts", {}).get("physical_curve", _empty_chart("Sem curva fisica P6")),
            "p6_area_performance": p6.get("charts", {}).get("area_performance", _empty_chart("Sem areas executivas P6")),
            "p6_monthly_units": p6.get("charts", {}).get("monthly_units", _empty_chart("Sem distribuicao mensal P6")),
            "line_readiness": material.get("charts", {}).get("line_readiness", _empty_chart()),
        },
    }
    cache.set(cache_key, result, 45)
    return result


def workspace_snapshot() -> dict:
    return {
        "datafy": datafy_dashboard(),
        "taskfy": taskfy_dashboard(),
    }


def json_default(value: Any):
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)
