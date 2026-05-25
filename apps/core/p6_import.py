from __future__ import annotations

import hashlib
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from django.core.cache import cache
from django.db import transaction

from .models import P6CurveImport, P6CurvePoint, P6ManagementSnapshot, P6ProgressRow


PROGRESS_SHEET = "Progress PMS (%)"
CURVE_SHEET = "DB (%)"
MANAGEMENT_LOAD_SHEETS = ("DB BLPMSACT", "DB BLNW", "DB FORECAST NW PMS")


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").strip().split())


def _float(value: Any) -> float:
    text = str(value or "").strip().replace("%", "").replace("'", "").replace("\xa0", "")
    if not text:
        return 0.0
    text = text.replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _ratio(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    parsed = _float(value)
    if isinstance(value, str) and "%" in value:
        parsed = parsed / 100
    elif abs(parsed) > 1.5:
        parsed = parsed / 100
    return max(0.0, min(1.0, parsed))


def _workbook_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y", "%d-%b-%y", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _normalize(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _level(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return int(float(text.replace(",", ".")))
    except ValueError:
        return None


def _find_sheet_row(ws: Any, wanted: str) -> int | None:
    wanted_norm = _normalize(wanted)
    for row_idx in range(1, ws.max_row + 1):
        if _normalize(ws.cell(row_idx, 1).value) == wanted_norm:
            return row_idx
    return None


def _parse_progress_rows(ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_idx in range(1, ws.max_row + 1):
        name = _clean_text(ws.cell(row_idx, 4).value)
        if not name:
            continue
        raw_level = ws.cell(row_idx, 3).value
        level = _level(raw_level)
        if level is None:
            continue
        rows.append({
            "row_number": row_idx,
            "level": level,
            "name": name,
            "weight_pct": _ratio(ws.cell(row_idx, 5).value) * 100,
            "baseline_pct": _ratio(ws.cell(row_idx, 6).value) * 100,
            "planned_pct": _ratio(ws.cell(row_idx, 7).value) * 100,
            "actual_pct": _ratio(ws.cell(row_idx, 8).value) * 100,
        })
    return rows


def _parse_curve_points(ws: Any) -> list[dict[str, Any]]:
    dated_columns: list[tuple[int, date]] = []
    for col_idx in range(2, ws.max_column + 1):
        current = _workbook_date(ws.cell(4, col_idx).value)
        if current:
            dated_columns.append((col_idx, current))

    planned_row = (
        _find_sheet_row(ws, "PROJECT ACCO(%)")
        or _find_sheet_row(ws, "PROJECT ACCO BL (%)")
    )
    actual_row = _find_sheet_row(ws, "PROJECT ACCO Real (%)")
    if not dated_columns or not planned_row or not actual_row:
        return []

    points: list[dict[str, Any]] = []
    for sequence, (col_idx, period) in enumerate(dated_columns, start=1):
        points.append({
            "sequence": sequence,
            "period": period,
            "planned_pct": _ratio(ws.cell(planned_row, col_idx).value) * 100,
            "actual_pct": _ratio(ws.cell(actual_row, col_idx).value) * 100,
        })
    return points


def _month_floor(value: date) -> date:
    return date(value.year, value.month, 1)


def _add_month(value: date) -> date:
    if value.month == 12:
        return date(value.year + 1, 1, 1)
    return date(value.year, value.month + 1, 1)


def _month_label(value: date) -> str:
    labels = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
    return f"{labels[value.month - 1]} {value.year}"


def _number_label(value: Any) -> str:
    return f"{float(value or 0):,.0f}".replace(",", ".")


def _pct_label(value: Any) -> str:
    return f"{float(value or 0):.2f}%"


def _pct_css(value: Any) -> str:
    return f"{max(0, min(100, float(value or 0))):.2f}"


def _parse_management_weekly_load(workbook: Any) -> tuple[str, list[dict[str, Any]], float]:
    for sheet_name in MANAGEMENT_LOAD_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        weekly_rows: list[dict[str, Any]] = []
        total_budget = _float(ws.cell(2, 4).value)
        for col_idx in range(9, ws.max_column + 1):
            period = _workbook_date(ws.cell(1, col_idx).value)
            if not period:
                continue
            value = _float(ws.cell(2, col_idx).value)
            if value <= 0:
                continue
            weekly_rows.append({"period": period, "value": value})
        if weekly_rows:
            return sheet_name, weekly_rows, total_budget
    return "", [], 0.0


def _build_management_payload(
    workbook: Any,
    progress_rows: list[dict[str, Any]],
    curve_points: list[dict[str, Any]],
) -> tuple[dict[str, Any], str, int, int]:
    today = date.today()
    source_sheet, weekly_load, total_budget = _parse_management_weekly_load(workbook)
    if total_budget <= 0 and weekly_load:
        total_budget = sum(row["value"] for row in weekly_load)

    month_values: dict[str, float] = defaultdict(float)
    for row in weekly_load:
        month_key = _month_floor(row["period"]).isoformat()
        month_values[month_key] += float(row["value"] or 0)

    if month_values:
        first_month = _month_floor(min(row["period"] for row in weekly_load))
        last_month = _month_floor(max(row["period"] for row in weekly_load))
    elif curve_points:
        first_month = _month_floor(min(row["period"] for row in curve_points))
        last_month = _month_floor(max(row["period"] for row in curve_points))
    else:
        first_month = _month_floor(today)
        last_month = first_month

    months = []
    current = first_month
    while current <= last_month:
        months.append({
            "iso": current.isoformat(),
            "label": _month_label(current),
            "value": round(month_values.get(current.isoformat(), 0.0), 2),
        })
        current = _add_month(current)

    current_month = _month_floor(today).isoformat()
    next_90 = today + timedelta(days=90)
    current_month_units = sum(row["value"] for row in weekly_load if _month_floor(row["period"]).isoformat() == current_month)
    next_90_units = sum(row["value"] for row in weekly_load if today <= row["period"] <= next_90)

    total_row = next((row for row in progress_rows if int(row["level"]) == 0), {})
    physical_pct = float(total_row.get("actual_pct") or 0)
    planned_pct = float(total_row.get("planned_pct") or 0)
    actual_units = round(total_budget * physical_pct / 100, 2) if total_budget else 0.0
    remaining_units = max(total_budget - actual_units, 0.0)
    area_rows = []
    for row in progress_rows:
        if int(row["level"]) != 1:
            continue
        planned = float(row.get("planned_pct") or 0)
        actual = float(row.get("actual_pct") or 0)
        delta = actual - planned
        area_rows.append({
            "level": int(row["level"]),
            "label": row["name"],
            "weight_pct": float(row.get("weight_pct") or 0),
            "weight_label": _pct_label(float(row.get("weight_pct") or 0)),
            "planned_pct": planned,
            "planned_pct_css": _pct_css(planned),
            "planned_label": _pct_label(planned),
            "actual_pct": actual,
            "actual_pct_css": _pct_css(actual),
            "actual_label": _pct_label(actual),
            "delta_pct": delta,
            "delta_label": f"{delta:+.2f} p.p.",
            "delta_abs_css": f"{min(abs(delta) * 8, 100):.2f}",
            "status": "ahead" if delta >= 0 else "late",
            "finish_label": "-",
            "budget_label": "-",
            "remaining_label": "-",
        })

    late_area_rows = [row for row in area_rows if float(row.get("delta_pct") or 0) < 0]
    worst_area = min(area_rows, key=lambda row: float(row.get("delta_pct") or 0), default={})
    consult_tree = []
    stack: list[dict[str, Any]] = []
    for index, row in enumerate(progress_rows):
        level = int(row["level"])
        while stack and int(stack[-1]["level"]) >= level:
            stack.pop()
        node_id = f"xlsx-p6-{index}"
        parent = stack[-1]["id"] if stack else ""
        consult_tree.append({
            "id": node_id,
            "parent": parent,
            "level": level,
            "label": row["name"],
            "code": "-",
            "start": "-",
            "finish": "-",
            "pct": _pct_label(float(row.get("actual_pct") or 0)),
            "pct_css": _pct_css(float(row.get("actual_pct") or 0)),
            "status": "late" if float(row.get("actual_pct") or 0) < float(row.get("planned_pct") or 0) else "neutral",
            "status_label": "Atrasado" if float(row.get("actual_pct") or 0) < float(row.get("planned_pct") or 0) else "Planejado",
        })
        stack.append({"id": node_id, "level": level})

    monthly_planned = {row["iso"]: row["value"] for row in months}
    monthly_actual = {
        row["iso"]: row["value"] if row["iso"] <= current_month else 0.0
        for row in months
    }
    payload = {
        "source": "Annex III XLSX",
        "source_sheet": source_sheet,
        "area_rows": area_rows,
        "monthly_rows": months,
        "management_kpis": {
            "late_areas": len(late_area_rows),
            "total_budget_units": total_budget,
            "actual_units": actual_units,
            "physical_pct": physical_pct,
            "planned_today_pct": planned_pct,
            "remaining_units": remaining_units,
            "remaining_units_label": _number_label(remaining_units),
            "remaining_pct": round(max(100 - physical_pct, 0), 2),
            "current_month_units": current_month_units,
            "current_month_units_label": _number_label(current_month_units),
            "next_90_units": next_90_units,
            "next_90_units_label": _number_label(next_90_units),
            "active_packages": round(current_month_units),
            "next_90_packages": round(next_90_units),
            "worst_area_label": worst_area.get("label", "-"),
            "worst_area_delta": worst_area.get("delta_label", "-"),
        },
        "timeline": {
            "start": first_month.isoformat(),
            "finish": last_month.isoformat(),
            "start_label": first_month.strftime("%d/%m/%y"),
            "finish_label": last_month.strftime("%d/%m/%y"),
            "months": [{"iso": row["iso"], "label": row["label"]} for row in months],
            "month_count": len(months),
        },
        "monthly_planned": monthly_planned,
        "monthly_actual": monthly_actual,
        "consult_tree": consult_tree,
    }
    return payload, source_sheet, len(months), len(area_rows)


def import_p6_curves_workbook(uploaded_file: Any, *, imported_by: Any = None) -> P6CurveImport:
    filename = getattr(uploaded_file, "name", "") or "p6-curves.xlsx"
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("Envie a planilha P6 em formato .xlsx ou .xlsm.")

    content = uploaded_file.read()
    if not content:
        raise ValueError("O arquivo enviado esta vazio.")

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    if PROGRESS_SHEET not in workbook.sheetnames:
        raise ValueError(f"A aba obrigatoria '{PROGRESS_SHEET}' nao foi encontrada.")
    if CURVE_SHEET not in workbook.sheetnames:
        raise ValueError(f"A aba obrigatoria '{CURVE_SHEET}' nao foi encontrada.")

    progress_rows = _parse_progress_rows(workbook[PROGRESS_SHEET])
    curve_points = _parse_curve_points(workbook[CURVE_SHEET])
    management_payload, management_sheet, monthly_count, area_count = _build_management_payload(
        workbook,
        progress_rows,
        curve_points,
    )
    if not progress_rows:
        raise ValueError(f"Nenhuma linha PMS valida foi encontrada na aba '{PROGRESS_SHEET}'.")
    if not curve_points:
        raise ValueError(f"Nenhum ponto semanal valido foi encontrado na aba '{CURVE_SHEET}'.")

    file_hash = hashlib.sha256(content).hexdigest()
    user = getattr(imported_by, "_wrapped", imported_by)
    if not getattr(user, "is_authenticated", False):
        user = None
    with transaction.atomic():
        P6CurveImport.objects.filter(is_active=True).update(is_active=False)
        batch = P6CurveImport.objects.create(
            original_filename=filename,
            file_size=len(content),
            file_hash=file_hash,
            imported_by=user,
            progress_sheet=PROGRESS_SHEET,
            curve_sheet=CURVE_SHEET,
            progress_row_count=len(progress_rows),
            curve_point_count=len(curve_points),
            executive_row_count=sum(1 for row in progress_rows if row["level"] == 1),
            metadata={
                "sheet_names": workbook.sheetnames,
                "total_sheet_count": len(workbook.sheetnames),
            },
        )
        P6ProgressRow.objects.bulk_create([
            P6ProgressRow(import_batch=batch, **row)
            for row in progress_rows
        ])
        P6CurvePoint.objects.bulk_create([
            P6CurvePoint(import_batch=batch, **point)
            for point in curve_points
        ])
        P6ManagementSnapshot.objects.create(
            import_batch=batch,
            payload=management_payload,
            monthly_point_count=monthly_count,
            area_count=area_count,
            source_sheet=management_sheet or "Annex III",
        )

    cache.clear()
    return batch
